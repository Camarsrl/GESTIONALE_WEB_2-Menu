# -*- coding: utf-8 -*-
"""
Modulo Trasporti - versione corretta.
Fix:
- niente nullslast() per evitare errori SQL su alcuni database
- filtri data/costo più robusti
- export Excel compatibile sia con Date sia con Text
- registrazione tabella/colonne trasporti sicura
"""

def register_trasporti_routes(app_obj, deps):
    globals().update(deps)
    globals()["app"] = app_obj

    def _ensure_trasporti_schema_safe():
        """Crea/aggiorna la tabella trasporti se il DB esiste già."""
        try:
            Base.metadata.create_all(engine)
            insp = inspect(engine)
            tables = set(insp.get_table_names())
            if "trasporti" not in tables:
                return

            cols = {c.get("name") for c in insp.get_columns("trasporti")}
            extra_cols = {
                "data": "DATE",
                "tipo_mezzo": "TEXT",
                "cliente": "TEXT",
                "trasportatore": "TEXT",
                "ddt_uscita": "TEXT",
                "magazzino": "TEXT",
                "consolidato": "TEXT",
                "costo": "FLOAT",
            }
            for col, typ in extra_cols.items():
                if col not in cols:
                    try:
                        with engine.begin() as conn:
                            conn.execute(text(f"ALTER TABLE trasporti ADD COLUMN {col} {typ}"))
                        print(f"[OK] aggiunta colonna trasporti.{col}")
                    except Exception as e:
                        print(f"[WARN] colonna trasporti.{col}: {e}")
        except Exception as e:
            print(f"[WARN] ensure trasporti schema: {e}")

    _ensure_trasporti_schema_safe()

    def _as_date_safe(v):
        try:
            d = parse_any_date(v)
            if d:
                return d
        except Exception:
            pass
        try:
            d = to_date_db(v)
            if d:
                return d
        except Exception:
            pass
        return None

    def _fmt_date_safe(v):
        d = _as_date_safe(v)
        if d:
            try:
                return d.strftime("%Y-%m-%d")
            except Exception:
                return str(d)[:10]
        return ""

    def _match_txt(value, filtro):
        filtro = (filtro or '').strip()
        if not filtro:
            return True
        v = str(value or '')
        try:
            return filtro.lower() in v.lower() or normalize_text_key(filtro) in normalize_text_key(v)
        except Exception:
            return filtro.lower() in v.lower()

    def _cliente_safe_from_form(value):
        raw = (value or "").strip()
        if not raw:
            return None
        try:
            return validate_cliente_or_raise(raw)
        except Exception:
            # Non blocchiamo la pagina trasporti se il cliente non è nella lista utenti
            return raw.upper()

    @app.route('/trasporti', methods=['GET', 'POST'])
    @login_required
    def trasporti():
        db = SessionLocal()
        try:
            # --- MODIFICA TRASPORTO ---
            if request.method == 'POST' and request.form.get('edit_trasporto'):
                if session.get('role') != 'admin':
                    flash("ACCESSO NEGATO: Solo Admin.", "danger")
                    return redirect(url_for('trasporti'))

                try:
                    tid = int(request.form.get('id') or 0)
                    rec = db.query(Trasporto).filter(Trasporto.id == tid).first()
                    if not rec:
                        flash("Trasporto non trovato.", "danger")
                        return redirect(url_for('trasporti'))

                    data_str = (request.form.get('data') or '').strip()
                    rec.data = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None

                    costo_str = (request.form.get('costo') or '').strip()
                    rec.costo = float(costo_str.replace(',', '.')) if costo_str != '' else None

                    rec.tipo_mezzo = (request.form.get('tipo_mezzo') or '').strip() or None
                    rec.cliente = _cliente_safe_from_form(request.form.get('cliente'))
                    rec.trasportatore = (request.form.get('trasportatore') or '').strip() or None
                    rec.ddt_uscita = (request.form.get('ddt_uscita') or '').strip() or None
                    rec.magazzino = (request.form.get('magazzino') or '').strip() or None
                    rec.consolidato = (request.form.get('consolidato') or '').strip() or None

                    db.commit()
                    flash("Trasporto modificato!", "success")
                except Exception as e:
                    db.rollback()
                    flash(f"Errore modifica trasporto: {e}", "danger")

                return redirect(url_for('trasporti'))

            # --- AGGIUNGI NUOVO TRASPORTO ---
            if request.method == 'POST' and request.form.get('add_trasporto'):
                if session.get('role') != 'admin':
                    flash("ACCESSO NEGATO: Solo Admin.", "danger")
                    return redirect(url_for('trasporti'))

                try:
                    data_str = (request.form.get('data') or '').strip()
                    data_val = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None

                    costo_str = (request.form.get('costo') or '').strip()
                    costo_val = float(costo_str.replace(',', '.')) if costo_str != '' else None

                    nuovo = Trasporto(
                        data=data_val,
                        tipo_mezzo=(request.form.get('tipo_mezzo') or '').strip() or None,
                        cliente=_cliente_safe_from_form(request.form.get('cliente')),
                        trasportatore=(request.form.get('trasportatore') or '').strip() or None,
                        ddt_uscita=(request.form.get('ddt_uscita') or '').strip() or None,
                        magazzino=(request.form.get('magazzino') or '').strip() or None,
                        consolidato=(request.form.get('consolidato') or '').strip() or None,
                        costo=costo_val
                    )

                    db.add(nuovo)
                    db.commit()
                    flash("Trasporto salvato!", "success")
                except Exception as e:
                    db.rollback()
                    flash(f"Errore salvataggio trasporto: {e}", "danger")

                return redirect(url_for('trasporti'))

            # --- EDIT MODE ---
            edit_id = request.args.get('edit_id')
            edit_row = None
            if edit_id and session.get('role') == 'admin':
                try:
                    edit_row = db.query(Trasporto).filter(Trasporto.id == int(edit_id)).first()
                except Exception:
                    edit_row = None

            filtri = {
                'data_da': (request.args.get('data_da') or '').strip(),
                'data_a': (request.args.get('data_a') or '').strip(),
                'cliente': (request.args.get('cliente') or '').strip(),
                'tipo_mezzo': (request.args.get('tipo_mezzo') or '').strip(),
                'trasportatore': (request.args.get('trasportatore') or '').strip(),
                'ddt_uscita': (request.args.get('ddt_uscita') or '').strip(),
                'magazzino': (request.args.get('magazzino') or '').strip(),
                'consolidato': (request.args.get('consolidato') or '').strip(),
                'costo_da': (request.args.get('costo_da') or '').strip(),
                'costo_a': (request.args.get('costo_a') or '').strip(),
            }

            data_da = _safe_date_ymd(filtri['data_da'])
            data_a = _safe_date_ymd(filtri['data_a'])
            costo_da = _safe_float_it(filtri['costo_da'])
            costo_a = _safe_float_it(filtri['costo_a'])

            # Evita nullslast(): in alcuni ambienti può generare errore.
            dati = db.query(Trasporto).order_by(Trasporto.id.desc()).all()

            filtered = []
            for rec in dati:
                rec_date = _as_date_safe(getattr(rec, 'data', None))
                if data_da and (not rec_date or rec_date < data_da):
                    continue
                if data_a and (not rec_date or rec_date > data_a):
                    continue
                if not _match_txt(rec.cliente, filtri['cliente']):
                    continue
                if not _match_txt(rec.tipo_mezzo, filtri['tipo_mezzo']):
                    continue
                if not _match_txt(rec.trasportatore, filtri['trasportatore']):
                    continue
                if not _match_txt(rec.ddt_uscita, filtri['ddt_uscita']):
                    continue
                if not _match_txt(rec.magazzino, filtri['magazzino']):
                    continue
                if not _match_txt(rec.consolidato, filtri['consolidato']):
                    continue
                costo_val = _safe_float_it(rec.costo)
                if costo_da is not None and (costo_val is None or costo_val < costo_da):
                    continue
                if costo_a is not None and (costo_val is None or costo_val > costo_a):
                    continue
                filtered.append(rec)

            # ordinamento Python robusto: data desc, id desc
            filtered.sort(key=lambda r: (_as_date_safe(getattr(r, 'data', None)) or date.min, getattr(r, 'id', 0) or 0), reverse=True)

            return render_template(
                'trasporti.html',
                trasporti=filtered,
                today=date.today(),
                edit_row=edit_row,
                filtri=filtri,
                clienti_validi=get_clienti_utenti()
            )

        except Exception as e:
            db.rollback()
            scrivi_log_errore("Errore pagina trasporti", e)
            flash(f"Errore apertura Trasporti: {e}", "danger")
            return redirect(url_for('home'))
        finally:
            db.close()


    @app.route('/report_trasporti', methods=['POST'])
    @login_required
    def report_trasporti():
        if session.get('role') != 'admin':
            return "No Access", 403

        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        mese = (request.form.get('mese') or '').strip()
        mezzo = (request.form.get('tipo_mezzo') or '').strip()
        cliente = (request.form.get('cliente') or '').strip()
        ddt_uscita = (request.form.get('ddt_uscita') or '').strip()
        consolidato = (request.form.get('consolidato') or '').strip()

        db = SessionLocal()
        try:
            rows = db.query(Trasporto).all()
            dati = []

            mese_start = mese_end = None
            if mese:
                try:
                    y, m = [int(x) for x in mese.split("-")]
                    mese_start = date(y, m, 1)
                    mese_end = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
                except Exception:
                    mese_start = mese_end = None

            for t in rows:
                d_val = _as_date_safe(getattr(t, "data", None))
                if mese_start and (not d_val or d_val < mese_start or d_val >= mese_end):
                    continue
                if mezzo and not _match_txt(t.tipo_mezzo, mezzo):
                    continue
                if cliente and not _match_txt(t.cliente, cliente):
                    continue
                if ddt_uscita and not _match_txt(t.ddt_uscita, ddt_uscita):
                    continue
                if consolidato and not _match_txt(t.consolidato, consolidato):
                    continue
                dati.append(t)

            dati.sort(key=lambda r: (_as_date_safe(getattr(r, 'data', None)) or date.min, getattr(r, 'id', 0) or 0))

            wb = Workbook()
            ws = wb.active
            ws.title = "Trasporti"

            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            header_fill = PatternFill("solid", fgColor="D9E1F2")

            ws["A1"] = "REPORT TRASPORTI"
            ws["A1"].font = Font(bold=True, size=16)
            ws.merge_cells("A1:H1")
            ws["A1"].alignment = center

            ws["A3"] = "Filtri:"
            ws["A3"].font = bold
            ws["B3"] = f"Mese={mese or 'Tutti'} | Cliente={cliente or 'Tutti'} | Mezzo={mezzo or 'Tutti'} | DDT={ddt_uscita or 'Tutti'} | Consolidato={consolidato or 'Tutti'}"
            ws.merge_cells("B3:H3")

            headers = ["Data", "Mezzo", "Cliente", "Trasportatore", "DDT", "Magazzino", "Consolidato", "Costo (€)"]
            start_row = 5
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col, value=h)
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = center

            riga = start_row + 1
            totale = 0.0

            for t in dati:
                costo_val = float(t.costo or 0.0)
                totale += costo_val

                ws.cell(riga, 1, _fmt_date_safe(t.data)).alignment = center
                ws.cell(riga, 2, (t.tipo_mezzo or "")).alignment = left
                ws.cell(riga, 3, (t.cliente or "")).alignment = left
                ws.cell(riga, 4, (t.trasportatore or "")).alignment = left
                ws.cell(riga, 5, (t.ddt_uscita or "")).alignment = center
                ws.cell(riga, 6, (t.magazzino or "")).alignment = center
                ws.cell(riga, 7, (t.consolidato or "")).alignment = center

                c = ws.cell(riga, 8, costo_val)
                c.number_format = '#,##0.00'
                c.alignment = center

                riga += 1

            ws.cell(riga, 1, "TOTALE").font = bold
            ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=7)
            ws.cell(riga, 1).alignment = Alignment(horizontal="right", vertical="center")
            tot_cell = ws.cell(riga, 8, totale)
            tot_cell.font = bold
            tot_cell.number_format = '#,##0.00'
            tot_cell.alignment = center

            for i, w in enumerate([12, 16, 22, 22, 14, 14, 16, 12], start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

            safe_mese = mese.replace("-", "_") if mese else "TUTTO"
            filename = f"Report_Trasporti_{safe_mese}.xlsx"

            return send_file(
                bio,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            scrivi_log_errore("Errore export Trasporti Excel", e)
            return f"Errore export Trasporti Excel: {e}", 500
        finally:
            db.close()
