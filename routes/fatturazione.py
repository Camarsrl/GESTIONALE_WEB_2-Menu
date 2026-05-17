# -*- coding: utf-8 -*-
"""
Modulo Fatturazione / Report - Step completo.
Contiene report fatturazione e calcolo costi/giacenze mensili.
"""


def register_fatturazione_routes(app_obj, deps):
    globals().update(deps)
    globals()["app"] = app_obj

    # ========================================================
    # REPORT FATTURAZIONE (SOLO ADMIN)
    # ========================================================
    FATTURAZIONE_SPECIAL_RULES = {
        normalize_text_key('GALVANO TECNICA'): {'mode': 'pallet', 'label': 'GALVANO TECNICA'},
    }


    def parse_any_date(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        s = str(value).strip().split(' ')[0]
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except Exception:
                pass
        return None


    def _cliente_report_config():
        configs = []
        for cli in get_clienti_utenti():
            norm = normalize_text_key(cli)
            special = FATTURAZIONE_SPECIAL_RULES.get(norm, {})
            mode = special.get('mode', 'm2')
            label = special.get('label', cli)
            configs.append({'cliente': cli, 'label': label, 'mode': mode, 'norm': norm})
        configs.sort(key=lambda x: (0 if x['mode'] == 'm2' else 1, x['label']))
        return configs


    def _safe_float(value):
        try:
            return float(str(value).replace(',', '.')) if value not in (None, '') else 0.0
        except Exception:
            return 0.0


    def _safe_int(value, default_if_blank=0):
        if value in (None, ''):
            return default_if_blank
        try:
            return int(float(str(value).replace(',', '.')))
        except Exception:
            return default_if_blank


    def _compute_report_fatturazione_data(mese: int, anno: int):
        mese = max(1, min(12, int(mese)))
        anno = int(anno)
        first_day = date(anno, mese, 1)
        last_day = date(anno, mese, calendar.monthrange(anno, mese)[1])

        db = SessionLocal()
        try:
            articoli = db.query(Articolo).all()
        finally:
            db.close()

        configs = _cliente_report_config()
        rows = []
        totals = {
            'm2_presenti': 0.0,
            'm2_fine_mese': 0.0,
            'm2_usciti': 0.0,
            'entrate_doganali_m2': 0.0,
            'picco_m2_occupati': 0.0,
            'pallet_giacenza': 0.0,
        }

        for conf in configs:
            row = {
                'cliente': conf['label'],
                'm2_presenti': 0.0,
                'm2_fine_mese': 0.0,
                'm2_usciti': 0.0,
                'entrate_doganali_m2': 0.0,
                'picco_m2_occupati': 0.0,
                'pallet_giacenza': 0.0,
            }
            norm_cli = conf['norm']

            cliente_articoli = [art for art in articoli if normalize_text_key(getattr(art, 'cliente', '')) == norm_cli]

            for art in cliente_articoli:
                d_ing = parse_any_date(getattr(art, 'data_ingresso', None))
                d_usc = parse_any_date(getattr(art, 'data_uscita', None))
                stato_norm = normalize_text_key(getattr(art, 'stato', ''))
                m2 = _safe_float(getattr(art, 'm2', 0))

                presente_nel_mese = d_ing is not None and d_ing <= last_day and (d_usc is None or d_usc >= first_day)
                presente_fine_mese = d_ing is not None and d_ing <= last_day and (d_usc is None or d_usc >= last_day)
                uscito_nel_mese = d_usc is not None and first_day <= d_usc <= last_day
                ingresso_doganale = d_ing is not None and first_day <= d_ing <= last_day and 'DOGAN' in stato_norm

                if conf['mode'] == 'pallet':
                    # GALVANO TECNICA:
                    # deve contare SOLO i pallet ancora in giacenza nel mese selezionato,
                    # cioè entrati entro la fine del mese e NON usciti entro la fine del mese.
                    # Se una riga è uscita durante il mese selezionato, non va conteggiata.
                    pallet_ancora_in_giacenza_fine_mese = (
                        d_ing is not None
                        and d_ing <= last_day
                        and (d_usc is None or d_usc > last_day)
                    )

                    # Per GALVANO TECNICA il conteggio deve usare la colonna N° Colli
                    # come nella schermata Giacenze. I campi vuoti NON devono valere 1,
                    # altrimenti il report conta una riga vuota come un bancale e gonfia il totale.
                    # Inoltre vengono esclusi tutti i record già usciti entro la fine del mese.
                    pallet_qty = _safe_int(getattr(art, 'n_colli', None), default_if_blank=0)
                    if pallet_qty < 0:
                        pallet_qty = 0

                    if pallet_ancora_in_giacenza_fine_mese:
                        row['pallet_giacenza'] += pallet_qty
                else:
                    if presente_nel_mese:
                        row['m2_presenti'] += m2
                    if presente_fine_mese:
                        row['m2_fine_mese'] += m2
                    if uscito_nel_mese:
                        row['m2_usciti'] += m2
                    if ingresso_doganale:
                        row['entrate_doganali_m2'] += m2

            if conf['mode'] != 'pallet':
                peak = 0.0
                for day_num in range(1, last_day.day + 1):
                    current_day = date(anno, mese, day_num)
                    occupied_m2 = 0.0
                    for art in cliente_articoli:
                        d_ing = parse_any_date(getattr(art, 'data_ingresso', None))
                        d_usc = parse_any_date(getattr(art, 'data_uscita', None))
                        if d_ing is not None and d_ing <= current_day and (d_usc is None or d_usc >= current_day):
                            occupied_m2 += _safe_float(getattr(art, 'm2', 0))
                    if occupied_m2 > peak:
                        peak = occupied_m2
                row['picco_m2_occupati'] = peak

            if any(abs(v) > 1e-9 for k, v in row.items() if k != 'cliente'):
                rows.append(row)
                for key in totals:
                    totals[key] += row[key]

        return rows, totals, first_day, last_day


    @app.route('/report_fatturazione')
    @login_required
    @require_admin
    def report_fatturazione():
        today = date.today()
        mese = request.args.get('mese', today.month, type=int)
        anno = request.args.get('anno', today.year, type=int)
        rows, totals, first_day, last_day = _compute_report_fatturazione_data(mese, anno)
        return render_template(
            'report_fatturazione.html',
            title='Report Fatturazione',
            mese=mese,
            anno=anno,
            rows=rows,
            totals=totals,
            periodo_da=first_day,
            periodo_a=last_day,
        )


    @app.route('/report_fatturazione/export_excel')
    @login_required
    @require_admin
    def export_report_fatturazione_excel():
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage

        today = date.today()
        mese = request.args.get('mese', today.month, type=int)
        anno = request.args.get('anno', today.year, type=int)
        rows, totals, first_day, last_day = _compute_report_fatturazione_data(mese, anno)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Fatturazione'

        current_row = 1
        if LOGO_PATH and Path(LOGO_PATH).exists():
            try:
                img = XLImage(str(LOGO_PATH))
                img.width = 180
                img.height = 55
                ws.add_image(img, 'A1')
                current_row = 5
            except Exception:
                current_row = 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.cell(current_row, 1).value = f'Report Fatturazione Camar - {mese:02d}/{anno}'
        ws.cell(current_row, 1).font = Font(bold=True, size=15)
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center')
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.cell(current_row, 1).value = f'Periodo: {first_day.strftime("%d/%m/%Y")} - {last_day.strftime("%d/%m/%Y")}'
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center')
        current_row += 2

        headers = ['Cliente', 'M2 presenti nel mese', 'M2 giacenza fine mese', 'M2 usciti nel mese', 'Entrate doganali M2', 'Picco M2 occupati', 'Pallet giacenza mese']
        header_fill = PatternFill('solid', fgColor='1F6FB2')
        thin = Side(style='thin', color='CCCCCC')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in rows:
            current_row += 1
            values = [
                row['cliente'], row['m2_presenti'], row['m2_fine_mese'], row['m2_usciti'], row['entrate_doganali_m2'], row['picco_m2_occupati'], row['pallet_giacenza']
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(current_row, col, value)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if col > 1:
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='right')

        if rows:
            current_row += 1
            total_values = ['TOTALE', totals['m2_presenti'], totals['m2_fine_mese'], totals['m2_usciti'], totals['entrate_doganali_m2'], totals['picco_m2_occupati'], totals['pallet_giacenza']]
            for col, value in enumerate(total_values, 1):
                cell = ws.cell(current_row, col, value)
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='D9EAF7')
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if col > 1:
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='right')

        widths = {1: 28, 2: 18, 3: 22, 4: 18, 5: 20, 6: 18, 7: 18}
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(
            bio,
            as_attachment=True,
            download_name=f'report_fatturazione_{anno}_{mese:02d}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def _parse_data_db_helper(val):
        """
        Accetta:
        - date / datetime
        - stringa 'YYYY-MM-DD'
        - stringa 'DD/MM/YYYY'
        - stringa con orario 'YYYY-MM-DD HH:MM:SS'
        Ritorna date oppure None.
        """
        if val is None:
            return None

        if isinstance(val, datetime):
            return val.date()

        if isinstance(val, date):
            return val

        s = str(val).strip()
        if not s:
            return None

        # prova YYYY-MM-DD
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
        except Exception:
            pass

        # prova DD/MM/YYYY
        try:
            return datetime.strptime(s[:10], "%d/%m/%Y").date()
        except Exception:
            pass

        return None

    # --- LOGICA CALCOLO COSTI (ROBUSTA) ---
    def _calcola_logica_costi(articoli, data_da, data_a, raggruppamento, m2_multiplier: float = 1.0, metric: str = "m2"):
        """
        metric:
          - "m2"    => usa art.m2
          - "colli" => usa art.n_colli
          - "pezzi" => usa art.pezzi / art.pezzo
        Ritorna SEMPRE anche m2_tot/m2_medio per compatibilità template.
        """
        from collections import defaultdict
        from datetime import timedelta, date, datetime

        val_per_giorno = defaultdict(float)

        def to_date_obj(d):
            if not d:
                return None
            if isinstance(d, datetime):
                return d.date()
            if isinstance(d, date):
                return d
            s = str(d).strip().split(" ")[0]
            if len(s) < 8 or not s[0].isdigit():
                return None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(s, fmt).date()
                except:
                    pass
            return None

        d_start = to_date_obj(data_da)
        d_end = to_date_obj(data_a)
        if not d_start or not d_end:
            return []

        metric = (metric or "m2").strip().lower()

        def get_qty(art):
            # ✅ M2
            if metric == "m2":
                try:
                    val_m2 = str(getattr(art, "m2", "") or "").replace(",", ".")
                    m2 = float(val_m2) if val_m2 else 0.0
                except:
                    m2 = 0.0
                if m2 <= 0:
                    return 0.0

                # Area manovra (solo se metric == m2)
                try:
                    m2 = m2 * float(m2_multiplier or 1.0)
                except:
                    pass

                return float(m2)

            # ✅ COLLI
            if metric == "colli":
                try:
                    return float(int(getattr(art, "n_colli", 0) or 0))
                except:
                    return 0.0

            # ✅ PEZZI
            if metric == "pezzi":
                raw = getattr(art, "pezzi", None)
                if raw is None:
                    raw = getattr(art, "pezzo", None)
                try:
                    return float(int(raw or 0))
                except:
                    return 0.0

            # fallback
            return 0.0

        for art in articoli:
            qty = get_qty(art)
            if qty <= 0:
                continue

            d_ingr = to_date_obj(getattr(art, "data_ingresso", None))
            if not d_ingr:
                continue

            d_usc = to_date_obj(getattr(art, "data_uscita", None))

            inizio = max(d_ingr, d_start)
            if d_usc:
                fine = min(d_usc - timedelta(days=1), d_end)
            else:
                fine = d_end

            if fine < inizio:
                continue

            cliente_key = (getattr(art, "cliente", None) or "SCONOSCIUTO").strip().upper()

            curr = inizio
            while curr <= fine:
                val_per_giorno[(cliente_key, curr)] += qty
                curr += timedelta(days=1)

        risultati_finali = []

        def pack_row(periodo, cliente, tot, medio, giorni):
            # ✅ compatibilità: restituisco SEMPRE anche m2_tot/m2_medio
            # così il template admin che stampa r.m2_tot / r.m2_medio funziona sempre.
            tot_s = f"{tot:.3f}" if isinstance(tot, (int, float)) else str(tot)
            med_s = f"{medio:.3f}" if isinstance(medio, (int, float)) else str(medio)

            return {
                "periodo": periodo,
                "cliente": cliente,
                # chiavi nuove "neutre"
                "tot": tot_s,
                "medio": med_s,
                "giorni": giorni,
                # chiavi legacy del template
                "m2_tot": tot_s,
                "m2_medio": med_s,
            }

        if raggruppamento == "giorno":
            sorted_keys = sorted(val_per_giorno.keys(), key=lambda k: (k[0], k[1]))
            for cliente, giorno in sorted_keys:
                val = val_per_giorno[(cliente, giorno)]
                risultati_finali.append(
                    pack_row(giorno.strftime("%d/%m/%Y"), cliente, val, val, 1)
                )
        else:
            agg = defaultdict(lambda: {"sum": 0.0, "days": set()})
            for (cli, day), val in val_per_giorno.items():
                k = (cli, day.year, day.month)
                agg[k]["sum"] += val
                agg[k]["days"].add(day)

            sorted_keys = sorted(agg.keys(), key=lambda k: (k[1], k[2], k[0]))
            for (cli, y, m) in sorted_keys:
                dati = agg[(cli, y, m)]
                n_days = len(dati["days"])
                tot = dati["sum"]

                # ✅ M² EFFETTIVI (non medi): valore reale sull'ULTIMO giorno del periodo considerato per quel mese
                if n_days > 0:
                    last_day = max(dati["days"])
                    eff = float(val_per_giorno.get((cli, last_day), 0.0))
                else:
                    eff = 0.0

                risultati_finali.append(
                    pack_row(f"{m:02d}/{y}", cli, tot, eff, n_days)
                )

        return risultati_finali


    def _calcola_logica_colli_giacenza(articoli, data_da, data_a, raggruppamento):
        """
        Calcola i COLLI in GIACENZA nel periodo (fotografia giornaliera o mensile),
        togliendo quelli già usciti.

        - Per ogni giorno: somma n_colli degli articoli che risultano "presenti" quel giorno.
        - Presente = data_ingresso <= giorno AND (data_uscita è vuota oppure data_uscita > giorno)
        """
        from collections import defaultdict
        from datetime import timedelta, date, datetime

        colli_per_giorno = defaultdict(float)

        def to_date_obj(d):
            if not d:
                return None
            if isinstance(d, datetime):
                return d.date()
            if isinstance(d, date):
                return d
            s = str(d).strip().split(' ')[0]
            if len(s) < 8 or not s[0].isdigit():
                return None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    pass
            return None

        d_start = to_date_obj(data_da)
        d_end = to_date_obj(data_a)
        if not d_start or not d_end:
            return []

        for art in articoli:
            try:
                colli = float(int(art.n_colli or 0))
            except Exception:
                colli = 0.0

            if colli <= 0:
                continue

            d_ingr = to_date_obj(art.data_ingresso)
            if not d_ingr:
                continue

            d_usc = to_date_obj(art.data_uscita)

            # Range di verifica nel periodo
            start = max(d_ingr, d_start)
            end = d_end

            if end < start:
                continue

            cliente_key = (art.cliente or "SCONOSCIUTO").strip().upper()

            curr = start
            while curr <= end:
                presente = (d_ingr <= curr) and ((d_usc is None) or (d_usc > curr))
                if presente:
                    colli_per_giorno[(cliente_key, curr)] += colli
                curr += timedelta(days=1)

        risultati = []

        if raggruppamento == "giorno":
            keys = sorted(colli_per_giorno.keys(), key=lambda k: (k[0], k[1]))
            for cli, day in keys:
                v = colli_per_giorno[(cli, day)]
                risultati.append({
                    "periodo": day.strftime("%d/%m/%Y"),
                    "cliente": cli,
                    "tot": f"{v:.0f}",
                    "medio": f"{v:.0f}",
                    "giorni": 1
                })
        else:
            agg = defaultdict(lambda: {"sum": 0.0, "days": set()})
            for (cli, day), v in colli_per_giorno.items():
                k = (cli, day.year, day.month)
                agg[k]["sum"] += v
                agg[k]["days"].add(day)

            keys = sorted(agg.keys(), key=lambda k: (k[1], k[2], k[0]))
            for cli, y, m in keys:
                dati = agg[(cli, y, m)]
                n_days = len(dati["days"])
                tot = dati["sum"]
                avg = tot / n_days if n_days else 0.0
                risultati.append({
                    "periodo": f"{m:02d}/{y}",
                    "cliente": cli,
                    "tot": f"{tot:.0f}",
                    "medio": f"{avg:.0f}",
                    "giorni": n_days
                })

        return risultati


    @app.route('/calcola_costi', methods=['GET', 'POST'])
    @login_required
    def calcola_costi():
        oggi = date.today()
        data_da_val = (oggi.replace(day=1)).strftime("%Y-%m-%d")
        data_a_val = oggi.strftime("%Y-%m-%d")

        # Admin: può filtrare + area manovra
        # Client: solo il proprio cliente e niente area manovra
        is_admin = (session.get('role') == 'admin')
        cliente_lock = current_cliente()  # stringa se role=client, altrimenti None

        cliente_val = (cliente_lock or "")
        raggruppamento = "mese"
        area_manovra_val = False
        risultati = []
        metric = "m2"

        def _metric_for_cliente(nome_cliente: str) -> str:
            s = (nome_cliente or "").strip().upper()
            # Per Galvano Tecnica calcoliamo COLLI in giacenza
            if "GALVANO" in s:
                return "colli"
            return "m2"

        if request.method == 'POST':
            data_da_str = request.form.get('data_da')
            data_a_str = request.form.get('data_a')
            raggruppamento = request.form.get('raggruppamento', 'mese')

            # Cliente + area manovra
            if is_admin:
                cliente_val = (request.form.get('cliente') or '').strip()
                area_manovra = (request.form.get('area_manovra') == '1')
            else:
                cliente_val = (cliente_lock or '').strip()
                area_manovra = False

            export_excel = ('export_excel' in request.form)

            # metrica (colli o m2)
            metric = _metric_for_cliente(cliente_val)

            try:
                db = SessionLocal()
                query = db.query(Articolo)

                # ✅ filtro sicuro
                if cliente_val:
                    cliente_norm = normalize_text_key(cliente_val)
                    if cliente_norm:
                        query = query.filter(normalized_sql_text(Articolo.cliente) == cliente_norm)

                articoli = query.all()
                db.close()

                # ✅ calcolo in base a metrica
                if metric == "colli":
                    risultati = _calcola_logica_colli_giacenza(
                        articoli,
                        data_da_str,
                        data_a_str,
                        raggruppamento
                    )
                else:
                    risultati = _calcola_logica_costi(
                        articoli,
                        data_da_str,
                        data_a_str,
                        raggruppamento,
                        m2_multiplier=(1.25 if (is_admin and area_manovra) else 1.0)
                    )

                data_da_val = data_da_str
                data_a_val = data_a_str
                area_manovra_val = bool(is_admin and area_manovra and metric == "m2")

                # ✅ Export Excel
                if export_excel:
                    try:
                        df = pd.DataFrame(risultati)

                        if metric == "colli":
                            # qui i risultati hanno chiavi: periodo, cliente, tot, medio, giorni
                            df = df.rename(columns={
                                'periodo': 'Periodo',
                                'cliente': 'Cliente',
                                'tot': 'Colli Giacenza (somma)',
                                'medio': 'Colli Medi',
                                'giorni': 'Giorni'
                            })
                            filename = f"Report_Colli_Giacenza_{data_da_val}_to_{data_a_val}.xlsx"
                        else:
                            # qui i risultati hanno chiavi: periodo, cliente, m2_tot, m2_medio, giorni
                            df = df.rename(columns={
                                'periodo': 'Periodo',
                                'cliente': 'Cliente',
                                'm2_tot': 'M2 Tot',
                                'm2_medio': 'M2 Medio',
                                'giorni': 'Giorni'
                            })
                            extra = '_AREA_MANOVRA' if area_manovra_val else ''
                            filename = f"Report_Costi{extra}_{data_da_val}_to_{data_a_val}.xlsx"

                        bio = io.BytesIO()
                        df.to_excel(bio, index=False, engine='openpyxl')
                        bio.seek(0)

                        return send_file(
                            bio,
                            as_attachment=True,
                            download_name=filename,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                    except Exception as e:
                        flash(f"Errore export Excel: {e}", "danger")

                if not risultati:
                    flash("Nessun dato valido trovato per i criteri selezionati.", "warning")

            except Exception as e:
                flash(f"Errore: {e}", "danger")

        # ✅ anche su GET metrica in base al cliente lock
        metric = _metric_for_cliente(cliente_val)

        return render_template(
            'calcoli.html',
            risultati=risultati,
            data_da=data_da_val,
            data_a=data_a_val,
            cliente_filtro=cliente_val,
            raggruppamento=raggruppamento,
            area_manovra=area_manovra_val,
            is_admin=is_admin,
            metric=metric
        )
