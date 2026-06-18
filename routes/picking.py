# -*- coding: utf-8 -*-
"""
Modulo Picking / Lavorazioni.

Contiene le route:
- /lavorazioni
- /stampa_picking_pdf

Spostato dal file principale per alleggerire gestionale_web_full.py.
"""

def register_picking_routes(app_obj, deps):
    globals().update(deps)
    globals()["app"] = app_obj

    import io
    import re
    from datetime import date, datetime
    from flask import request, redirect, url_for, flash, render_template, send_file, session
    from flask_login import login_required


    def _parse_picking_date_safe(value):
        """Converte la data picking in date senza dipendere da helper esterni mancanti."""
        if not value:
            return None
        try:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
        except Exception:
            pass
        try:
            if 'parse_any_date' in globals():
                d = parse_any_date(value)
                if d:
                    return d.date() if isinstance(d, datetime) else d
        except Exception:
            pass
        try:
            if 'to_date_db' in globals():
                d = to_date_db(value)
                if d:
                    return d.date() if isinstance(d, datetime) else d
        except Exception:
            pass
        s = str(value or '').strip().split(' ')[0][:10]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        return None

    def _month_bounds_from_request_lavorazioni():
        """Mese iniziale: se non ci sono filtri mostra il mese corrente.
        Con ?mese=YYYY-MM mostra quel mese; con ?view=tutti mostra tutto.
        """
        mese = (request.args.get('mese') or '').strip()
        view = (request.args.get('view') or '').strip().lower()
        if view == 'tutti':
            return '', None, None, 'tutti'

        if not mese:
            has_filters = any((request.args.get(k) or '').strip() for k in [
                'data_da','data_a','cliente','descrizione','richiesta_di','seriali','n_arrivo',
                'colli_da','colli_a','pallet_forniti_da','pallet_forniti_a',
                'pallet_uscita_da','pallet_uscita_a','ore_blue_da','ore_blue_a',
                'ore_white_da','ore_white_a'
            ])
            if has_filters:
                return '', None, None, ''
            mese = date.today().strftime('%Y-%m')

        try:
            y, m = [int(x) for x in mese.split('-', 1)]
            start = date(y, m, 1)
            end = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
            return mese, start, end, 'mese'
        except Exception:
            return '', None, None, ''

    def _prev_next_month(mese):
        try:
            y, m = [int(x) for x in (mese or '').split('-', 1)]
            prev_y, prev_m = (y - 1, 12) if m == 1 else (y, m - 1)
            next_y, next_m = (y + 1, 1) if m == 12 else (y, m + 1)
            return f"{prev_y:04d}-{prev_m:02d}", f"{next_y:04d}-{next_m:02d}"
        except Exception:
            cur = date.today().strftime('%Y-%m')
            return cur, cur


    @app.route('/lavorazioni', methods=['GET', 'POST'])
    @login_required
    def lavorazioni():
        db = SessionLocal()

        # --- MODIFICA LAVORAZIONE ---
        if request.method == 'POST' and request.form.get('edit_lavorazione'):
            if session.get('role') != 'admin':
                flash("ACCESSO NEGATO: Solo Admin.", "danger")
                return redirect(url_for('lavorazioni', mese=date.today().strftime('%Y-%m')))

            try:
                lid = int(request.form.get('id') or 0)
                rec = db.query(Lavorazione).filter(Lavorazione.id == lid).first()
                if not rec:
                    flash("Record non trovato.", "danger")
                    return redirect(url_for('lavorazioni', mese=date.today().strftime('%Y-%m')))

                d_val = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
                rec.data = d_val
                rec.cliente = canonical_cliente_picking(request.form.get('cliente'))
                rec.descrizione = request.form.get('descrizione')
                rec.richiesta_di = request.form.get('richiesta_di')
                rec.seriali = request.form.get('seriali')
                rec.colli = int(request.form.get('colli') or 0)
                rec.pallet_forniti = int(request.form.get('pallet_forniti') or 0)
                rec.pallet_uscita = int(request.form.get('pallet_uscita') or 0)
                rec.ore_blue_collar = float(request.form.get('ore_blue_collar') or 0)
                rec.ore_white_collar = float(request.form.get('ore_white_collar') or 0)

                db.commit()
                flash("Picking modificato!", "success")
            except Exception as e:
                db.rollback()
                flash(f"Errore modifica: {e}", "danger")

            return redirect(url_for('lavorazioni', mese=date.today().strftime('%Y-%m')))

        # --- INSERIMENTO ---
        if request.method == 'POST' and request.form.get('add_lavorazione'):
            if session.get('role') != 'admin':
                flash("ACCESSO NEGATO: Solo Admin.", "danger")
                return redirect(url_for('lavorazioni', mese=date.today().strftime('%Y-%m')))

            try:
                d_val = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
                nuovo = Lavorazione(
                    data=d_val,
                    cliente=canonical_cliente_picking(request.form.get('cliente')),
                    descrizione=request.form.get('descrizione'),
                    richiesta_di=request.form.get('richiesta_di'),
                    seriali=request.form.get('seriali'),
                    n_arrivo=request.form.get('n_arrivo'),
                    colli=int(request.form.get('colli') or 0),
                    pallet_forniti=int(request.form.get('pallet_forniti') or 0),
                    pallet_uscita=int(request.form.get('pallet_uscita') or 0),
                    ore_blue_collar=float(request.form.get('ore_blue_collar') or 0),
                    ore_white_collar=float(request.form.get('ore_white_collar') or 0)
                )
                db.add(nuovo)
                db.commit()
                try:
                    totale_cliente = db.query(Lavorazione).filter(
                        normalized_sql_text(Lavorazione.cliente) == normalize_text_key(nuovo.cliente)
                    ).count()
                    flash(f"Picking aggiunto per {nuovo.cliente} in data {nuovo.data}. Record presenti per questo cliente: {totale_cliente}.", "success")
                except Exception:
                    flash(f"Picking aggiunto per {nuovo.cliente} in data {nuovo.data}.", "success")
            except Exception as e:
                db.rollback()
                flash(f"Errore inserimento: {e}", "danger")
            return redirect(url_for('lavorazioni', mese=date.today().strftime('%Y-%m')))

        # --- EDIT MODE (GET ?edit_id=) ---
        edit_id = request.args.get('edit_id')
        edit_row = None
        if edit_id and session.get('role') == 'admin':
            try:
                edit_row = db.query(Lavorazione).filter(Lavorazione.id == int(edit_id)).first()
            except:
                edit_row = None

        # --- VISUALIZZAZIONE ---
        filtri = {
            'data_da': (request.args.get('data_da') or '').strip(),
            'data_a': (request.args.get('data_a') or '').strip(),
            'cliente': (request.args.get('cliente') or '').strip(),
            'descrizione': (request.args.get('descrizione') or '').strip(),
            'richiesta_di': (request.args.get('richiesta_di') or '').strip(),
            'seriali': (request.args.get('seriali') or '').strip(),
            'n_arrivo': (request.args.get('n_arrivo') or '').strip(),
            'colli_da': (request.args.get('colli_da') or '').strip(),
            'colli_a': (request.args.get('colli_a') or '').strip(),
            'pallet_forniti_da': (request.args.get('pallet_forniti_da') or '').strip(),
            'pallet_forniti_a': (request.args.get('pallet_forniti_a') or '').strip(),
            'pallet_uscita_da': (request.args.get('pallet_uscita_da') or '').strip(),
            'pallet_uscita_a': (request.args.get('pallet_uscita_a') or '').strip(),
            'ore_blue_da': (request.args.get('ore_blue_da') or '').strip(),
            'ore_blue_a': (request.args.get('ore_blue_a') or '').strip(),
            'ore_white_da': (request.args.get('ore_white_da') or '').strip(),
            'ore_white_a': (request.args.get('ore_white_a') or '').strip(),
            'mese': (request.args.get('mese') or '').strip(),
            'view': (request.args.get('view') or '').strip(),
        }

        mese_attivo, mese_start, mese_end, vista_mese = _month_bounds_from_request_lavorazioni()
        filtri['mese'] = mese_attivo
        filtri['view'] = vista_mese

        data_da = _safe_date_ymd(filtri['data_da'])
        data_a = _safe_date_ymd(filtri['data_a'])
        colli_da = _safe_int(filtri['colli_da'])
        colli_a = _safe_int(filtri['colli_a'])
        pallet_forniti_da = _safe_int(filtri['pallet_forniti_da'])
        pallet_forniti_a = _safe_int(filtri['pallet_forniti_a'])
        pallet_uscita_da = _safe_int(filtri['pallet_uscita_da'])
        pallet_uscita_a = _safe_int(filtri['pallet_uscita_a'])
        # Se il browser/telefono lascia 0 o 0,0 nei campi "a", non deve diventare
        # un filtro massimo attivo: altrimenti i picking con ore/colli > 0 spariscono
        # (caso riscontrato su GALVANO TECNICA).
        for _k in ['colli_a', 'pallet_forniti_a', 'pallet_uscita_a', 'ore_blue_a', 'ore_white_a']:
            filtri[_k] = _clean_picking_upper_bound(filtri.get(_k))

        colli_a = _safe_int(filtri['colli_a'])
        pallet_forniti_a = _safe_int(filtri['pallet_forniti_a'])
        pallet_uscita_a = _safe_int(filtri['pallet_uscita_a'])

        ore_blue_da = _safe_float_it(filtri['ore_blue_da'])
        ore_blue_a = _safe_float_it(filtri['ore_blue_a'])
        ore_white_da = _safe_float_it(filtri['ore_white_da'])
        ore_white_a = _safe_float_it(filtri['ore_white_a'])

        def _match_txt(value, filtro):
            filtro = (filtro or '').strip()
            if not filtro:
                return True
            v = str(value or '')
            nf = normalize_text_key(filtro)
            nv = normalize_text_key(v)
            if filtro.lower() in v.lower() or nf in nv:
                return True
            if nf in {'GALVANOTECNICA', 'COTUGNOGALVANOTECNICA', 'GALVANO'} and nv == 'GALVANOTECNICA':
                return True
            return False

        # Normalizza i vecchi record Galvano e poi ricarica la lista.
        _normalize_existing_galvano_picking(db)

        dati = (
            db.query(Lavorazione)
            .order_by(Lavorazione.data.desc(), Lavorazione.id.desc())
            .all()
        )

        # Pagina iniziale: solo mese corrente. Con ?mese=YYYY-MM si cambia mese.
        # Con ?view=tutti si vede tutto l'archivio.
        filtered = []
        for rec in dati:
            rec_date = _parse_picking_date_safe(getattr(rec, 'data', None))
            if mese_start and (not rec_date or rec_date < mese_start or rec_date >= mese_end):
                continue
            if data_da and (not rec_date or rec_date < data_da):
                continue
            if data_a and (not rec_date or rec_date > data_a):
                continue
            if not _match_txt(rec.cliente, filtri['cliente']):
                continue
            if not _match_txt(rec.descrizione, filtri['descrizione']):
                continue
            if not _match_txt(rec.richiesta_di, filtri['richiesta_di']):
                continue
            if not _match_txt(rec.seriali, filtri['seriali']):
                continue
            if not _match_txt(getattr(rec, 'n_arrivo', ''), filtri.get('n_arrivo','')):
                continue
            if colli_da is not None and (rec.colli is None or int(rec.colli or 0) < colli_da):
                continue
            if colli_a is not None and (rec.colli is None or int(rec.colli or 0) > colli_a):
                continue
            if pallet_forniti_da is not None and int(rec.pallet_forniti or 0) < pallet_forniti_da:
                continue
            if pallet_forniti_a is not None and int(rec.pallet_forniti or 0) > pallet_forniti_a:
                continue
            if pallet_uscita_da is not None and int(rec.pallet_uscita or 0) < pallet_uscita_da:
                continue
            if pallet_uscita_a is not None and int(rec.pallet_uscita or 0) > pallet_uscita_a:
                continue
            if ore_blue_da is not None and float(rec.ore_blue_collar or 0) < ore_blue_da:
                continue
            if ore_blue_a is not None and float(rec.ore_blue_collar or 0) > ore_blue_a:
                continue
            if ore_white_da is not None and float(rec.ore_white_collar or 0) < ore_white_da:
                continue
            if ore_white_a is not None and float(rec.ore_white_collar or 0) > ore_white_a:
                continue
            filtered.append(rec)


        dati = filtered

        return render_template('lavorazioni.html', lavorazioni=dati, today=date.today(), edit_row=edit_row, filtri=filtri, clienti_validi=get_clienti_utenti(), mese_corrente=date.today().strftime('%Y-%m'), mese_precedente=_prev_next_month(mese_attivo or date.today().strftime('%Y-%m'))[0], mese_successivo=_prev_next_month(mese_attivo or date.today().strftime('%Y-%m'))[1])



    @app.route('/stampa_picking_pdf', methods=['POST'])
    @login_required
    def stampa_picking_pdf():
        if session.get('role') != 'admin':
            return "No Access", 403

        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from sqlalchemy import func

        mese = (request.form.get('mese') or '').strip()       # es '2026-01'
        cliente = (request.form.get('cliente') or '').strip()

        db = SessionLocal()
        try:
            query = db.query(Lavorazione)

            # ✅ Qui trattiamo lavorazioni.data come TEXT -> la convertiamo in DATE (Postgres)
            data_as_date = func.to_date(func.left(Lavorazione.data, 10), 'YYYY-MM-DD')

            # ✅ FILTRO MESE (con range corretto)
            if mese:
                try:
                    year, month = mese.split("-")
                    y = int(year)
                    m = int(month)

                    start_str = f"{y:04d}-{m:02d}-01"
                    if m == 12:
                        end_str = f"{y+1:04d}-01-01"
                    else:
                        end_str = f"{y:04d}-{m+1:02d}-01"

                    query = query.filter(
                        data_as_date >= func.to_date(start_str, 'YYYY-MM-DD'),
                        data_as_date < func.to_date(end_str, 'YYYY-MM-DD')
                    )
                except Exception:
                    # fallback se mese non è valido
                    pass

            # ✅ FILTRO CLIENTE
            if cliente:
                query = query.filter(Lavorazione.cliente.ilike(f"%{cliente}%"))

            # ✅ ORDINAMENTO SICURO (per data convertita)
            rows = query.order_by(data_as_date.asc().nullslast(), Lavorazione.id.asc()).all()

            # --- CREA EXCEL ---
            wb = Workbook()
            ws = wb.active
            ws.title = "Picking"

            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            header_fill = PatternFill("solid", fgColor="D9E1F2")

            ws["A1"] = "REPORT PICKING / LAVORAZIONI"
            ws["A1"].font = Font(bold=True, size=16)
            ws.merge_cells("A1:K1")
            ws["A1"].alignment = center

            ws["A3"] = "Filtri:"
            ws["A3"].font = bold
            ws["B3"] = f"Mese={mese or 'Tutti'} | Cliente={cliente or 'Tutti'}"
            ws.merge_cells("B3:K3")

            headers = [
                "Data", "Cliente", "Descrizione", "Richiesta di", "Seriali/Buono", "N. Arrivo",
                "Colli", "Pallet Entrati", "Pallet Usciti", "Ore Blue", "Ore White"
            ]

            start_row = 5
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col, value=h)
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = center

            riga = start_row + 1

            # Totali
            t_colli = 0
            t_pin = 0
            t_pout = 0
            t_blue = 0.0
            t_white = 0.0

            for r in rows:
                d_str = (str(r.data)[:10] if r.data else "")

                colli = int(r.colli or 0)
                pin = int(r.pallet_forniti or 0)
                pout = int(r.pallet_uscita or 0)
                blue = float(r.ore_blue_collar or 0.0)
                white = float(r.ore_white_collar or 0.0)

                t_colli += colli
                t_pin += pin
                t_pout += pout
                t_blue += blue
                t_white += white

                ws.cell(riga, 1, d_str).alignment = center
                ws.cell(riga, 2, (r.cliente or "")).alignment = left
                ws.cell(riga, 3, (r.descrizione or "")).alignment = left
                ws.cell(riga, 4, (r.richiesta_di or "")).alignment = left
                ws.cell(riga, 5, (r.seriali or "")).alignment = left
                ws.cell(riga, 6, (getattr(r, 'n_arrivo', '') or "")).alignment = left
                ws.cell(riga, 7, colli).alignment = center
                ws.cell(riga, 8, pin).alignment = center
                ws.cell(riga, 9, pout).alignment = center

                c10 = ws.cell(riga, 10, blue);  c10.number_format = '0.00'; c10.alignment = center
                c11 = ws.cell(riga, 11, white); c11.number_format = '0.00'; c11.alignment = center

                riga += 1

            # Riga Totali
            ws.cell(riga, 1, "TOTALI").font = bold
            ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=6)
            ws.cell(riga, 1).alignment = Alignment(horizontal="right", vertical="center")

            ws.cell(riga, 7, t_colli).font = bold
            ws.cell(riga, 8, t_pin).font = bold
            ws.cell(riga, 9, t_pout).font = bold

            tc10 = ws.cell(riga, 10, t_blue); tc10.font = bold; tc10.number_format = '0.00'; tc10.alignment = center
            tc11 = ws.cell(riga, 11, t_white); tc11.font = bold; tc11.number_format = '0.00'; tc11.alignment = center

            widths = [12, 18, 40, 20, 22, 18, 10, 14, 14, 10, 10]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            ws.freeze_panes = "A6"

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

            safe_mese = mese.replace("-", "_") if mese else "TUTTO"
            filename = f"Report_Picking_{safe_mese}.xlsx"

            return send_file(
                bio,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            return f"Errore export Picking Excel: {e}", 500
        finally:
            db.close()


    # --- NUOVO: EXPORT INVENTARIO EXCEL ---

    @app.post('/report_inventario_excel')
    @login_required
    def report_inventario_excel():
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter
        import io
        from datetime import datetime, date
        from collections import defaultdict

        # Client bloccato sul proprio account; Admin seleziona un cliente valido dal form.
        if session.get('role') == 'client':
            cliente_rif = (current_user.id or '').strip()
        else:
            cliente_rif = validate_cliente_or_raise(request.form.get('cliente_inventario'))

        if not cliente_rif:
            return "Cliente mancante", 400

        data_rif_str = (request.form.get('data_inventario') or '').strip()

        def parse_d(v):
            if not v:
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            s = str(v).strip().split(' ')[0][:10]
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    pass
            return None

        if data_rif_str:
            d_limit = parse_d(data_rif_str)
            if not d_limit:
                return "Formato data inventario non valido", 400
        else:
            d_limit = date.today()

        cliente_key = normalize_text_key(cliente_rif)
        usa_pezzi = (cliente_key == normalize_text_key("GALVANO TECNICA"))
        is_duferco = (cliente_key == normalize_text_key("DUFERCO"))

        db = SessionLocal()
        try:
            # Match ESATTO e normalizzato: evita che l'admin mescoli clienti simili
            articoli = (
                db.query(Articolo)
                .filter(normalized_sql_text(Articolo.cliente) == cliente_key)
                .all()
            )

            agg = defaultdict(lambda: {
                "descrizione": "",
                "serial_number": "",
                "entrata": 0,
                "uscita": 0
            }) if is_duferco else defaultdict(lambda: {
                "descrizione": "",
                "entrata": 0,
                "uscita": 0
            })

            for art in articoli:
                codice = (art.codice_articolo or "").strip()
                if not codice:
                    continue

                serial = (getattr(art, 'serial_number', None) or '').strip()
                key = (codice, serial) if is_duferco else codice
                descr = (art.descrizione or "").strip()

                if usa_pezzi:
                    q_raw = getattr(art, "pezzi", None)
                    if q_raw is None:
                        q_raw = getattr(art, "pezzo", None)
                else:
                    q_raw = getattr(art, "n_colli", None)

                try:
                    qty = int(float(str(q_raw).replace(',', '.'))) if str(q_raw).strip() != '' else 0
                except Exception:
                    qty = 0

                if descr and not agg[key]["descrizione"]:
                    agg[key]["descrizione"] = descr

                if is_duferco and serial and not agg[key]["serial_number"]:
                    agg[key]["serial_number"] = serial

                d_ing = parse_d(getattr(art, "data_ingresso", None))
                d_usc = parse_d(getattr(art, "data_uscita", None))

                if d_ing and d_ing <= d_limit:
                    agg[key]["entrata"] += qty

                if d_usc and d_usc <= d_limit:
                    agg[key]["uscita"] += qty

            righe = []
            for k in sorted(agg.keys()):
                data = agg[k]
                entrata = int(data.get("entrata", 0) or 0)
                uscita = int(data.get("uscita", 0) or 0)
                rimanenza = entrata - uscita

                if entrata == 0 and uscita == 0 and rimanenza == 0:
                    continue

                if is_duferco:
                    codice, serial = k
                else:
                    codice, serial = k, ""

                righe.append({
                    "codice": codice,
                    "serial_number": serial,
                    "descrizione": data.get("descrizione", ""),
                    "entrata": entrata,
                    "uscita": uscita,
                    "rimanenza": rimanenza
                })

            wb = Workbook()
            ws = wb.active
            ws.title = "INVENTARIO"

            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")
            left = Alignment(horizontal="left", vertical="center")
            header_fill = PatternFill("solid", fgColor="D9E1F2")
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            oggi_str = datetime.now().strftime("%Y-%m-%d")
            data_limite_str = d_limit.strftime("%Y-%m-%d")
            tipo = "PEZZI" if usa_pezzi else "COLLI"

            ws["A1"] = "ELENCO ARTICOLI"
            ws["A2"] = f"Cliente: {cliente_rif}"
            ws["A3"] = f"Inventario basato su: {tipo}"
            ws["A4"] = f"Inventario al: {data_limite_str}"
            ws["A5"] = f"Generato il: {oggi_str}"

            ws["A1"].font = Font(bold=True, size=14)
            ws["A2"].font = bold
            ws["A3"].font = bold
            ws["A4"].font = bold
            ws["A5"].font = bold

            headers = ["ID", "CODICE ARTICOLO"]
            if is_duferco:
                headers.append("SERIAL NUMBER")
            headers += [
                "DESCRIZIONE",
                f"Q.TA ENTRATA ({tipo})",
                f"Q.TA USCITA ({tipo})",
                f"RIMANENZA ({tipo})"
            ]

            start_row = 7
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=c, value=h)
                cell.font = bold
                cell.alignment = center
                cell.fill = header_fill
                cell.border = border

            r = start_row + 1
            idx = 1
            for row in righe:
                values = [str(idx).zfill(3), row["codice"]]
                if is_duferco:
                    values.append(row.get("serial_number", ""))
                values += [row["descrizione"], row["entrata"], row["uscita"], row["rimanenza"]]

                for c, v in enumerate(values, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.alignment = left if c in (2, 3, 4 if is_duferco else 3) else center
                    cell.border = border
                r += 1
                idx += 1

            ws.freeze_panes = f"A{start_row + 1}"

            if r > start_row + 1:
                last_col = get_column_letter(len(headers))
                tab = Table(displayName="TabInventario", ref=f"A{start_row}:{last_col}{r-1}")
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                ws.add_table(tab)

            widths = [8, 24]
            if is_duferco:
                widths.append(22)
            widths += [55, 22, 22, 22]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

            filename = f"Inventario_{cliente_rif.replace(' ', '_')}_{data_limite_str}.xlsx"
            return send_file(
                bio,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        finally:
            db.close()

    # =========================
    # IMPORT EXCEL (con log)
    # =========================
