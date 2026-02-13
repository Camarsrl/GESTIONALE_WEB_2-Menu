# -*- coding: utf-8 -*-
"""
Camar • Gestionale Web – build aggiornata (Ottobre 2025)
© Copyright Alessia Moncalvo
Tutti i diritti riservati.
"""
import os
import io
import re
import uuid
import json
import logging
import calendar
import smtplib
import hashlib  # <--- QUESTO MANCAVA E CAUSA ERRORI
import math
import time
import mimetypes
from urllib.parse import unquote
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import defaultdict
from functools import wraps

# Excel e PDF
import pandas as pd
import pdfplumber

# Email
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email.mime.application import MIMEApplication
from email import encoders

# Flask
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify, render_template_string, abort
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_mail import Mail
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

# Database (SQLAlchemy)
from sqlalchemy import create_engine, Column, Integer, String, Text, Float, Date, ForeignKey, Boolean, or_, Identity, text, Index, inspect
from sqlalchemy.orm import declarative_base, sessionmaker, relationship, scoped_session, selectinload
from sqlalchemy.sql import func
from sqlalchemy.exc import IntegrityError

# ReportLab (PDF)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak

# Jinja
from jinja2 import DictLoader
# ========================================================
# 1. INIZIALIZZAZIONE APP E LOGIN MANAGER (ORDINE CORRETTO)
# ========================================================
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "chiave_segreta_super_sicura")

# Inizializza LoginManager SUBITO
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# =========================
# Formattazione numeri ITA
# =========================
def it_num(value, decimals=2):
    """Formatta un numero con la virgola come separatore decimale (stile IT)."""
    if value is None or value == "":
        return ""
    try:
        v = float(value)
    except Exception:
        return str(value)

    s = f"{v:.{int(decimals)}f}"
    return s.replace('.', ',')


# filtro Jinja: {{ value|it_num(2) }}
app.jinja_env.filters['it_num'] = it_num


# =========================
# Helpers debug mappe file
# =========================


def to_date_db(val):
    """
    Converte val in datetime.date (per DB).
    Gestisce: datetime/date, pandas Timestamp, numeri Excel (seriali), stringhe.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)) or pd.isna(val) or val == '':
        return None

    # pandas Timestamp / datetime / date
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    # Excel serial date (es. 45234)
    if isinstance(val, (int, float)):
        try:
            # Excel origin: 1899-12-30
            return (datetime(1899, 12, 30) + timedelta(days=int(val))).date()
        except Exception:
            return None

    # Stringa
    s = str(val).strip()
    if not s:
        return None

    # Tentativi di parsing formati comuni
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    # Ultima spiaggia: pandas to_datetime
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if not pd.isna(dt):
            return dt.date()
    except Exception:
        pass
        
    return None

def _file_digest(p: Path) -> str:
    """MD5 del file (per capire se su Render stai usando davvero la versione corretta)."""
    try:
        data = p.read_bytes()
        return hashlib.md5(data).hexdigest()
    except Exception:
        return "N/A"


app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', '587'))

app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() == 'true'
app.config['MAIL_USE_SSL'] = os.environ.get('MAIL_USE_SSL', 'false').lower() == 'true'

app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = app.config['MAIL_USERNAME']

mail = Mail(app)

# ========================================================
# 2. CONFIGURAZIONE PATH E FILES
# ========================================================
APP_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = APP_DIR / "static"

# --- GESTIONE DISCO PERSISTENTE (Percorso Forzato) ---
persistent_path = "/var/data/app"

if os.path.exists(persistent_path):
    # Se la cartella del disco esiste fisicamente, usala!
    MEDIA_DIR = Path(persistent_path)
    print(f"✅ USO DISCO PERSISTENTE RENDER: {MEDIA_DIR}")
else:
    # Altrimenti usa cartella locale
    MEDIA_DIR = APP_DIR / "media"
    print(f"⚠️ USO DISCO LOCALE (Temporaneo): {MEDIA_DIR}")

DOCS_DIR = MEDIA_DIR / "docs"
PHOTOS_DIR = MEDIA_DIR / "photos"


# ========================================================
#  BACKUP (DB + JSON + Media) - crea ZIP in /media/backups
# ========================================================
BACKUP_DIR = MEDIA_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

def create_backup_zip(include_media: bool = True) -> Path:
    """Crea un backup ZIP e ritorna il path."""
    import zipfile
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = BACKUP_DIR / f"backup_camar_{ts}.zip"

    def _safe_add(zf, p: Path, arcname: str):
        try:
            if p.exists():
                zf.write(p, arcname=arcname)
        except Exception as e:
            print(f"[WARN] backup skip {p}: {e}")

    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        # DB
        _safe_add(zf, APP_DIR / "magazzino.db", "magazzino.db")

        # Config / JSON
        for name in ["mappe_excel.json", "destinatari_saved.json", "progressivi_ddt.json"]:
            _safe_add(zf, APP_DIR / name, f"config/{name}")
            _safe_add(zf, MEDIA_DIR / name, f"config/{name}")  # se sta sul disco

        _safe_add(zf, _rubrica_email_path(), "config/rubrica_email.json")

        # Media (docs + photos)
        if include_media:
            for folder, arcroot in [(DOCS_DIR, "media/docs"), (PHOTOS_DIR, "media/photos")]:
                if folder.exists():
                    for p in folder.rglob("*"):
                        if p.is_file():
                            _safe_add(zf, p, f"{arcroot}/{p.name}")

    return out


# --- CONFIGURAZIONE FILE MAPPE EXCEL ---
# Definiamo qui i percorsi esatti per evitare confusione
MAPPE_FILE_PERSISTENT = MEDIA_DIR / "mappe_excel.json"        # File modificabile (nel disco dati)
MAPPE_FILE_ORIGINAL = APP_DIR / "config." / "mappe_excel.json" # File originale (da GitHub)

# Crea le cartelle se non esistono
for d in (STATIC_DIR, MEDIA_DIR, DOCS_DIR, PHOTOS_DIR):
    d.mkdir(parents=True, exist_ok=True)


# ================================
# BACKUP AUTOMATICO (OGNI 2 ORE)
# ================================

import time
import os

_AUTO_BACKUP_LAST_CHECK = 0


def auto_backup_if_due():
    """
    Backup automatico:
    - controllo max ogni 10 minuti
    - crea backup se ultimo è più vecchio di 2 ore
    - mantiene solo gli ultimi 50 backup
    """

    global _AUTO_BACKUP_LAST_CHECK

    try:
        now = time.time()

        # ✅ evita controllo continuo: max ogni 10 minuti
        if _AUTO_BACKUP_LAST_CHECK and (now - _AUTO_BACKUP_LAST_CHECK) < 600:
            return

        _AUTO_BACKUP_LAST_CHECK = now

        # ✅ possibilità di disattivare via ENV
        if str(os.environ.get("AUTO_BACKUP", "1")).lower() in ("0", "false", "no", "off"):
            app.logger.info("[AUTO_BACKUP] disabilitato via AUTO_BACKUP=0")
            return

        # ✅ crea cartella backup se manca
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)

        # ✅ trova ultimo backup
        backups = sorted(
            BACKUP_DIR.glob("backup_camar_*.zip"),
            key=lambda p: p.stat().st_mtime,
            reverse=True
        )

        latest = backups[0] if backups else None

        if latest:
            ore_passate = (now - latest.stat().st_mtime) / 3600.0
            app.logger.info(f"[AUTO_BACKUP] ultimo backup {latest.name} ({ore_passate:.1f} ore fa)")
        else:
            app.logger.info("[AUTO_BACKUP] nessun backup trovato, ne creo uno ora")

        # ✅ intervallo BACKUP: ogni 2 ore
        INTERVALLO = 2 * 3600  # 2 ore

        if (latest is None) or ((now - latest.stat().st_mtime) > INTERVALLO):
            app.logger.warning("[AUTO_BACKUP] CREAZIONE backup automatico in corso...")

            zip_path = create_backup_zip(include_media=True)

            app.logger.warning(f"[AUTO_BACKUP] OK creato: {zip_path}")

            # ✅ mantiene solo ultimi 50 backup
            MAX_FILES = 50
            if len(backups) > MAX_FILES:
                for old in backups[MAX_FILES:]:
                    try:
                        old.unlink()
                        app.logger.info(f"[AUTO_BACKUP] eliminato backup vecchio: {old.name}")
                    except:
                        pass

        else:
            app.logger.info("[AUTO_BACKUP] skip: non sono passate 2 ore")

    except Exception as e:
        app.logger.warning(f"[AUTO_BACKUP] fallito: {e}")


# ✅ Hook automatico su ogni request (non blocca mai)
@app.before_request
def _auto_backup_hook():
    try:
        auto_backup_if_due()
    except Exception:
        pass

def pulisci_backup_vecchi(max_files=50):
    files = sorted(
        Path(BACKUP_DIR).glob("backup_*.zip"),
        key=os.path.getmtime,
        reverse=True
    )
    for f in files[max_files:]:
        f.unlink()


from pathlib import Path
import zipfile
import tempfile
import shutil
from datetime import datetime

# Assumo che tu abbia già:
# BACKUP_DIR = Path("/var/data/app/backups")
# MEDIA_DIR = Path("/var/data/app")
# e che magazzino.db stia in MEDIA_DIR

def _get_db_path():
    # Percorso DB (modifica qui se nel tuo progetto è diverso)
    return (MEDIA_DIR / "magazzino.db")

def list_backups():
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(BACKUP_DIR.glob("backup_camar_*.zip"), key=lambda p: p.stat().st_mtime, reverse=True)
    out = []
    for p in files:
        out.append({
            "name": p.name,
            "path": p,
            "size_mb": round(p.stat().st_size / (1024 * 1024), 2),
            "mtime": datetime.fromtimestamp(p.stat().st_mtime).strftime("%d/%m/%Y %H:%M")
        })
    return out

def restore_from_backup_zip(zip_filename: str, restore_media: bool = False):
    """
    Ripristino sicuro:
    - valida che il file stia dentro BACKUP_DIR
    - crea una copia di emergenza del DB attuale
    - estrae lo zip in temp
    - ripristina magazzino.db + JSON
    - opzionale: ripristina cartelle docs/photos
    """
    # ✅ sicurezza: niente path traversal
    zip_path = (BACKUP_DIR / zip_filename).resolve()
    if not str(zip_path).startswith(str(BACKUP_DIR.resolve())):
        raise Exception("Backup non valido (path non consentito).")
    if not zip_path.exists():
        raise Exception("Backup non trovato.")

    db_path = _get_db_path()
    MEDIA_DIR.mkdir(parents=True, exist_ok=True)

    # ✅ copia emergenza DB attuale
    if db_path.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        emergency = db_path.with_suffix(f".pre_restore_{ts}.bak")
        shutil.copy2(db_path, emergency)

    # ✅ estrai in temp e ripristina
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(tmpdir)

        # --- ripristina DB ---
        extracted_db = tmpdir / "magazzino.db"
        if extracted_db.exists():
            shutil.copy2(extracted_db, db_path)
        else:
            raise Exception("Nel backup non c'è magazzino.db")

        # --- ripristina JSON (se presenti) ---
        for json_name in ["mappe_excel.json", "destinatari_saved.json", "rubrica_email.json"]:
            src = tmpdir / json_name
            if src.exists():
                shutil.copy2(src, MEDIA_DIR / json_name)

        # --- ripristina media (opzionale) ---
        if restore_media:
            for folder in ["docs", "photos"]:
                src_folder = tmpdir / folder
                dst_folder = MEDIA_DIR / folder
                if src_folder.exists():
                    if dst_folder.exists():
                        shutil.rmtree(dst_folder)
                    shutil.copytree(src_folder, dst_folder)

    return True

def _discover_logo_path():
    # Lista aggiornata con il nome corretto del tuo file
    possible_names = [
        "logo camar.jpg",  # <--- Questo è quello che hai su GitHub
        "logo.png", 
        "logo.jpg", 
        "logo.jpeg", 
        "logo_camar.png"
    ]
    
    print(f"DEBUG: Cerco logo in {STATIC_DIR}")
    
    for name in possible_names:
        p = STATIC_DIR / name
        if p.exists(): 
            print(f"DEBUG: Logo TROVATO: {p}")
            return str(p)
            
    print("DEBUG: NESSUN logo trovato.")
    return None

LOGO_PATH = _discover_logo_path()

# ========================================================
# 3. CONFIGURAZIONE DATABASE (Render-safe)
# ========================================================
import os
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, scoped_session, declarative_base

DB_URL = os.environ.get("DATABASE_URL", "").strip()

# ✅ In produzione: MAI fallback silenzioso a sqlite
IS_RENDER = bool(os.environ.get("RENDER")) or bool(os.environ.get("RENDER_SERVICE_ID"))

if not DB_URL:
    if IS_RENDER:
        raise RuntimeError("DATABASE_URL non è impostata su Render! Controlla le Environment Variables del service.")
    DB_URL = f"sqlite:///{APP_DIR / 'magazzino.db'}"

# Render a volte usa postgres:// -> SQLAlchemy vuole postgresql://
if DB_URL.startswith("postgres://"):
    DB_URL = DB_URL.replace("postgres://", "postgresql://", 1)

def _normalize_db_url(u: str) -> str:
    # opzionale: mysql
    if u.startswith("mysql://"):
        u = "mysql+pymysql://" + u[len("mysql://"):]
    return u

DB_URL = _normalize_db_url(DB_URL)

# ✅ pool_pre_ping evita connessioni "morte" (tipico su hosting)
engine = create_engine(
    DB_URL,
    future=True,
    echo=False,
    pool_pre_ping=True
)

SessionLocal = scoped_session(
    sessionmaker(bind=engine, autoflush=False, autocommit=False, expire_on_commit=False)
)

Base = declarative_base()

# ✅ IMPORTANTISSIMO: rimuove la sessione a fine request (evita problemi con più utenti/worker)
@app.teardown_appcontext
def remove_scoped_session(exception=None):
    SessionLocal.remove()

# ========================================================
# 4. DEFINIZIONE MODELLI
# ========================================================
class Articolo(Base):
    __tablename__ = "articoli"
    id_articolo = Column(Integer, Identity(start=1), primary_key=True)
    codice_articolo = Column(Text); descrizione = Column(Text)
    cliente = Column(Text); fornitore = Column(Text); magazzino = Column(String(255))
    protocollo = Column(Text); ordine = Column(Text); commessa = Column(Text)
    buono_n = Column(Text); n_arrivo = Column(Text); ns_rif = Column(String(255))
    serial_number = Column(String(255)); pezzo = Column(String(255))
    n_colli = Column(Integer); peso = Column(Float)
    larghezza = Column(Float); lunghezza = Column(Float); altezza = Column(Float)
    m2 = Column(Float); m3 = Column(Float); posizione = Column(String(255))
    stato = Column(String(255)); note = Column(Text); mezzi_in_uscita = Column(String(255))
    data_ingresso = Column(String(32)); n_ddt_ingresso = Column(Text)
    data_uscita = Column(String(32)); n_ddt_uscita = Column(Text)
    attachments = relationship("Attachment", back_populates="articolo", cascade="all, delete-orphan", passive_deletes=True)
    lotto = Column(Text) # <--- AGGIUNGI QUESTA

class Attachment(Base):
    __tablename__ = "attachments"
    id = Column(Integer, Identity(start=1), primary_key=True)
    articolo_id = Column(Integer, ForeignKey("articoli.id_articolo", ondelete='CASCADE'), nullable=False)
    kind = Column(String(10)); filename = Column(String(512))
    articolo = relationship("Articolo", back_populates="attachments")

Base.metadata.create_all(engine)


# ========================================================
# 4b. INDICI DB (performance query) - SAFE (checkfirst)
# ========================================================

def ensure_db_indexes(engine):
    # Crea indici importanti se mancano (checkfirst).
    # Nota: su MySQL alcuni campi Text richiedono una lunghezza; la impostiamo solo su MySQL.
    try:
        dialect = engine.dialect.name
        mysql_len = 191 if dialect == 'mysql' else None

        idx_specs = []
        # Campi molto usati nei filtri/ricerche
        idx_specs.append(('ix_articoli_id_articolo', [Articolo.id_articolo], {}))
        idx_specs.append(('ix_articoli_magazzino', [Articolo.magazzino], {}))
        idx_specs.append(('ix_articoli_posizione', [Articolo.posizione], {}))
        idx_specs.append(('ix_articoli_serial_number', [Articolo.serial_number], {}))
        idx_specs.append(('ix_articoli_ns_rif', [Articolo.ns_rif], {}))

        # Text: cliente/codice/protocollo/ordine spesso ricercati
        if mysql_len:
            idx_specs.append(('ix_articoli_cliente', [Articolo.cliente], {'mysql_length': mysql_len}))
            idx_specs.append(('ix_articoli_codice_articolo', [Articolo.codice_articolo], {'mysql_length': mysql_len}))
            idx_specs.append(('ix_articoli_protocollo', [Articolo.protocollo], {'mysql_length': mysql_len}))
            idx_specs.append(('ix_articoli_ordine', [Articolo.ordine], {'mysql_length': mysql_len}))
            idx_specs.append(('ix_articoli_buono_n', [Articolo.buono_n], {'mysql_length': mysql_len}))
            idx_specs.append(('ix_articoli_n_arrivo', [Articolo.n_arrivo], {'mysql_length': mysql_len}))
            idx_specs.append(('ix_articoli_ddt_ingresso', [Articolo.n_ddt_ingresso], {'mysql_length': mysql_len}))
            idx_specs.append(('ix_articoli_ddt_uscita', [Articolo.n_ddt_uscita], {'mysql_length': mysql_len}))
        else:
            idx_specs.append(('ix_articoli_cliente', [Articolo.cliente], {}))
            idx_specs.append(('ix_articoli_codice_articolo', [Articolo.codice_articolo], {}))
            idx_specs.append(('ix_articoli_protocollo', [Articolo.protocollo], {}))
            idx_specs.append(('ix_articoli_ordine', [Articolo.ordine], {}))
            idx_specs.append(('ix_articoli_buono_n', [Articolo.buono_n], {}))
            idx_specs.append(('ix_articoli_n_arrivo', [Articolo.n_arrivo], {}))
            idx_specs.append(('ix_articoli_ddt_ingresso', [Articolo.n_ddt_ingresso], {}))
            idx_specs.append(('ix_articoli_ddt_uscita', [Articolo.n_ddt_uscita], {}))

        # Date come stringa: indice comunque utile per ordinamenti/filtri grezzi
        idx_specs.append(('ix_articoli_data_ingresso', [Articolo.data_ingresso], {}))
        idx_specs.append(('ix_articoli_data_uscita', [Articolo.data_uscita], {}))

        for name, cols, kwargs in idx_specs:
            try:
                Index(name, *cols, **kwargs).create(bind=engine, checkfirst=True)
            except Exception as e:
                print(f"[WARN] indice non creato {name}: {e}")
    except Exception as e:
        print(f"[WARN] ensure_db_indexes fallita: {e}")

# Crea indici (se possibile) all'avvio
ensure_db_indexes(engine)


# --- MODELLO TABELLA TRASPORTI (Separata da Articoli) ---
class Trasporto(Base):
    __tablename__ = 'trasporti'
    id = Column(Integer, primary_key=True)
    data = Column(Date)
    tipo_mezzo = Column(Text)       # Es. Motrice, Bilico
    cliente = Column(Text)
    trasportatore = Column(Text)
    ddt_uscita = Column(Text)       # N. DDT
    magazzino = Column(Text)        # Nuovo
    consolidato = Column(Text)      # Nuovo
    costo = Column(Float)

# --- MODELLO TABELLA PICKING / LAVORAZIONI (Separata da Articoli) ---
class Lavorazione(Base):
    __tablename__ = 'lavorazioni'
    id = Column(Integer, primary_key=True)
    data = Column(Date)
    cliente = Column(Text)
    descrizione = Column(Text)
    richiesta_di = Column(Text)     # Nuovo
    seriali = Column(Text)          # Nuovo
    colli = Column(Integer)
    pallet_forniti = Column(Integer) # Pallet IN
    pallet_uscita = Column(Integer)  # Pallet OUT
    ore_blue_collar = Column(Float)  # Ore Blue
    ore_white_collar = Column(Float) # Ore White

# ========================================================
# 5. GESTIONE UTENTI (Definizione PRIMA dell'uso)
# ========================================================
DEFAULT_USERS = {
    'DE WAVE': 'Struppa01', 'FINCANTIERI': 'Struppa02', 'FINCANTIERI SCOPERTO': 'Struppa12',
    'FINCANTIERI ELETTRICO': 'Struppa13','RF-DE WAVE': 'Struppa03',
    'SGDP': 'Struppa04', 'WINGECO': 'Struppa05', 'AMICO': 'Struppa06', 'DUFERCO': 'Struppa07',
    'SCORZA': 'Struppa08', 'MARINE INTERIORS': 'Struppa09', 'GALVANO TECNICA': 'Struppa10', 'DE WAVE SAMA': 'Struppa11','OPS': '271214',
    'CUSTOMS': 'Balleydier01', 'TAZIO': 'Balleydier02', 'DIEGO': 'Balleydier03', 'ADMIN': 'admin123'
}
ADMIN_USERS = {'ADMIN', 'OPS', 'CUSTOMS', 'TAZIO', 'DIEGO'}


def require_admin(view_func):
    """Decorator: allow only admin users."""
    @wraps(view_func)
    def _wrapped(*args, **kwargs):
        if session.get('role') != 'admin':
            flash("Accesso negato.", "danger")
            return redirect(url_for('giacenze'))
        return view_func(*args, **kwargs)
    return _wrapped

@app.route("/admin/backups", methods=["GET", "POST"])
@login_required
@require_admin
def admin_backups():

    if request.method == "POST":
        action = request.form.get("action")
        filename = request.form.get("filename", "")
        restore_media = (request.form.get("restore_media") == "1")

        try:
            if action == "restore":
                restore_from_backup_zip(filename, restore_media=restore_media)
                flash("✅ Ripristino completato!", "success")
            else:
                flash("Azione non valida.", "warning")

        except Exception as e:
            flash(f"Errore ripristino: {e}", "danger")

        return redirect(url_for("admin_backups"))

    backups = list_backups()
    return render_template_string(ADMIN_BACKUPS_HTML, backups=backups)


@app.route("/admin/backups/download/<path:filename>")
@login_required
@require_admin
def admin_backup_download(filename):
    # ✅ sicurezza path
    p = (BACKUP_DIR / filename).resolve()
    if not str(p).startswith(str(BACKUP_DIR.resolve())) or not p.exists():
        flash("Backup non trovato.", "danger")
        return redirect(url_for("admin_backups"))

    return send_file(p, as_attachment=True, download_name=p.name)

def current_cliente():
    """Cliente associato all'utente corrente (per i client è bloccato)."""
    if session.get('role') == 'client':
        return (current_user.id or '').strip()
    return None

def get_users():
    """Legge utenti dal file txt o usa i default."""
    try:
        fp = APP_DIR / "password Utenti Gestionale.txt"
        if fp.exists():
            content = fp.read_text(encoding="utf-8", errors="ignore")
            pairs = re.findall(r"'([^']+)'\s*[:=]\s*'?([^']+)'?", content)
            if pairs:
                return {k.strip().upper(): v.strip().replace("'", "") for k, v in pairs}
    except Exception as e:
        print(f"Errore lettura file utenti: {e}")
    return DEFAULT_USERS

# ORA possiamo chiamarla, perché è stata definita sopra
USERS_DB = get_users()


def _is_werkzeug_hash(s: str) -> bool:
    if not isinstance(s, str):
        return False
    s = s.strip()
    return s.startswith('pbkdf2:') or s.startswith('scrypt:') or s.startswith('argon2:')


def verify_password(stored: str, provided: str) -> bool:
    # Supporta sia password legacy in chiaro sia hash werkzeug.
    if stored is None:
        return False
    stored = str(stored).strip()
    provided = (provided or '').strip()
    if not stored or not provided:
        return False
    if _is_werkzeug_hash(stored):
        try:
            return check_password_hash(stored, provided)
        except Exception:
            return False
    return stored == provided


class User(UserMixin):
    def __init__(self, id, role):
        self.id = id; self.role = role

@login_manager.user_loader
def load_user(user_id):
    users_db = get_users() 
    if user_id in users_db:
        role = 'admin' if user_id in ADMIN_USERS else 'client'
        return User(user_id, role)
    return None

# --- UTILS ---


# --- HELPER ESTRAZIONE PDF (Necessario per Import PDF) ---

def extract_data_from_ddt_pdf(path):
    import pdfplumber
    import re
    from datetime import date

    def _to_float_it(s):
        s = (s or "").strip()
        if not s:
            return None
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except:
            return None

    def _to_int(s):
        try:
            return int(str(s).strip())
        except:
            return None

    meta = {
        "cliente": "",
        "fornitore": "",
        "commessa": "",
        "n_ddt": "",
        "data_ingresso": date.today().strftime("%Y-%m-%d"),
    }
    extracted_rows = []

    with pdfplumber.open(path) as pdf:
        full_text = ""
        for page in pdf.pages:
            txt = page.extract_text() or ""
            full_text += txt + "\n"

    # -----------------------
    # META (testata)
    # -----------------------

    # Cliente: prende la prima riga utile dopo "Destinatario merci"
    m = re.search(r"Destinatario\s+merci\s*\n([^\n]+)", full_text, flags=re.IGNORECASE)
    if m:
        meta["cliente"] = m.group(1).strip()

    # Fornitore: prova "Merce di proprieta di" + riga sotto, altrimenti prima ragione sociale in alto
    m = re.search(r"Merce\s+di\s+propriet[aà]\s+di\s*\n([^\n]+)", full_text, flags=re.IGNORECASE)
    if m:
        meta["fornitore"] = m.group(1).strip()
    else:
        # fallback: prima riga "tipo ragione sociale" (molto conservativo)
        top_lines = [l.strip() for l in full_text.splitlines() if l.strip()]
        if top_lines:
            meta["fornitore"] = top_lines[0][:80]

    # Numero Bolla / DDT
    m = re.search(r"Numero\s+Bolla\s+([A-Z0-9\/\-]+)", full_text, flags=re.IGNORECASE)
    if m:
        meta["n_ddt"] = m.group(1).strip()
    else:
        # fallback: cerca pattern tipo AT260209
        m2 = re.search(r"\b[A-Z]{1,3}\d{4,10}\b", full_text)
        if m2:
            meta["n_ddt"] = m2.group(0).strip()

    # Data Bolla -> data_ingresso
    m = re.search(r"Data\s+Bolla\s+(\d{2}\/\d{2}\/\d{4})", full_text, flags=re.IGNORECASE)
    if m:
        ddmmyyyy = m.group(1).strip()
        # converte in YYYY-MM-DD
        try:
            d, mth, y = ddmmyyyy.split("/")
            meta["data_ingresso"] = f"{y}-{mth}-{d}"
        except:
            pass

    # Commessa / riferimenti (se presenti)
    # qui puoi aggiungere regole specifiche quando vuoi

    # -----------------------
    # RIGHE ARTICOLI
    # -----------------------
    # Pattern codice: molto tollerante ma evita di prendere parole normali
    code_re = re.compile(r"\b[0-9A-Z]{3,}(?:[-\/][0-9A-Z]{2,}){1,}\b")

    lines = [l.strip() for l in full_text.splitlines() if l.strip()]

    for line in lines:
        # cerca un codice all'inizio riga (o comunque presente)
        code_m = code_re.search(line)
        if not code_m:
            continue

        codice = code_m.group(0).strip()

        # euristica: righe "Lotto ..." non sono righe articolo principali
        if line.lower().startswith("lotto"):
            continue

        # prova a catturare: ... Imballo <TXT> Colli <INT> ... UM <TXT> Qta <NUM>
        # es: ".... CAN 2 50,00 52,20 KG 50,00"
        # prendiamo colli = primo intero "piccolo" dopo il codice; qta = ultimo numero con virgola/punto
        rest = line.replace(codice, "", 1).strip()
        parts = rest.split()

        colli = None
        qta = None
        um = ""

        # qta: ultimo numero nel testo
        nums = re.findall(r"\d+(?:[.,]\d+)?", line)
        if nums:
            qta = _to_float_it(nums[-1])

        # colli: cerca un intero plausibile (1..999) nel mezzo
        for p in parts:
            if p.isdigit():
                v = _to_int(p)
                if v is not None and 1 <= v <= 999:
                    colli = v
                    break

        # UM: prova a prendere token tipo KG/PZ/NR
        um_m = re.search(r"\b(KG|PZ|NR|N|UN)\b", line, flags=re.IGNORECASE)
        if um_m:
            um = um_m.group(1).upper()

        # descrizione: togliamo pezzi “tecnici” (imballo/um/valori) ma restiamo conservativi
        descrizione = rest
        # rimuovi imballo comune tipo CAN, PAL, BOX, ecc. (opzionale)
        descrizione = re.sub(r"\b(CAN|PAL|BOX|CRT|CASS)\b", "", descrizione, flags=re.IGNORECASE).strip()
        # se colli è presente, togli la prima occorrenza
        if colli is not None:
            descrizione = re.sub(rf"\b{colli}\b", "", descrizione, count=1).strip()
        # se UM trovato, toglilo
        if um:
            descrizione = re.sub(rf"\b{um}\b", "", descrizione, flags=re.IGNORECASE).strip()

        extracted_rows.append({
            "codice": codice,
            "descrizione": descrizione,
            "colli": colli if colli is not None else 1,
            "pezzi": qta if qta is not None else 1,
            "um": um
        })

    # de-dup: se stesso codice+descrizione ripetuto identico, somma colli/pezzi
    merged = {}
    for r in extracted_rows:
        key = (r["codice"], r["descrizione"])
        if key not in merged:
            merged[key] = r
        else:
            merged[key]["colli"] = int(merged[key]["colli"] or 0) + int(r["colli"] or 0)
            try:
                merged[key]["pezzi"] = float(merged[key]["pezzi"] or 0) + float(r["pezzi"] or 0)
            except:
                pass

    return meta, list(merged.values())


def is_blank(v):
    try:
        if pd.isna(v): return True
    except Exception: pass
    return (v is None) or (isinstance(v, str) and not v.strip())

def to_float_eu(val):
    """Converte stringa '1,2' in float 1.2. Se vuoto o errore, restituisce 0.0"""
    if not val:
        return 0.0
    if isinstance(val, (float, int)):
        return float(val)
    try:
        # Sostituisce virgola con punto e rimuove spazi
        return float(str(val).replace(',', '.').strip())
    except:
        return 0.0

def to_int_eu(val):
    """Converte stringa in intero sicuro."""
    if not val:
        return 0
    try:
        return int(float(str(val).replace(',', '.')))
    except:
        return 0


def parse_date_ui(s):
    s = (s or "").strip()
    if not s:
        return None

    # supporta sia YYYY-MM-DD (input date) sia DD/MM/YYYY
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def fmt_date(d):
    if not d: return ""
    try:
        if isinstance(d, (datetime, date)):
            return d.strftime("%d/%m/%Y")
        return datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception: return d

def calc_m2_m3(L, P, H, colli=1):
    """Calcola M2 e M3 basandosi su metri o centimetri."""
    # Convertiamo tutto in float
    l = to_float_eu(L)
    p = to_float_eu(P)
    h = to_float_eu(H)
    c = max(1, to_int_eu(colli))

    # Logica intelligente: se le misure sono grandi (>10), assumiamo siano CM e convertiamo in METRI
    # Esempio: 120 -> 1.20
    if l > 10: l /= 100.0
    if p > 10: p /= 100.0
    if h > 10: h /= 100.0

    m3 = l * p * h * c
    # M2: Spesso si intende l'ingombro a terra (L x P) oppure la superficie. 
    # Qui calcoliamo L * P * Colli (Floor space)
    m2 = l * p * c
    
    return round(m2, 4), round(m3, 4)

def load_destinatari():
    DESTINATARI_JSON = APP_DIR / "destinatari_saved.json"
    data = {}
    if DESTINATARI_JSON.exists():
        try:
            content = DESTINATARI_JSON.read_text(encoding="utf-8")
            raw_data = json.loads(content)
            
            # Se il JSON è una lista (vecchio formato), lo convertiamo in dizionario
            if isinstance(raw_data, list):
                for item in raw_data:
                    # Usa il campo 'Cliente' come chiave, o genera un nome se manca
                    key = item.get("Cliente") or item.get("ragione_sociale") or "Destinatario Sconosciuto"
                    data[key] = {
                        "ragione_sociale": item.get("ragione_sociale") or item.get("Cliente", ""),
                        "indirizzo": item.get("indirizzo", ""),
                        "piva": item.get("piva", "")
                    }
            else:
                data = raw_data
        except Exception:
            data = {}
            
    if not data:
        # Dati di default se il file è vuoto o corrotto
        data = {
            "Sede Cliente": {
                "ragione_sociale": "Cliente S.p.A.", 
                "indirizzo": "Via Esempio 1, 16100 Genova", 
                "piva": "IT00000000000"
            }
        }
    return data

# ========================================================
#  RUBRICA EMAIL (contatti + gruppi)
#  File: rubrica_email.json (su disco persistente se presente)
# ========================================================

def _rubrica_email_path() -> Path:
    # Su Render MEDIA_DIR punta a /var/data/app (persistente) se esiste
    return (MEDIA_DIR / "rubrica_email.json") if 'MEDIA_DIR' in globals() else (APP_DIR / "rubrica_email.json")

def load_rubrica_email():
    fp = _rubrica_email_path()
    if fp.exists():
        try:
            data = json.loads(fp.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                # Normalizza struttura
                data.setdefault("contatti", {})
                data.setdefault("gruppi", {})
                return data
        except Exception:
            pass
    # default
    return {"contatti": {}, "gruppi": {}}

def save_rubrica_email(data: dict):
    fp = _rubrica_email_path()
    fp.parent.mkdir(parents=True, exist_ok=True)
    fp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def _parse_emails(raw: str):
    # accetta ; oppure , e ripulisce
    emails = []
    for e in (raw or "").replace(";", ",").split(","):
        e = e.strip()
        if e:
            emails.append(e)
    # de-dup preservando ordine
    seen = set()
    out = []
    for e in emails:
        if e.lower() in seen: 
            continue
        seen.add(e.lower())
        out.append(e)
    return out

def _ensure_progressivi_ddt_table(db):
    """Crea (se necessario) la tabella progressivi_ddt nel DB.
    Usiamo il DB invece di un file JSON così il progressivo resta memorizzato anche su server (Render/VPS).
    """
    try:
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS progressivi_ddt (
                anno VARCHAR(4) PRIMARY KEY,
                last_num INTEGER NOT NULL
            )
        """))
        db.commit()
    except Exception:
        # su alcuni DB/permessi, può fallire ma non blocchiamo l'app
        try:
            db.rollback()
        except Exception:
            pass

def peek_next_ddt_number():
    """Restituisce il prossimo progressivo SENZA incrementarlo (anteprima)."""
    y = str(date.today().year)[-2:]
    # 1) prova DB
    try:
        db = SessionLocal()
        try:
            _ensure_progressivi_ddt_table(db)
            row = db.execute(text("SELECT last_num FROM progressivi_ddt WHERE anno=:y"), {"y": y}).fetchone()
            last_num = int(row[0]) if row and row[0] is not None else 0
            n = last_num + 1
            return f"{n:02d}/{y}"
        finally:
            db.close()
    except Exception:
        pass

    # 2) fallback file (compatibilità)
    PROG_FILE = APP_DIR / "progressivi_ddt.json"
    prog = {}
    if PROG_FILE.exists():
        try:
            prog = json.loads(PROG_FILE.read_text(encoding="utf-8"))
        except Exception:
            prog = {}
    n = int(prog.get(y, 0)) + 1
    return f"{n:02d}/{y}"

def next_ddt_number():
    """Incrementa e memorizza il progressivo (solo in Finalizza)."""
    y = str(date.today().year)[-2:]

    # 1) DB (preferito)
    try:
        db = SessionLocal()
        try:
            _ensure_progressivi_ddt_table(db)

            dialect = getattr(engine.dialect, "name", "")
            if dialect and dialect.lower().startswith("mysql"):
                # lock riga anno per evitare doppioni
                row = db.execute(
                    text("SELECT last_num FROM progressivi_ddt WHERE anno=:y FOR UPDATE"),
                    {"y": y}
                ).fetchone()
            else:
                row = db.execute(
                    text("SELECT last_num FROM progressivi_ddt WHERE anno=:y"),
                    {"y": y}
                ).fetchone()

            last_num = int(row[0]) if row and row[0] is not None else 0
            n = last_num + 1

            if row:
                db.execute(
                    text("UPDATE progressivi_ddt SET last_num=:n WHERE anno=:y"),
                    {"n": n, "y": y}
                )
            else:
                db.execute(
                    text("INSERT INTO progressivi_ddt (anno, last_num) VALUES (:y, :n)"),
                    {"y": y, "n": n}
                )

            db.commit()
            return f"{n:02d}/{y}"
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()
    except Exception:
        # 2) fallback file (compatibilità)
        PROG_FILE = APP_DIR / "progressivi_ddt.json"
        prog = {}
        if PROG_FILE.exists():
            try:
                prog = json.loads(PROG_FILE.read_text(encoding="utf-8"))
            except Exception:
                prog = {}
        n = int(prog.get(y, 0)) + 1
        prog[y] = n
        try:
            PROG_FILE.write_text(json.dumps(prog, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass
        return f"{n:02d}/{y}"




def consume_specific_ddt_number(n_ddt: str) -> None:
    """
    Aggiorna il progressivo salvato (tabella progressivi_ddt o file json) per evitare
    che un numero scelto manualmente (con le frecce) venga riutilizzato in futuro.

    - NON cambia il valore del DDT passato.
    - Aggiorna last_num = max(last_num, numero_scelto) per l'anno del DDT.
    """
    if not n_ddt:
        return
    n_ddt = str(n_ddt).strip()

    m = re.match(r'^(\d+)\s*/\s*(\d{2})$', n_ddt)
    if not m:
        return

    try:
        chosen_num = int(m.group(1))
        chosen_year = m.group(2)
    except Exception:
        return

    if chosen_num < 1:
        return

    # 1) DB (preferito)
    try:
        db = SessionLocal()
        try:
            _ensure_progressivi_ddt_table(db)
            row = db.execute(text("SELECT last_num FROM progressivi_ddt WHERE anno=:y"), {"y": chosen_year}).fetchone()
            last_num = int(row[0]) if row and row[0] is not None else 0
            new_last = max(last_num, chosen_num)

            if row:
                db.execute(
                    text("UPDATE progressivi_ddt SET last_num=:n WHERE anno=:y"),
                    {"n": new_last, "y": chosen_year}
                )
            else:
                db.execute(
                    text("INSERT INTO progressivi_ddt (anno, last_num) VALUES (:y, :n)"),
                    {"y": chosen_year, "n": new_last}
                )
            db.commit()
            return
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()
    except Exception:
        pass

    # 2) fallback file (compatibilità)
    try:
        PROG_FILE = APP_DIR / "progressivi_ddt.json"
        prog = {}
        if PROG_FILE.exists():
            try:
                prog = json.loads(PROG_FILE.read_text(encoding="utf-8"))
            except Exception:
                prog = {}
        last_num = int(prog.get(chosen_year, 0) or 0)
        prog[chosen_year] = max(last_num, chosen_num)
        PROG_FILE.write_text(json.dumps(prog, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


# --- SEZIONE TEMPLATES HTML ---
BASE_HTML = """
<!doctype html>
<html lang="it">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{{ title or "Camar • Gestionale Web" }}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
    <style>
        body { background: #f8f9fa; font-size: 14px; }
        .navbar { background-color: #1f6fb2; } /* Blu Camar */
        .navbar-brand, .nav-link, .navbar-text { color: white !important; }
        .nav-link:hover { opacity: 0.8; }
        
        .card { border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,.08); border: none; }
        .table-container { overflow: auto; max-height: 65vh; }
        .table thead th { position: sticky; top: 0; background: #f0f2f5; z-index: 2; }
        .dropzone { border: 2px dashed #0d6efd; background: #eef4ff; padding: 20px; border-radius: 12px; text-align: center; color: #0d6efd; cursor: pointer; }
        .logo { height: 32px; width: auto; }
        .table-compact th, .table-compact td { font-size: 11px; padding: 4px 5px; white-space: normal; word-wrap: break-word; vertical-align: middle; }
        .table-striped tbody tr:nth-of-type(odd) { background-color: rgba(0,0,0,.03); }
        
        /* Stile Bottoni Admin nel Menu */
        .btn-nav-admin { font-weight: bold; border-radius: 6px; box-shadow: 0 2px 4px rgba(0,0,0,0.2); }
        
        @media print { .no-print { display: none !important; } }
    </style>
</head>
<body>
<nav class="navbar navbar-expand-lg navbar-dark shadow-sm no-print">
    <div class="container-fluid">
        <a class="navbar-brand d-flex align-items-center gap-2" href="{{ url_for('home') }}">
            {% if logo_url %}<img src="{{ logo_url }}" class="logo" alt="logo">{% endif %}
            Camar • Gestionale
        </a>

        <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
            <span class="navbar-toggler-icon"></span>
        </button>

        <div class="collapse navbar-collapse" id="navbarNav">
            <ul class="navbar-nav ms-auto align-items-center gap-2">
                
                <li class="nav-item"><a class="nav-link" href="{{ url_for('giacenze') }}">📦 Magazzino</a></li>
                {% if session.get('role') == 'admin' %}
                <li class="nav-item"><a class="nav-link" href="{{ url_for('import_excel') }}">📥 Import Excel</a></li>
                {% endif %}

                {% if session.get('role') == 'admin' %}
                    <li class="nav-item border-start border-light ps-2 ms-2 d-none d-lg-block"></li> <li class="nav-item">
                        <a class="nav-link btn btn-danger text-white px-3 ms-2 btn-nav-admin" href="{{ url_for('trasporti') }}">
                            <i class="bi bi-truck"></i> TRASPORTI
                        </a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link btn btn-success text-white px-3 ms-2 btn-nav-admin" href="{{ url_for('lavorazioni') }}">
                            <i class="bi bi-box-seam"></i> PICKING
                        </a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link btn btn-outline-light text-white px-3 ms-2 btn-nav-admin" href="{{ url_for('rubrica_email') }}">
                            <i class="bi bi-journal-bookmark"></i> RUBRICA EMAIL
                        </a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link btn btn-outline-warning text-dark px-3 ms-2 btn-nav-admin" href="{{ url_for('backup_download') }}">
                            <i class="bi bi-download"></i> BACKUP
                        </a>
                    </li>

                    <li class="nav-item">
                         <a class="nav-link text-white-50 ms-1" href="{{ url_for('manage_mappe') }}" title="Gestione Mappe"><i class="bi bi-gear"></i></a>
                    </li>
                {% endif %}

                {% if session.get('user') %}
                    <li class="nav-item ms-4 text-white-50 small d-none d-lg-block">Utente: <b>{{ session['user'] }}</b></li>
                    <li class="nav-item">
                        <a class="btn btn-outline-light btn-sm ms-2" href="{{ url_for('logout') }}"><i class="bi bi-box-arrow-right"></i> Esci</a>
                    </li>
                {% endif %}
            </ul>
        </div>
    </div>
</nav>

<main class="container-fluid my-4">
    {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
            {% for category, message in messages %}
                <div class="alert alert-{{ category }} alert-dismissible fade show no-print" role="alert">
                    {{ message|safe }}
                    <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Close"></button>
                </div>
            {% endfor %}
        {% endif %}
    {% endwith %}

    {% block content %}{% endblock %}
</main>

<footer class="text-center text-white py-3 small no-print" style="background-color: #1f6fb2; margin-top: auto;">
    © Alessia Moncalvo – Gestionale Camar Web Edition • Tutti i diritti riservati.
</footer>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
{% block extra_js %}{% endblock %}
</body>
</html>
"""
REPORT_INVENTARIO_HTML = """
<!DOCTYPE html>
<html>
<head>
  <title>Inventario Totale - {{ data_rif }}</title>
  <meta charset="utf-8">
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    @media print {
      .pagebreak { page-break-after: always; }
    }
    table { font-size: 12px; }
    h1 { font-size: 26px; margin-bottom: 10px; }
    h3 { font-size: 18px; }
    .thead-custom th { background: #d9e1f2 !important; }
  </style>
</head>

<body onload="window.print()">
  <div class="container mt-4">
    <h1 class="text-center">Inventario Totale</h1>
    <div class="text-center text-muted" style="margin-top:-6px;">
      Generato il {{ data_rif }}
    </div>
    <hr>

    {% if not inventario or inventario|length == 0 %}
      <div class="alert alert-warning">
        Nessun articolo trovato.
      </div>
    {% endif %}

    {% for cliente, righe in inventario.items() %}
      <h3 class="mt-4 bg-light p-2">{{ cliente }}</h3>

      <table class="table table-sm table-bordered">
        <thead class="thead-custom">
          <tr>
            <th style="width:60px;">ID</th>
            <th style="width:220px;">CODICE ARTICOLO</th>
            <th>DESCRIZIONE</th>
            <th class="text-center" style="width:130px;">Q.TA ENTRATA</th>
            <th class="text-center" style="width:130px;">Q.TA USCITA</th>
            <th class="text-center" style="width:130px;">RIMANENZA</th>
          </tr>
        </thead>
        <tbody>
          {% for r in righe %}
          <tr>
            <td class="text-center">{{ r.idx }}</td>
            <td>{{ r.codice }}</td>
            <td>{{ r.descrizione }}</td>
            <td class="text-center">{{ r.entrata }}</td>
            <td class="text-center">{{ r.uscita }}</td>
            <td class="text-center">{{ r.rimanenza }}</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>

      <div class="pagebreak"></div>
    {% endfor %}
  </div>
</body>
</html>
"""


REPORT_TRASPORTI_HTML = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Report Trasporti</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { font-size: 12px; }
        h1 { font-size: 22px; }
        .table th { background-color: #f0f0f0; }
        .wrap { white-space: normal; }
    </style>
</head>
<body onload="window.print()">
    <div class="container mt-4">
        <h1 class="mb-3">Report Trasporti</h1>

        <div class="mb-3 p-2 bg-light border">
            <strong>Filtri applicati:</strong><br>
            Periodo: {{ mese }} |
            Cliente: {{ cliente }} |
            Mezzo: {{ mezzo }} |
            DDT: {{ ddt_uscita }} |
            Consolidato: {{ consolidato }}
        </div>

        <table class="table table-bordered table-sm align-middle">
            <thead>
                <tr>
                    <th>Data</th>
                    <th>Mezzo</th>
                    <th>Cliente</th>
                    <th>Trasportatore</th>
                    <th>DDT</th>
                    <th>Mag.</th>
                    <th>Consolidato</th>
                    <th class="text-end">Costo</th>
                </tr>
            </thead>
            <tbody>
                {% for t in dati %}
                <tr>
                    <td>{{ t.data or '' }}</td>
                    <td>{{ t.tipo_mezzo or '' }}</td>
                    <td class="wrap">{{ t.cliente or '' }}</td>
                    <td class="wrap">{{ t.trasportatore or '' }}</td>
                    <td>{{ t.ddt_uscita or '' }}</td>
                    <td>{{ t.magazzino or '' }}</td>
                    <td>{{ t.consolidato or '' }}</td>
                    <td class="text-end">€ {{ '%.2f'|format(t.costo) if t.costo else '0.00' }}</td>
                </tr>
                {% else %}
                <tr>
                    <td colspan="8" class="text-center">Nessun dato trovato per i criteri selezionati.</td>
                </tr>
                {% endfor %}
            </tbody>
            <tfoot>
                <tr class="table-dark fw-bold">
                    <td colspan="7" class="text-end">TOTALE COMPLESSIVO</td>
                    <td class="text-end">€ {{ totale }}</td>
                </tr>
            </tfoot>
        </table>

        <div class="text-muted small mt-4">
            Generato il: <script>document.write(new Date().toLocaleDateString())</script>
        </div>
    </div>
</body>
</html>
"""

LOGIN_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="row justify-content-center mt-5">
    <div class="col-md-5 col-lg-4">
        <div class="card p-4 text-center">
            {% if logo_url %}<img src="{{ logo_url }}" class="mb-3 mx-auto" style="height:56px; width: auto;">{% endif %}
            <h4 class="mb-3">Login al gestionale</h4>
            <form method="post" class="text-start">
                <div class="mb-3">
                    <label class="form-label">Utente</label>
                    <input name="user" class="form-control" required>
                </div>
                <div class="mb-3">
                    <label class="form-label">Password</label>
                    <input type="password" name="pwd" class="form-control" required>
                </div>
                <button class="btn btn-primary w-100">Accedi</button>
            </form>
        </div>
    </div>
</div>
{% endblock %}
"""

HOME_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="row g-3">
    <div class="col-lg-3">
        <div class="card p-3">
            <h6 class="mb-3">Menu Principale</h6>
            <div class="d-grid gap-2">
                <a class="btn btn-primary" href="{{ url_for('giacenze') }}"><i class="bi bi-grid-3x3-gap-fill"></i> Visualizza Giacenze</a>
                {% if session.get('role') == 'admin' %}
                <a class="btn btn-success" href="{{ url_for('nuovo_articolo') }}"><i class="bi bi-plus-circle"></i> Nuovo Articolo</a>
                {% endif %}
                {% if session.get('role') == 'admin' %}
                <a class="btn btn-outline-secondary" href="{{ url_for('labels_form') }}"><i class="bi bi-tag"></i> Stampa Etichette</a>
                {% endif %}
                <hr>
                {% if session.get('role') == 'admin' %}
                <a class="btn btn-outline-secondary btn-sm" href="{{ url_for('import_excel') }}"><i class="bi bi-file-earmark-arrow-up"></i> Import Excel</a>
                {% endif %}
                {% if session.get('role') == 'admin' %}
                <a class="btn btn-outline-secondary btn-sm" href="{{ url_for('export_excel') }}"><i class="bi bi-file-earmark-arrow-down"></i> Export Excel Totale</a>
                {% endif %}
                <a class="btn btn-outline-secondary btn-sm" href="{{ url_for('export_client') }}"><i class="bi bi-people"></i> Export per Cliente</a>
                <a class="btn btn-outline-secondary btn-sm" href="{{ url_for('calcola_costi') }}"><i class="bi bi-calculator"></i> Calcola Giacenze Mensili</a>
            </div>
        </div>
    </div>
    <div class="col-lg-9">
        <div class="card p-4">
            <div class="d-flex align-items-center gap-3">
                {% if logo_url %}<img src="{{ logo_url }}" style="height:48px">{% endif %}
                <div>
                    <h4 class="m-0">Benvenuto nel Gestionale Camar</h4>
                    <p class="text-muted m-0">Gestione completa di giacenze, DDT, buoni di prelievo e stampa PDF.</p>
                </div>
            </div>
        </div>
    </div>
</div>
{% endblock %}
"""
IMPORT_PDF_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="card p-4">
    <h3><i class="bi bi-file-earmark-pdf"></i> Importa Articoli da PDF (DDT/Bolla)</h3>
    
    {% if not rows %}
    <div class="alert alert-info">
        Carica un DDT in formato PDF digitale. Il sistema tenterà di leggere codici e quantità.<br>
        <b>Nota:</b> Funziona meglio con PDF generati da computer, non scansioni.
    </div>
    <form method="post" enctype="multipart/form-data" class="mt-4">
        <div class="mb-3">
            <label class="form-label">Seleziona File PDF</label>
            <input type="file" name="file" class="form-control" accept=".pdf" required>
        </div>
        <button type="submit" class="btn btn-primary">Analizza PDF</button>
        <a href="{{ url_for('giacenze') }}" class="btn btn-secondary">Annulla</a>
    </form>
    {% endif %}

    {% if rows %}
    <form action="{{ url_for('save_pdf_import') }}" method="post">
        <div class="row g-3 mb-3 bg-light p-3 rounded border">
            <h5 class="mb-3">Dati Testata (Rilevati o da compilare)</h5>
            <div class="col-md-3">
                <label>Cliente</label>
                <input name="cliente" class="form-control" value="{{ meta.cliente or '' }}">
            </div>
            <div class="col-md-3">
                <label>Fornitore</label>
                <input name="fornitore" class="form-control" value="{{ meta.fornitore or '' }}">
            </div>
            <div class="col-md-2">
                <label>Commessa</label>
                <input name="commessa" class="form-control" value="{{ meta.commessa or '' }}">
            </div>
            <div class="col-md-2">
                <label>N. DDT</label>
                <input name="n_ddt" class="form-control" value="{{ meta.n_ddt or '' }}">
            </div>
            <div class="col-md-2">
                <label>Data Ingresso</label>
                <input type="date" name="data_ingresso" class="form-control" value="{{ meta.data_ingresso or '' }}">
            </div>
        </div>

        <div class="table-responsive">
            <table class="table table-striped table-sm align-middle">
                <thead class="table-dark">
                    <tr>
                        <th style="width:70px">Rimuovi</th>
                        <th>Codice Articolo</th>
                        <th>Descrizione</th>
                        <th style="width:120px">Colli</th>
                        <th style="width:160px">Pezzi / Q.tà</th>
                    </tr>
                </thead>
                <tbody id="rowsBody">
                    {% for r in rows %}
                    <tr>
                        <td class="text-center">
                            <button type="button" class="btn btn-danger btn-sm py-0" onclick="this.closest('tr').remove()">X</button>
                        </td>
                        <td>
                            <input name="codice[]" class="form-control form-control-sm" value="{{ r.codice or '' }}">
                        </td>
                        <td>
                            <input name="descrizione[]" class="form-control form-control-sm" value="{{ r.descrizione or '' }}">
                        </td>
                        <td>
                            <input name="colli[]" type="number" min="0" class="form-control form-control-sm" value="{{ r.colli or r.qta or 1 }}">
                        </td>
                        <td>
                            <input name="pezzi[]" type="number" step="0.01" class="form-control form-control-sm" value="{{ r.pezzi or 1 }}">
                        </td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        
        <div class="d-flex justify-content-between mt-3">
            <button type="button" class="btn btn-secondary btn-sm" onclick="addRow()">+ Aggiungi Riga Vuota</button>
            <div>
                <a href="{{ url_for('import_pdf') }}" class="btn btn-outline-secondary">Ricomincia</a>
                <button type="submit" class="btn btn-success fw-bold px-4">CONFERMA E IMPORTA</button>
            </div>
        </div>
    </form>

    <script>
    function addRow() {
        const tbody = document.getElementById('rowsBody');
        const tr = document.createElement('tr');
        tr.innerHTML = `
            <td class="text-center">
                <button type="button" class="btn btn-danger btn-sm py-0" onclick="this.closest('tr').remove()">X</button>
            </td>
            <td><input name="codice[]" class="form-control form-control-sm"></td>
            <td><input name="descrizione[]" class="form-control form-control-sm"></td>
            <td><input name="colli[]" type="number" min="0" class="form-control form-control-sm" value="1"></td>
            <td><input name="pezzi[]" type="number" step="0.01" class="form-control form-control-sm" value="1"></td>
        `;
        tbody.appendChild(tr);
    }
    </script>
    {% endif %}
</div>
{% endblock %}
"""


    
CALCOLI_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="container-fluid">
    <h3><i class="bi bi-calculator"></i> Report Costi Magazzino (M² per cliente)</h3>
    
    <div class="card mb-4" style="background-color: #f0f0f0; border: 1px solid #ccc;">
        <div class="card-body py-3">
            <form method="POST">
                <div class="row g-3 align-items-center">
                    
                    <div class="col-auto">
                        <label class="fw-bold">Data da:</label>
                    </div>
                    <div class="col-auto">
                        <input type="date" name="data_da" class="form-control form-control-sm" required value="{{ data_da }}">
                    </div>
                    
                    <div class="col-auto">
                        <label class="fw-bold">Data a:</label>
                    </div>
                    <div class="col-auto">
                        <input type="date" name="data_a" class="form-control form-control-sm" required value="{{ data_a }}">
                    </div>

                    {% if is_admin %}
                    <div class="col-auto ms-4">
                        <label class="fw-bold">Cliente (contiene):</label>
                    </div>
                    <div class="col-auto">
                        <input type="text" name="cliente" class="form-control form-control-sm" value="{{ cliente_filtro }}">
                    </div>
                    {% else %}
                    <div class="col-auto ms-4">
                        <label class="fw-bold">Cliente:</label>
                    </div>
                    <div class="col-auto">
                        <input type="text" class="form-control form-control-sm" value="{{ cliente_filtro }}" readonly>
                    </div>
                    {% endif %}

                    <div class="col-auto ms-4">
                        <div class="form-check form-check-inline">
                            <input class="form-check-input" type="radio" name="raggruppamento" value="mese" id="rmese" {% if raggruppamento=='mese' %}checked{% endif %}>
                            <label class="form-check-label" for="rmese">Per mese</label>
                        </div>
                        <div class="form-check form-check-inline">
                            <input class="form-check-input" type="radio" name="raggruppamento" value="giorno" id="rgiorno" {% if raggruppamento=='giorno' %}checked{% endif %}>
                            <label class="form-check-label" for="rgiorno">Per giorno</label>
                        </div>
                    </div>

                    
                    <input type="hidden" name="area_manovra" id="area_manovra" value="{{ '1' if (is_admin and area_manovra) else '0' }}">
                    <div class="col-auto ms-auto d-flex gap-2">
                        <button type="button" class="btn btn-secondary border-dark px-4" style="background-color: #e0e0e0; color: black;" onclick="document.getElementById('area_manovra').value='0'; this.closest('form').submit();">Calcola</button>
                        {% if is_admin %}
                        <button type="button" class="btn btn-warning border-dark px-4" onclick="document.getElementById('area_manovra').value='1'; this.closest('form').submit();">
                            Area Manovra {% if area_manovra %}<i class="bi bi-check2-circle"></i>{% endif %}
                        </button>
                        {% endif %}
                        <button type="submit" name="export_excel" value="1" class="btn btn-success border-dark px-4">
                            <i class="bi bi-file-earmark-excel"></i> Excel
                        </button>
                    </div>

                </div>
            </form>
        </div>
    </div>

    {% if risultati %}
    <div class="table-responsive bg-white border">
        <table class="table table-bordered table-striped table-hover table-sm mb-0">
            <thead class="table-light">
                <tr class="text-center align-middle">
                    <th>Mese/Giorno</th>
                    <th>Cliente</th>
                    <th>M² * giorni</th>
                    <th>M² effettivi</th>
                    <th>Giorni</th>
                </tr>
            </thead>
            <tbody>
                {% for row in risultati %}
                <tr>
                    <td class="text-center">{{ row.periodo }}</td>
                    <td>{{ row.cliente }}</td>
                    <td class="text-end">{{ row.m2_tot }}</td>
                    <td class="text-end fw-bold">{{ row.m2_medio }}</td>
                    <td class="text-center">{{ row.giorni }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    {% elif request.method == 'POST' %}
    <div class="alert alert-warning">Nessun dato trovato per i criteri selezionati.</div>
    {% endif %}
</div>
{% endblock %}
"""

GIACENZE_HTML = """
{% extends 'base.html' %}
{% block content %}
<style>
    .table-compact td, .table-compact th { font-size: 0.78rem; padding: 4px 6px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 160px; vertical-align: middle; }
    .table-compact th { background-color: #f0f0f0; font-weight: 700; text-align: center; position: sticky; top: 0; z-index: 10; }
    .fw-buono { font-weight: bold; color: #000; }
    .att-link { text-decoration: none; font-size: 1.3em; cursor: pointer; margin: 0 3px; }
    .att-link:hover { transform: scale(1.2); display:inline-block; }
    /* Paginazione */
    .pagination { margin-bottom: 0; }
    .page-link { color: #333; text-decoration: none; padding: 0.3rem 0.6rem; }
    .page-item.active .page-link { background-color: #0d6efd; border-color: #0d6efd; color: white; }
    .page-item.disabled .page-link { color: #ccc; pointer-events: none; }
</style>

<div class="d-flex justify-content-between align-items-center mb-2">
    <h4 class="mb-0"><i class="bi bi-box-seam"></i> Magazzino <small class="text-muted fs-6">({{ total_items }} articoli)</small></h4>
    <div class="d-flex gap-2 flex-wrap">
        {% if session.get('role') == 'admin' %}
        <a href="{{ url_for('nuovo_articolo') }}" class="btn btn-sm btn-success"><i class="bi bi-plus-lg"></i> Nuovo</a>
        <a href="{{ url_for('import_pdf') }}" class="btn btn-sm btn-dark"><i class="bi bi-file-earmark-pdf"></i> Import PDF</a>
        <form action="{{ url_for('labels_pdf') }}" method="POST" target="_blank" class="d-inline">
            <button class="btn btn-sm btn-info text-white"><i class="bi bi-tag"></i> Etichette</button>
        </form>
        {% endif %}
        <a href="{{ url_for('calcola_costi') }}" class="btn btn-sm btn-warning"><i class="bi bi-calculator"></i> Calcoli</a>

        <form action="{{ url_for('report_inventario_excel') }}" method="POST" class="d-inline-block">
            <div class="input-group input-group-sm">
                <input type="date" name="data_inventario" class="form-control" required value="{{ today }}">
                {% if session.get('role') == 'admin' %}
                    <input type="text" name="cliente_inventario" class="form-control" placeholder="Cliente (es. FINCANTIERI)" style="max-width: 200px;">
                {% else %}
                    <input type="hidden" name="cliente_inventario" value="{{ session.get('user') }}">
                {% endif %}
                <button class="btn btn-success" type="submit" title="Scarica Excel">📥 Excel</button>
            </div>
        </form>
    </div>
</div>

<div class="card mb-2 bg-light shadow-sm">
    <div class="card-header py-1" data-bs-toggle="collapse" data-bs-target="#filterBody" style="cursor:pointer">
        <small><i class="bi bi-funnel"></i> <b>Filtri Avanzati</b></small>
    </div>
    <div id="filterBody" class="collapse {% if request.args %}show{% endif %}">
        <div class="card-body py-2">
            <form method="get">
                <div class="row g-1 mb-1">
                    <div class="col-md-1"><input name="id" class="form-control form-control-sm" placeholder="ID" value="{{ request.args.get('id','') }}"></div>
                    <div class="col-md-2"><input name="cliente" class="form-control form-control-sm" placeholder="Cliente" value="{{ request.args.get('cliente','') }}"></div>
                    <div class="col-md-2"><input name="fornitore" class="form-control form-control-sm" placeholder="Fornitore" value="{{ request.args.get('fornitore','') }}"></div>
                    <div class="col-md-2"><input name="codice_articolo" class="form-control form-control-sm" placeholder="Codice" value="{{ request.args.get('codice_articolo','') }}"></div>
                    <div class="col-md-2"><input name="serial_number" class="form-control form-control-sm" placeholder="Serial" value="{{ request.args.get('serial_number','') }}"></div>
                    <div class="col-md-2"><input name="ordine" class="form-control form-control-sm" placeholder="Ordine" value="{{ request.args.get('ordine','') }}"></div>

                    <!-- ✅ NUOVO FILTRO: SOLO IN GIACENZA -->
                    <div class="col-md-2 d-flex align-items-center">
                        <div class="form-check">
                            <input class="form-check-input" type="checkbox" value="1" id="solo_giacenza" name="solo_giacenza"
                                   {% if request.args.get('solo_giacenza') == '1' %}checked{% endif %}>
                            <label class="form-check-label" for="solo_giacenza">
                                Solo in giacenza
                            </label>
                        </div>
                    </div>

                    <div class="col-md-1"><button type="submit" class="btn btn-primary btn-sm w-100">Cerca</button></div>
                </div>

                <div class="row g-1 mb-1">
                    <div class="col-md-2"><input name="lotto" class="form-control form-control-sm" placeholder="Lotto" value="{{ request.args.get('lotto','') }}"></div>
                    <div class="col-md-2"><input name="commessa" class="form-control form-control-sm" placeholder="Commessa" value="{{ request.args.get('commessa','') }}"></div>
                    <div class="col-md-2"><input name="protocollo" class="form-control form-control-sm" placeholder="Protocollo" value="{{ request.args.get('protocollo','') }}"></div>
                    <div class="col-md-2"><input name="magazzino" class="form-control form-control-sm" placeholder="Magazzino" value="{{ request.args.get('magazzino','') }}"></div>
                    <div class="col-md-2"><input name="descrizione" class="form-control form-control-sm" placeholder="Descrizione" value="{{ request.args.get('descrizione','') }}"></div>
                    <div class="col-md-2"><input name="buono_n" class="form-control form-control-sm" placeholder="N. Buono" value="{{ request.args.get('buono_n','') }}"></div>
                    <div class="col-md-2"><input name="n_arrivo" class="form-control form-control-sm" placeholder="N. Arrivo" value="{{ request.args.get('n_arrivo','') }}"></div>
                    <div class="col-md-2"><input name="mezzi_in_uscita" class="form-control form-control-sm" placeholder="Mezzo Uscita" value="{{ request.args.get('mezzi_in_uscita','') }}"></div>
                    <div class="col-md-2"><input name="stato" class="form-control form-control-sm" placeholder="Stato" value="{{ request.args.get('stato','') }}"></div>
                </div>

                <div class="row g-1">
                    <div class="col-md-2"><input name="n_ddt_ingresso" class="form-control form-control-sm" placeholder="DDT Ing." value="{{ request.args.get('n_ddt_ingresso','') }}"></div>
                    <div class="col-md-2"><input name="n_ddt_uscita" class="form-control form-control-sm" placeholder="DDT Usc." value="{{ request.args.get('n_ddt_uscita','') }}"></div>
                    <div class="col-md-4">
                        <div class="input-group input-group-sm">
                            <span class="input-group-text">Ingr</span>
                            <input name="data_ing_da" type="date" class="form-control" value="{{ request.args.get('data_ing_da','') }}">
                            <span class="input-group-text">-</span>
                            <input name="data_ing_a" type="date" class="form-control" value="{{ request.args.get('data_ing_a','') }}">
                        </div>
                    </div>
                    <div class="col-md-4 d-flex gap-1">
                        <div class="input-group input-group-sm">
                            <span class="input-group-text">Usc</span>
                            <input name="data_usc_da" type="date" class="form-control" value="{{ request.args.get('data_usc_da','') }}">
                            <span class="input-group-text">-</span>
                            <input name="data_usc_a" type="date" class="form-control" value="{{ request.args.get('data_usc_a','') }}">
                        </div>
                        <a href="{{ url_for('giacenze') }}" class="btn btn-outline-secondary btn-sm" onclick="localStorage.removeItem('camar_selected_articles');">Reset</a>
                    </div>
                </div>
            </form>
        </div>
    </div>
</div>

<form method="POST">
    <div class="btn-toolbar mb-2 gap-1 flex-wrap">
        {% if session.get('role') == 'admin' %}
        <button type="submit" formaction="{{ url_for('buono_preview') }}" class="btn btn-outline-dark btn-sm">Buono</button>
        <button type="submit" formaction="{{ url_for('ddt_preview') }}" class="btn btn-outline-dark btn-sm">DDT</button>
        <button type="submit" formaction="{{ url_for('invia_email') }}" formmethod="get" class="btn btn-success btn-sm"><i class="bi bi-envelope"></i> Email</button>
        <button type="submit" formaction="{{ url_for('bulk_edit') }}" class="btn btn-info btn-sm text-white">Modifica</button>
        <button type="submit" formaction="{{ url_for('labels_pdf') }}" formtarget="_blank" class="btn btn-warning btn-sm"><i class="bi bi-download"></i> Etichette</button>
        <button type="submit" formaction="{{ url_for('delete_rows') }}" class="btn btn-danger btn-sm" onclick="return confirm('Eliminare SELEZIONATI?')">Elimina</button>
        <button type="submit" formaction="{{ url_for('bulk_duplicate') }}" class="btn btn-primary btn-sm" onclick="return confirm('Duplicare?')">Duplica</button>
        {% endif %}
    </div>

    <div class="table-responsive shadow-sm" style="max-height: 65vh;">
        <table class="table table-striped table-bordered table-hover table-compact mb-0">
            <thead class="sticky-top" style="top:0; z-index:5;">
                <tr>
                    <th><input type="checkbox" onclick="toggleAll(this)"></th>
                    <th>ID</th> <th>Codice</th> <th>Pz</th> <th>Larg</th> <th>Lung</th> <th>Alt</th> <th>M2</th> <th>M3</th>
                    <th>Descrizione</th> <th>Protocollo</th> <th>Commessa</th> <th>Ordine</th> <th>Colli</th> <th>Fornitore</th> <th>Magazzino</th>
                    <th>Data Ing</th> <th>DDT Ing</th> <th>DDT Usc</th> <th>Data Usc</th> <th>Mezzo Usc</th>
                    <th>Cliente</th> <th>Kg</th> <th>Posiz</th> <th>N.Arr</th> <th>N.Buono</th> <th>Note</th> 
                    <th>Lotto</th> <th>Ns.Rif</th> <th>Serial</th> <th>Stato</th>
                    <th>Doc</th> <th>Foto</th> <th>Act</th>
                </tr>
            </thead>
            <tbody>
                {% for r in rows %}
                <tr>
                    <td class="text-center"><input type="checkbox" name="ids" value="{{ r.id_articolo }}" class="row-checkbox"></td>
                    <td>{{ r.id_articolo }}</td>
                    <td title="{{ r.codice_articolo }}">{{ r.codice_articolo or '' }}</td>
                    <td>{{ r.pezzo or '' }}</td>
                    <td>{{ r.larghezza|it_num(2) if r.larghezza else '' }}</td>
                    <td>{{ r.lunghezza|it_num(2) if r.lunghezza else '' }}</td>
                    <td>{{ r.altezza|it_num(2) if r.altezza else '' }}</td>
                    <td>{{ r.m2|it_num(3) if r.m2 else '' }}</td>
                    <td>{{ r.m3|it_num(3) if r.m3 else '' }}</td>
                    <td title="{{ r.descrizione }}">{{ (r.descrizione or '')[:25] }}...</td>
                    <td>{{ r.protocollo or '' }}</td> <td>{{ r.commessa or '' }}</td> <td>{{ r.ordine or '' }}</td>
                    <td>{{ r.n_colli or '' }}</td> <td>{{ r.fornitore or '' }}</td> <td>{{ r.magazzino or '' }}</td>
                    <td>{{ r.data_ingresso or '' }}</td>
                    <td>{{ r.n_ddt_ingresso or '' }}</td> <td>{{ r.n_ddt_uscita or '' }}</td>
                    <td>{{ r.data_uscita or '' }}</td>
                    <td>{{ r.mezzi_in_uscita or '' }}</td> <td>{{ r.cliente or '' }}</td> <td>{{ r.peso|it_num(2) if r.peso else '' }}</td>
                    <td>{{ r.posizione or '' }}</td> <td>{{ r.n_arrivo or '' }}</td> <td class="fw-buono">{{ r.buono_n or '' }}</td>
                    <td title="{{ r.note }}">{{ (r.note or '')[:15] }}...</td>
                    <td>{{ r.lotto or '' }}</td> <td>{{ r.ns_rif or '' }}</td> <td>{{ r.serial_number or '' }}</td>
                    <td>{{ r.stato or '' }}</td>
                    <td class="text-center">
                        {% for a in r.attachments if a.kind=='doc' %}
                        <a href="{{ url_for('serve_uploaded_file', filename=a.filename) }}" target="_blank" class="att-link">📄</a>
                        {% endfor %}
                    </td>
                    <td class="text-center">
                        {% for a in r.attachments if a.kind=='photo' %}
                        <a href="{{ url_for('serve_uploaded_file', filename=a.filename) }}" target="_blank" class="att-link">📷</a>
                        {% endfor %}
                    </td>
                    <td class="text-center">
                        {% if session.get('role') == 'admin' %}
                        <a href="{{ url_for('edit_articolo', id=r.id_articolo) }}" class="text-decoration-none">✏️</a>
                        <a href="{{ url_for('delete_articolo', id=r.id_articolo) }}" class="text-decoration-none text-danger" onclick="return confirm('Eliminare?')">🗑️</a>
                        {% else %}-{% endif %}
                    </td>
                </tr>
                {% else %}
                <tr><td colspan="34" class="text-center p-3 text-muted">Nessun articolo trovato.</td></tr>
                {% endfor %}
            </tbody>
        </table>
    </div>

    {% if total_pages > 1 %}
    <nav class="mt-2 bg-white p-2 border-top d-flex justify-content-between align-items-center shadow-sm">
        <div>
            <span class="fw-bold">Pagina {{ page }} di {{ total_pages }}</span> 
            <small class="text-muted">({{ total_items }} articoli totali)</small>
        </div>
        <ul class="pagination pagination-sm m-0">
            <li class="page-item {% if page == 1 %}disabled{% endif %}">
                <a class="page-link" href="{{ url_for('giacenze', page=page-1, **search_params) }}">
                    <i class="bi bi-chevron-left"></i> Precedente
                </a>
            </li>
            
            <li class="page-item {% if page == total_pages %}disabled{% endif %}">
                <a class="page-link" href="{{ url_for('giacenze', page=page+1, **search_params) }}">
                    Successivo <i class="bi bi-chevron-right"></i>
                </a>
            </li>
        </ul>
    </nav>
    {% endif %}

    <div class="text-end mt-2 text-muted small bg-light p-2 border-top fw-bold">
        Totali: Colli {{ total_colli }} | M2 {{ total_m2 }} | Peso {{ total_peso }}
    </div>
</form>

<script>
    const STORAGE_KEY = 'camar_selected_articles';
    
    function saveSelection() {
        const checked = Array.from(document.querySelectorAll('input[name="ids"]:checked')).map(cb => cb.value);
        let saved = JSON.parse(localStorage.getItem(STORAGE_KEY) || "[]");
        checked.forEach(id => { if(!saved.includes(id)) saved.push(id); });
        const unchecked = Array.from(document.querySelectorAll('input[name="ids"]:not(:checked)')).map(cb => cb.value);
        saved = saved.filter(id => !unchecked.includes(id));
        localStorage.setItem(STORAGE_KEY, JSON.stringify(saved));
    }

    function toggleAll(source) {
        document.querySelectorAll('input[name="ids"]').forEach(c => c.checked = source.checked);
        saveSelection();
    }

    document.addEventListener("DOMContentLoaded", function() {
        const savedIds = JSON.parse(localStorage.getItem(STORAGE_KEY) || "[]");
        document.querySelectorAll('input[name="ids"]').forEach(cb => {
            if (savedIds.includes(cb.value)) cb.checked = true;
        });
        document.addEventListener("change", function(e) {
            if (e.target && e.target.matches('input[name="ids"]')) saveSelection();
        });
    });
</script>
{% endblock %}
"""

EDIT_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="d-flex justify-content-between align-items-center mb-4">
    <h3>
        <i class="bi bi-pencil-square"></i> 
        {% if row.id_articolo %}Modifica Articolo #{{ row.id_articolo }}{% else %}Nuovo Articolo{% endif %}
    </h3>
    <a href="{{ url_for('giacenze') }}" class="btn btn-secondary">Torna alla Lista</a>
</div>

<form method="post" enctype="multipart/form-data" class="card p-4 shadow-sm mb-4">
    <div class="row g-3">
        <div class="col-md-3">
            <label class="form-label fw-bold">Codice Articolo</label>
            <input type="text" name="codice_articolo" class="form-control" value="{{ row.codice_articolo or '' }}">
        </div>
        <div class="col-md-5">
            <label class="form-label fw-bold">Descrizione</label>
            <input type="text" name="descrizione" class="form-control" value="{{ row.descrizione or '' }}">
        </div>
        <div class="col-md-2">
            <label class="form-label">Stato</label>
            <input class="form-control" list="statoList" name="stato" value="{{ row.stato or '' }}" placeholder="Seleziona...">
            <datalist id="statoList">
                <option value="NAZIONALE">
                <option value="DOGANALE">
                <option value="ESTERO">
                <option value="USCITO">
                <option value="FINCANTIERI SCOPERTO">
                <option value="AGGIUNTO A MANO">
            </datalist>
        </div>
        <div class="col-md-2"><label class="form-label">Commessa</label><input type="text" name="commessa" class="form-control" value="{{ row.commessa or '' }}"></div>

        <div class="col-md-4"><label class="form-label">Cliente</label><input type="text" name="cliente" class="form-control" value="{{ row.cliente or '' }}"></div>
        <div class="col-md-4"><label class="form-label">Fornitore</label><input type="text" name="fornitore" class="form-control" value="{{ row.fornitore or '' }}"></div>
        <div class="col-md-4"><label class="form-label">Protocollo</label><input type="text" name="protocollo" class="form-control" value="{{ row.protocollo or '' }}"></div>
        
        <div class="col-md-3"><label class="form-label">N. Buono</label><input type="text" name="buono_n" class="form-control" value="{{ row.buono_n or '' }}"></div>
        <div class="col-md-3"><label class="form-label">Magazzino</label><input type="text" name="magazzino" class="form-control" value="{{ row.magazzino or 'STRUPPA' }}"></div>
        <div class="col-md-3"><label class="form-label">Posizione</label><input type="text" name="posizione" class="form-control" value="{{ row.posizione or '' }}"></div>
        <div class="col-md-3"><label class="form-label">Ordine</label><input type="text" name="ordine" class="form-control" value="{{ row.ordine or '' }}"></div>
        
        <div class="col-md-3"><label class="form-label">Data Ingresso</label><input type="date" name="data_ingresso" class="form-control" value="{{ row.data_ingresso or '' }}"></div>
        <div class="col-md-3"><label class="form-label">DDT Ingresso</label><input type="text" name="n_ddt_ingresso" class="form-control" value="{{ row.n_ddt_ingresso or '' }}"></div>
        <div class="col-md-3"><label class="form-label">Data Uscita</label><input type="date" name="data_uscita" class="form-control" value="{{ row.data_uscita or '' }}"></div>
        <div class="col-md-3"><label class="form-label">DDT Uscita</label><input type="text" name="n_ddt_uscita" class="form-control" value="{{ row.n_ddt_uscita or '' }}"></div>
        
        <div class="col-md-2"><label class="form-label">Pezzi</label><input type="number" name="pezzo" class="form-control" value="{{ row.pezzo or '' }}"></div>
        <div class="col-md-2 bg-warning bg-opacity-10 rounded border border-warning">
            <label class="form-label fw-bold text-dark">N° Colli</label>
            <input type="number" name="n_colli" class="form-control fw-bold" value="{{ row.n_colli or 1 }}">
            <small class="text-muted" style="font-size:10px">Se > 1, crea N righe separate!</small>
        </div>
        <div class="col-md-2"><label class="form-label">Peso (Kg)</label><input type="number" step="0.01" name="peso" class="form-control" value="{{ row.peso or '' }}"></div>
        <div class="col-md-2"><label class="form-label">M³</label><input type="number" step="0.001" name="m3" class="form-control" value="{{ row.m3 or '' }}"></div>
        <div class="col-md-2"><label class="form-label">N. Arrivo</label><input type="text" name="n_arrivo" class="form-control" value="{{ row.n_arrivo or '' }}"></div>
        
        <div class="col-md-4">
            <label class="form-label">Dimensioni (LxPxH)</label>
            <div class="input-group">
                <input type="number" step="0.01" name="lunghezza" class="form-control" placeholder="L" value="{{ row.lunghezza or '' }}">
                <span class="input-group-text">x</span>
                <input type="number" step="0.01" name="larghezza" class="form-control" placeholder="P" value="{{ row.larghezza or '' }}">
                <span class="input-group-text">x</span>
                <input type="number" step="0.01" name="altezza" class="form-control" placeholder="H" value="{{ row.altezza or '' }}">
            </div>
        </div>
        
        <div class="col-md-2"><label class="form-label">Serial Number</label><input type="text" name="serial_number" class="form-control" value="{{ row.serial_number or '' }}"></div>
        <div class="col-md-2"><label class="form-label">Lotto</label><input type="text" name="lotto" class="form-control" value="{{ row.lotto or '' }}"></div>
        <div class="col-md-2"><label class="form-label">Ns. Rif</label><input type="text" name="ns_rif" class="form-control" value="{{ row.ns_rif or '' }}"></div>
        <div class="col-md-2"><label class="form-label">Mezzi Uscita</label><input type="text" name="mezzi_in_uscita" class="form-control" value="{{ row.mezzi_in_uscita or '' }}"></div>
        
        <div class="col-12"><label class="form-label">Note</label><textarea name="note" class="form-control" rows="3">{{ row.note or '' }}</textarea></div>

        {% if not row.id_articolo %}
        <div class="col-12 mt-3">
            <div class="card bg-white border border-primary p-3 shadow-sm">
                 <label class="form-label fw-bold text-primary fs-5"><i class="bi bi-cloud-upload"></i> Carica Documenti e Foto</label>
                 
                 <input type="file" name="new_files" class="form-control form-control-lg" multiple>
                 
                 <div class="alert alert-info mt-2 mb-0 py-2 small">
                    <i class="bi bi-info-circle-fill"></i> 
                    Puoi selezionare <strong>più file contemporaneamente</strong> (es. il PDF del documento e la FOTO del pacco).<br>
                    Tieni premuto il tasto <b>CTRL</b> (o CMD su Mac) mentre clicchi sui file nella finestra di selezione.
                 </div>
            </div>
        </div>
        {% endif %}
    </div>

    <div class="mt-4 text-end">
        <button type="submit" class="btn btn-primary px-5 btn-lg"><i class="bi bi-save"></i> {% if row.id_articolo %}Salva Modifiche{% else %}Crea Articolo{% endif %}</button>
    </div>
</form>

{% if row and row.id_articolo %}
<div class="card p-4 shadow-sm mb-5 border-top border-4 border-primary">
    <div class="d-flex justify-content-between align-items-center mb-3">
        <h5 class="m-0"><i class="bi bi-paperclip"></i> Allegati Salvati</h5>
        
        <form action="{{ url_for('upload_file', id_articolo=row.id_articolo) }}" method="post" enctype="multipart/form-data" class="d-flex gap-2 align-items-center">
            <input type="file" name="file" class="form-control" multiple required>
            <button type="submit" class="btn btn-success fw-bold"><i class="bi bi-cloud-upload"></i> Aggiungi File</button>
        </form>
    </div>
    <div class="small text-muted mb-3">Puoi caricare foto e PDF aggiuntivi selezionandoli insieme (tieni premuto CTRL).</div>
    <hr>
    
    <div class="row g-3">
        {% for att in row.attachments %}
        <div class="col-md-2 col-6">
            <div class="card h-100 text-center p-2 border bg-light position-relative shadow-sm">
                <div class="mb-2 text-primary" style="font-size:2.5em;">
                    {% if att.kind == 'photo' %}
                    <i class="bi bi-file-earmark-image"></i>
                    {% else %}
                    <i class="bi bi-file-earmark-pdf text-danger"></i>
                    {% endif %}
                </div>
                <div class="text-truncate small fw-bold mb-2 text-dark" title="{{ att.filename }}">
                    {{ att.filename.split('_', 2)[-1] }} 
                </div>
                
                <div class="btn-group btn-group-sm w-100">
                    <a href="{{ url_for('serve_uploaded_file', filename=att.filename) }}" target="_blank" class="btn btn-outline-primary">Apri</a>
                    <a href="{{ url_for('delete_attachment', id_attachment=att.id) }}" class="btn btn-outline-danger" onclick="return confirm('Sicuro di eliminare questo file?')">Elimina</a>
                </div>
            </div>
        </div>
        {% else %}
        <div class="col-12 text-center text-muted py-4 border border-dashed rounded">
            <i class="bi bi-inbox fs-3 d-block mb-2"></i>
            Nessun allegato presente per questo articolo.
        </div>
        {% endfor %}
    </div>
</div>
{% endif %}
{% endblock %}
"""


BULK_EDIT_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="container mt-4">
    <div class="d-flex justify-content-between align-items-center mb-3">
        <h3><i class="bi bi-ui-checks"></i> Modifica Multipla ({{ rows|length }} articoli)</h3>
        <a href="{{ url_for('giacenze') }}" class="btn btn-secondary">Annulla</a>
    </div>

    <div class="alert alert-warning shadow-sm">
        <i class="bi bi-exclamation-triangle-fill me-2"></i>
        <strong>Attenzione:</strong> Attiva la spunta accanto ai campi che vuoi modificare. 
        Il valore inserito verrà applicato a <b>TUTTI</b> gli articoli selezionati.
    </div>

    <form method="POST" enctype="multipart/form-data">
        <input type="hidden" name="save_bulk" value="true">
        {% for id in ids_csv.split(',') %}
        <input type="hidden" name="ids" value="{{ id }}">
        {% endfor %}

        <div class="card p-4 mb-4 bg-light border-dashed shadow-sm">
            <h5 class="text-primary"><i class="bi bi-cloud-upload"></i> Caricamento Allegati Massivo</h5>
            <div class="d-flex gap-2">
                <!-- ✅ CORRETTO: name="bulk_files" -->
                <input type="file" name="bulk_files" class="form-control" multiple>
            </div>
            <small class="text-muted">I file selezionati verranno allegati a ciascuno degli articoli.</small>
        </div>

        <div class="row g-3">
            {% for label, field_name in fields %}
            <div class="col-md-3 col-sm-6">
                <div class="card h-100 shadow-sm border-0">
                    <div class="card-header py-2 bg-white border-bottom-0 d-flex align-items-center gap-2">
                        <div class="form-check form-switch m-0">
                            <input class="form-check-input" type="checkbox" name="chk_{{ field_name }}" id="chk_{{ field_name }}" 
                                   onchange="document.getElementById('in_{{ field_name }}').disabled = !this.checked; 
                                             document.getElementById('in_{{ field_name }}').focus();">
                        </div>
                        <label for="chk_{{ field_name }}" class="m-0 fw-bold text-dark w-100" style="cursor:pointer; font-size:0.9rem;">
                            {{ label }}
                        </label>
                    </div>
                    <div class="card-body p-2 bg-light rounded-bottom">

                        {% if field_name == 'stato' %}
                            <input list="statoOptions" name="{{ field_name }}" id="in_{{ field_name }}" class="form-control form-control-sm" disabled placeholder="Seleziona...">
                            <datalist id="statoOptions">
                                <option value="NAZIONALE">
                                <option value="DOGANALE">
                                <option value="ESTERO">
                                <option value="USCITO">
                                <option value="FINCANTIERI SCOPERTO">
                                <option value="AGGIUNTO A MANO">
                            </datalist>

                        {% elif 'data' in field_name %}
                            <input type="date" name="{{ field_name }}" id="in_{{ field_name }}" class="form-control form-control-sm" disabled>

                        {% elif field_name in ['pezzo','n_colli','lunghezza','larghezza','altezza','peso'] %}
                            <input type="number" step="0.01" name="{{ field_name }}" id="in_{{ field_name }}" class="form-control form-control-sm" disabled>

                        {% else %}
                            <input type="text" name="{{ field_name }}" id="in_{{ field_name }}" class="form-control form-control-sm" disabled>
                        {% endif %}

                    </div>
                </div>
            </div>
            {% endfor %}
        </div>

        <div class="mt-5 text-center mb-5">
            <button type="submit" class="btn btn-warning btn-lg px-5 fw-bold shadow">
                <i class="bi bi-check-circle-fill"></i> Applica Modifiche a Tutti
            </button>
        </div>
    </form>
</div>
{% endblock %}
"""


BUONO_PREVIEW_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="card p-3 shadow-sm">
    <div class="d-flex align-items-center gap-3 mb-3">
        {% if logo_url %}<img src="{{ logo_url }}" style="height:40px">{% endif %}
        <h5 class="flex-grow-1 text-center m-0 text-uppercase fw-bold">Buono di Prelievo</h5>
        
        <div class="btn-group">
            <button type="button" class="btn btn-outline-primary" onclick="submitBuono('preview')">
                <i class="bi bi-eye"></i> Anteprima PDF
            </button>
            <button type="button" class="btn btn-success fw-bold" onclick="submitBuono('save')">
                <i class="bi bi-file-earmark-arrow-down"></i> Genera e Salva
            </button>
            <a href="{{ url_for('giacenze') }}" class="btn btn-secondary">Annulla</a>
        </div>
    </div>

    <form id="buono-form" method="POST" action="{{ url_for('buono_finalize_and_get_pdf') }}">
        <input type="hidden" name="ids" value="{{ ids }}">
        <input type="hidden" name="action" id="action_field" value="preview">

        <div class="row g-3 bg-light p-3 rounded border mb-3">
            <div class="col-md-2">
                <label class="form-label small fw-bold">N. Buono</label>
                <input name="buono_n" class="form-control" value="{{ meta.buono_n }}">
            </div>
            <div class="col-md-2">
                <label class="form-label small fw-bold">Data Em.</label>
                <input name="data_em" class="form-control" value="{{ meta.data_em }}" readonly>
            </div>
            <div class="col-md-2">
                <label class="form-label small fw-bold">Ordine</label>
                <input name="ordine" class="form-control" value="{{ meta.ordine }}">
            </div>
            <div class="col-md-2">
                <label class="form-label small fw-bold">Commessa</label>
                <input name="commessa" class="form-control" value="{{ meta.commessa }}">
            </div>
            <div class="col-md-2">
                <label class="form-label small fw-bold">Fornitore</label>
                <input name="fornitore" class="form-control" value="{{ meta.fornitore }}">
            </div>
            <div class="col-md-2">
                <label class="form-label small fw-bold">Protocollo</label>
                <input name="protocollo" class="form-control" value="{{ meta.protocollo }}">
            </div>
        </div>

        <div class="table-responsive">
            <table class="table table-sm table-bordered align-middle table-hover">
                <thead class="table-dark text-black" style="color:black !important;">
                    <tr>
                        <th style="width:10%">Ordine Orig.</th>
                        <th style="width:15%">Codice</th>
                        <th style="width:35%">Descrizione</th>
                        <th style="width:10%">Q.tà</th>
                        <th style="width:10%">N.Arr</th>
                    </tr>
                </thead>
                <tbody>
                    {% for r in rows %}
                    <tr class="table-light">
                        <td class="small">{{ r.ordine or '' }}</td>
                        <td class="fw-bold">{{ r.codice_articolo or '' }}</td>
                        <td class="small">{{ r.descrizione or '' }}</td>
                        <td>
                            <!-- ✅ Q.tà prende PEZZI -->
                            <input name="q_{{ r.id_articolo }}" type="number"
                                   class="form-control form-control-sm text-center fw-bold"
                                   value="{{ r.pezzo or 1 }}">
                        </td>
                        <td class="small text-center">{{ r.n_arrivo or '' }}</td>
                    </tr>
                    <tr>
                        <td colspan="2" class="text-end small text-muted align-middle" style="border-top:none;">Note:</td>
                        <td colspan="3" style="border-top:none;">
                            <textarea class="form-control form-control-sm border-0 bg-white text-primary"
                                      name="note_{{ r.id_articolo }}" rows="1"
                                      placeholder="Inserisci note aggiuntive...">{{ r.note or '' }}</textarea>
                        </td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
    </form>
</div>

<script>
function submitBuono(actionType) {
    const form = document.getElementById('buono-form');
    document.getElementById('action_field').value = actionType;

    if (actionType === 'preview') {
        form.target = '_blank';
        form.submit();
    } else {
        form.target = '_self';
        const formData = new FormData(form);
        const url = "{{ url_for('buono_finalize_and_get_pdf') }}"; 
        
        fetch(url, { method: 'POST', body: formData })
        .then(resp => {
            if (resp.ok) return resp.blob();
            return resp.text().then(text => { throw new Error(text) });
        })
        .then(blob => {
            const urlBlob = window.URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = urlBlob;
            a.download = 'Buono_Prelievo.pdf';
            document.body.appendChild(a);
            a.click();
            window.URL.revokeObjectURL(urlBlob);
            setTimeout(() => { window.location.href = '{{ url_for("giacenze") }}'; }, 1500);
        })
        .catch(err => alert("Errore durante il salvataggio:\\n" + err));
    }
}
</script>
{% endblock %}
"""

DDT_PREVIEW_HTML = """ 
{% extends 'base.html' %}
{% block content %}
<div class="card p-3">

    <!-- ✅ HEADER -->
    <div class="d-flex align-items-center gap-3 mb-4" style="padding-bottom:10px;">
        {% if logo_url %}
            <img src="{{ logo_url }}" style="height:70px; margin-bottom:10px;">
        {% endif %}

        <h5 class="flex-grow-1 text-center m-0" style="padding-top:10px;">
            DOCUMENTO DI TRASPORTO
        </h5>

        <div class="btn-group">
            <button type="button" class="btn btn-outline-primary" onclick="submitDdt('preview')">
                <i class="bi bi-printer"></i> Anteprima PDF
            </button>
            <button type="button" class="btn btn-success" onclick="submitDdt('finalize')">
                <i class="bi bi-check-circle-fill"></i> Finalizza e Scarica
            </button>
            <a href="{{ url_for('invia_email', ids=ids) }}" class="btn btn-warning">
                <i class="bi bi-envelope"></i> Invia Email
            </a>
            <a href="{{ url_for('giacenze') }}" class="btn btn-secondary">Annulla</a>
        </div>
    </div>

    <form id="ddt-form" method="POST" action="{{ url_for('ddt_finalize') }}">
        <input type="hidden" name="ids" value="{{ ids }}">
        <input type="hidden" name="action" id="action_field" value="preview">

        <div class="row g-3">
            <div class="col-md-4">
                <label class="form-label">Destinatario</label>
                <div class="input-group">
                    <select class="form-select" name="dest_key" required>
                        {% for k, v in destinatari.items() %}
                        <option value="{{ k }}">{{ k }} - {{ v.ragione_sociale }}</option>
                        {% endfor %}
                    </select>
                    <a href="{{ url_for('manage_destinatari') }}" class="btn btn-outline-secondary" target="_blank">
                        <i class="bi bi-pencil"></i>
                    </a>
                </div>
            </div>

            <div class="col-md-3">
                <label class="form-label">N. DDT</label>
                <div class="input-group">
                    <!-- ✅ PREV -->
                    <button class="btn btn-outline-secondary" type="button" id="get-prev-ddt" title="Numero precedente">
                        <i class="bi bi-arrow-left"></i>
                    </button>

                    <input name="n_ddt" id="n_ddt_input" class="form-control text-center" value="{{ n_ddt }}" required>

                    <!-- ✅ NEXT -->
                    <button class="btn btn-outline-secondary" type="button" id="get-next-ddt" title="Numero successivo">
                        <i class="bi bi-arrow-right"></i>
                    </button>
                </div>
                <div class="form-text">Usa ⬅️/➡️ per cambiare progressivo.</div>
            </div>

            <div class="col-md-2">
                <label class="form-label">Data DDT</label>
                <input name="data_ddt" type="date" class="form-control" value="{{ oggi }}" required>
            </div>

            <div class="col-md-3">
                <label class="form-label">Targa</label>
                <input name="targa" class="form-control">
            </div>

            <div class="col-md-4">
                <label class="form-label">Causale</label>
                <input name="causale" class="form-control" value="TRASFERIMENTO">
            </div>

            <div class="col-md-4">
                <label class="form-label">Porto</label>
                <input name="porto" class="form-control" value="FRANCO">
            </div>

            <div class="col-md-4">
                <label class="form-label">Aspetto</label>
                <input name="aspetto" class="form-control" value="A VISTA">
            </div>

            <!-- ✅ MEZZO IN USCITA -->
            <div class="col-md-4">
                <label class="form-label">Mezzo in uscita *</label>
                <select name="mezzi_in_uscita" id="mezzi_in_uscita" class="form-select">
                    <option value="" selected>-- Seleziona --</option>
                    <option value="MOTRICE">Motrice</option>
                    <option value="BILICO">Bilico</option>
                    <option value="FURGONE">Furgone</option>
                </select>
                <div class="form-text">Obbligatorio quando fai “Finalizza”.</div>
            </div>
        </div>

        <hr style="margin-top:18px; margin-bottom:18px;">

        <div class="mt-3" style="margin-bottom:18px;">
            <h5 class="mb-0" style="font-weight:600;">Articoli nel DDT</h5>
        </div>

        <div class="table-responsive" style="margin-top:8px;">
            <table class="table table-sm table-bordered align-middle">
                <thead class="table-light">
                    <tr>
                        <th>ID</th>
                        <th>Cod.Art.</th>
                        <th>Descrizione</th>
                        <th style="width: 250px;">Note (Editabili)</th>
                        <th style="width: 70px;">Pezzi</th>
                        <th style="width: 70px;">Colli</th>
                        <th style="width: 80px;">Peso</th>
                        <th>N.Arrivo</th>
                    </tr>
                </thead>

                <tbody>
                    {% for r in rows %}
                    <tr>
                        <td>{{ r.id_articolo }}</td>
                        <td>{{ r.codice_articolo or '' }}</td>
                        <td>{{ r.descrizione or '' }}</td>
                        <td>
                            <textarea class="form-control form-control-sm" name="note_{{ r.id_articolo }}" rows="1">{{ r.note or '' }}</textarea>
                        </td>
                        <td><input class="form-control form-control-sm" name="pezzi_{{ r.id_articolo }}" value="{{ r.pezzo or 1 }}"></td>
                        <td><input class="form-control form-control-sm" name="colli_{{ r.id_articolo }}" value="{{ r.n_colli or 1 }}"></td>
                        <td><input class="form-control form-control-sm" name="peso_{{ r.id_articolo }}" value="{{ r.peso or '' }}"></td>
                        <td>{{ r.n_arrivo or '' }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
    </form>
</div>
{% endblock %}

{% block extra_js %}
<script>
const nDdtInput = document.getElementById('n_ddt_input');

document.getElementById('get-next-ddt').addEventListener('click', function() {
    const current = (nDdtInput.value || '').trim();
    fetch('{{ url_for("get_next_ddt_number") }}?current=' + encodeURIComponent(current))
      .then(r => r.json())
      .then(d => {
          if (d.next_ddt) nDdtInput.value = d.next_ddt;
      });
});

document.getElementById('get-prev-ddt').addEventListener('click', function() {
    const current = (nDdtInput.value || '').trim();
    fetch('{{ url_for("get_prev_ddt_number") }}?current=' + encodeURIComponent(current))
      .then(r => r.json())
      .then(d => {
          if (d.prev_ddt) nDdtInput.value = d.prev_ddt;
      });
});

function submitDdt(actionType) {
    const form = document.getElementById('ddt-form');
    document.getElementById('action_field').value = actionType;

    // ✅ obbligatorio SOLO in finalize
    if (actionType === 'finalize') {
        const mezzo = (document.getElementById('mezzi_in_uscita').value || '').trim();
        if (!mezzo) {
            alert("Seleziona il Mezzo in uscita (Motrice / Bilico / Furgone) prima di finalizzare.");
            document.getElementById('mezzi_in_uscita').focus();
            return;
        }
    }

    if (actionType === 'preview') {
        form.target = '_blank';
        form.submit();
    } else {
        form.target = '_self';
        const formData = new FormData(form);
        const url = form.getAttribute('action');

        fetch(url, { method: 'POST', body: formData })
        .then(resp => {
            if (resp.ok) return resp.blob();
            return resp.text().then(t => { throw new Error(t || 'Errore finalizzazione'); });
        })
        .then(blob => {
            const urlBlob = window.URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = urlBlob;
            a.download = 'DDT_Finale.pdf';
            document.body.appendChild(a);
            a.click();
            window.URL.revokeObjectURL(urlBlob);
            setTimeout(() => { window.location.href = '{{ url_for("giacenze") }}'; }, 1500);
        })
        .catch(err => alert("Errore: " + err.message));
    }
}
</script>
{% endblock %}
"""



DDT_MEZZO_USCITA_HTML = """
{% extends "base.html" %}
{% block content %}
<div class="container" style="max-width:520px; margin-top:30px;">
  <div class="card shadow-sm">
    <div class="card-body">
      <h5 class="mb-2"><i class="bi bi-truck"></i> Mezzo in uscita</h5>

      {% if n_ddt %}
      <div class="alert alert-info py-2">
        DDT Uscita: <b>{{ n_ddt }}</b>
      </div>
      {% endif %}

      <form method="POST">
        <input type="hidden" name="ids" value="{{ ids }}">
        <input type="hidden" name="n_ddt" value="{{ n_ddt }}">

        <label class="form-label fw-bold">Seleziona mezzo *</label>
        <select name="mezzo" class="form-select" required>
          <option value="" selected disabled>-- Seleziona --</option>
          <option value="Motrice">Motrice</option>
          <option value="Bilico">Bilico</option>
          <option value="Furgone">Furgone</option>
        </select>

        <div class="d-flex justify-content-end mt-3">
          <button type="submit" class="btn btn-primary">
            <i class="bi bi-save"></i> Salva
          </button>
        </div>
      </form>

      <div class="text-muted small mt-3">
        Compilazione obbligatoria per aggiornare la colonna <b>Mezzo in uscita</b> nelle righe selezionate.
      </div>
    </div>
  </div>
</div>
{% endblock %}
"""
DDT_MEZZO_USCITA_OK_HTML = """
{% extends "base.html" %}
{% block content %}
<div class="container" style="max-width:520px; margin-top:30px;">
  <div class="alert alert-success shadow-sm">
    <b>Salvato!</b><br>
    Mezzo in uscita: <b>{{ mezzo }}</b><br>
    Righe aggiornate: <b>{{ count }}</b>
    {% if n_ddt %}<br>DDT: <b>{{ n_ddt }}</b>{% endif %}
  </div>

  <script>
    setTimeout(function(){ window.close(); }, 900);
  </script>
</div>
{% endblock %}
"""


LABELS_FORM_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="card p-4">
    <h3><i class="bi bi-tag"></i> Nuova Etichetta</h3>
    <hr>

    <div class="alert alert-info py-2">
        <i class="bi bi-info-circle"></i>
        Il PDF verrà <b>scaricato</b> automaticamente: aprilo e stampa dal file scaricato per mantenere il formato <b>100x62mm</b>.
    </div>

    <form method="post" action="{{ url_for('labels_pdf') }}">
        <div class="row g-3">
            <div class="col-md-4">
                <label class="form-label">Cliente</label>
                <input class="form-control" list="clienti-datalist" name="cliente" placeholder="Digita o seleziona un cliente...">
                <datalist id="clienti-datalist">
                    {% for c in clienti %}
                    <option value="{{ c }}">
                    {% endfor %}
                </datalist>
            </div>
            <div class="col-md-4"><label class="form-label">Fornitore</label><input name="fornitore" class="form-control"></div>
            <div class="col-md-4"><label class="form-label">Ordine</label><input name="ordine" class="form-control"></div>
            <div class="col-md-4"><label class="form-label">Commessa</label><input name="commessa" class="form-control"></div>
            <div class="col-md-4"><label class="form-label">DDT Ingresso</label><input name="ddt_ingresso" class="form-control"></div>
            <div class="col-md-4"><label class="form-label">Data Ingresso</label><input name="data_ingresso" class="form-control" placeholder="gg/mm/aaaa"></div>
            <div class="col-md-4"><label class="form-label">Arrivo (es. 01/25)</label><input name="arrivo" class="form-control"></div>
            <div class="col-md-4"><label class="form-label">N. Colli</label><input name="n_colli" class="form-control"></div>
            <div class="col-md-4"><label class="form-label">Posizione</label><input name="posizione" class="form-control"></div>
        </div>

        <div class="mt-4 d-flex gap-2">
            <button type="submit" class="btn btn-primary">
                <i class="bi bi-printer"></i> Scarica PDF Etichetta
            </button>
        </div>
    </form>
</div>
{% endblock %}
"""

LABELS_PREVIEW_HTML = " " # Non più utilizzato

IMPORT_EXCEL_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="row justify-content-center">
    <div class="col-md-8 col-lg-6">
        <div class="card p-4">
            <h3><i class="bi bi-file-earmark-arrow-up"></i> Importa Articoli da Excel</h3>
            <hr>
            
            {% if not profiles %}
            <div class="alert alert-warning">
                <i class="bi bi-exclamation-triangle"></i> <strong>Attenzione:</strong> Nessun profilo di importazione trovato. 
                <br>Carica prima il file <code>mappe_excel.json</code> nella sezione <a href="{{ url_for('manage_mappe') }}">Gestisci Mappe</a>.
            </div>
            {% else %}
            <p class="text-muted">Seleziona il profilo di mappatura corretto per il tuo file.</p>
            <form method="post" enctype="multipart/form-data">
                <div class="mb-3">
                    <label class="form-label fw-bold">1. Seleziona Profilo Mappa</label>
                    <select name="profile" class="form-select" required>
                        <option value="" disabled selected>-- Scegli dalla lista --</option>
                        {% for p in profiles %}
                        <option value="{{ p }}">{{ p }}</option>
                        {% endfor %}
                    </select>
                    <div class="form-text">Profili caricati: {{ profiles|length }}</div>
                </div>
                
                <div class="mb-3">
                    <label for="excel_file" class="form-label fw-bold">2. Carica File Excel</label>
                    <input class="form-control" type="file" id="excel_file" name="excel_file" accept=".xlsx,.xls,.xlsm" required>
                </div>
                
                <div class="d-grid gap-2">
                    <button type="submit" class="btn btn-primary btn-lg">Avvia Importazione</button>
                    <a href="{{ url_for('home') }}" class="btn btn-outline-secondary">Annulla</a>
                </div>
            </form>
            {% endif %}
            
            <div class="mt-4 text-center">
                <a href="{{ url_for('manage_mappe') }}" class="small text-decoration-none"><i class="bi bi-gear"></i> Gestisci file mappe_excel.json</a>
            </div>
        </div>
    </div>
</div>
{% endblock %}
"""
MAPPE_EXCEL_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="container">
    <div class="card p-4">
        <h3><i class="bi bi-gear"></i> Gestione Mappe Excel (JSON)</h3>
        <hr>
        <p>Qui puoi visualizzare o aggiornare il file <code>mappe_excel.json</code> che definisce come leggere i file Excel.</p>
        
        <form method="post">
            <div class="mb-3">
                <label class="form-label">Contenuto JSON corrente:</label>
                <textarea name="json_content" class="form-control" rows="15" style="font-family: monospace; font-size: 12px;">{{ content }}</textarea>
            </div>
            <button type="submit" class="btn btn-success">Salva Modifiche</button>
            <a href="{{ url_for('import_excel') }}" class="btn btn-secondary">Torna all'Import</a>
        </form>

        <hr class="my-4">
        <h5>Oppure carica un nuovo file .json</h5>
        <form method="post" enctype="multipart/form-data" action="{{ url_for('upload_mappe_json') }}">
             <div class="input-group mb-3">
                <input type="file" class="form-control" name="json_file" accept=".json" required>
                <button class="btn btn-outline-primary" type="submit">Carica File</button>
             </div>
        </form>
    </div>
</div>
{% endblock %}
"""
EXPORT_CLIENT_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="row justify-content-center">
    <div class="col-md-8 col-lg-6">
        <div class="card p-4">
            <h3><i class="bi bi-people"></i> Export Excel per Cliente</h3>
            <hr>
            <p>Seleziona un cliente dall'elenco per scaricare il file Excel con solo le sue giacenze.</p>
            <form method="post">
                <div class="mb-3">
                    <label for="cliente" class="form-label">Cliente</label>
                    <select class="form-select" id="cliente" name="cliente" required>
                        <option value="" disabled selected>-- Seleziona un cliente --</option>
                        {% for c in clienti %}
                        <option value="{{ c }}">{{ c }}</option>
                        {% endfor %}
                    </select>
                </div>
                <button type="submit" class="btn btn-primary">Scarica Excel</button>
                <a href="{{ url_for('home') }}" class="btn btn-secondary">Annulla</a>
            </form>
        </div>
    </div>
</div>
{% endblock %}
"""
DESTINATARI_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-md-10 col-lg-8">
    <div class="card p-4">
      <h3><i class="bi bi-person-rolodex"></i> Gestione Destinatari</h3>
      <hr>

      <h5>Aggiungi Nuovo Destinatario</h5>
      <form method="post" class="mb-4">
        <div class="row g-3">
          <div class="col-md-6">
            <label class="form-label">Nome Chiave (es. FINCANTIERI)</label>
            <input name="key_name" class="form-control" required>
          </div>
          <div class="col-md-6">
            <label class="form-label">Ragione Sociale</label>
            <input name="ragione_sociale" class="form-control" required>
          </div>
          <div class="col-md-6">
            <label class="form-label">Indirizzo Completo</label>
            <input name="indirizzo" class="form-control">
          </div>
          <div class="col-md-6">
            <label class="form-label">Partita IVA</label>
            <input name="piva" class="form-control">
          </div>
        </div>

        <button type="submit" class="btn btn-primary mt-3">
          <i class="bi bi-plus-lg"></i> Aggiungi
        </button>
      </form>

      <hr>
      <h5>Destinatari Esistenti</h5>

      <ul class="list-group">
        {% for key, details in destinatari.items() %}
        <li class="list-group-item d-flex justify-content-between align-items-center">
          <div>
            <strong>{{ key }}</strong><br>
            <small class="text-muted">
              {{ details.ragione_sociale or '' }}{% if details.indirizzo %} - {{ details.indirizzo }}{% endif %}
            </small>
          </div>

          <!-- ✅ ELIMINAZIONE CORRETTA: POST alla stessa pagina -->
          <form method="post" class="m-0">
            <input type="hidden" name="delete_key" value="{{ key }}">
            <button type="submit"
                    class="btn btn-sm btn-outline-danger"
                    onclick="return confirm('Sei sicuro di voler eliminare questo destinatario?')">
              <i class="bi bi-trash"></i>
            </button>
          </form>
        </li>
        {% else %}
        <li class="list-group-item">Nessun destinatario salvato.</li>
        {% endfor %}
      </ul>

      <a href="{{ request.referrer or url_for('home') }}" class="btn btn-secondary mt-4">Indietro</a>
    </div>
  </div>
</div>
{% endblock %}
"""

RUBRICA_EMAIL_HTML = """
{% extends 'base.html' %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-md-10 col-lg-9">
    <div class="card p-4 shadow-sm">
      <div class="d-flex justify-content-between align-items-center">
        <h3 class="mb-0"><i class="bi bi-journal-bookmark"></i> Rubrica Email</h3>
        <a href="{{ url_for('home') }}" class="btn btn-outline-secondary btn-sm">Torna alla Home</a>
      </div>
      <hr>

      <div class="row g-4">
        <div class="col-md-6">
          <h5>Contatti</h5>
          <form method="post" class="row g-2 mb-3">
            <input type="hidden" name="action" value="save_contact">
            <div class="col-5">
              <input class="form-control" name="nome" placeholder="Nome (es. Ufficio Genova)">
            </div>
            <div class="col-7">
              <input class="form-control" name="email" placeholder="email@dominio.it">
            </div>
            <div class="col-12 d-grid">
              <button class="btn btn-primary btn-sm">Salva Contatto</button>
            </div>
          </form>

          <div class="table-responsive">
            <table class="table table-sm table-striped align-middle">
              <thead class="table-light">
                <tr><th>Nome</th><th>Email</th><th class="text-end">Azioni</th></tr>
              </thead>
              <tbody>
              {% for nome, info in rubrica.contatti.items() %}
                <tr>
                  <td>{{ nome }}</td>
                  <td>{{ info.email }}</td>
                  <td class="text-end">
                    <form method="post" class="d-inline">
                      <input type="hidden" name="action" value="delete_contact">
                      <input type="hidden" name="nome" value="{{ nome }}">
                      <button class="btn btn-outline-danger btn-sm" onclick="return confirm('Eliminare contatto?')">
                        <i class="bi bi-trash"></i>
                      </button>
                    </form>
                  </td>
                </tr>
              {% else %}
                <tr><td colspan="3" class="text-center text-muted">Nessun contatto.</td></tr>
              {% endfor %}
              </tbody>
            </table>
          </div>
          <div class="form-text">
            Suggerimento: nei gruppi puoi incollare una lista di email separate da <b>;</b> o <b>,</b>.
          </div>
        </div>

        <div class="col-md-6">
          <h5>Gruppi</h5>
          <form method="post" class="mb-3">
            <input type="hidden" name="action" value="save_group">
            <div class="mb-2">
              <input class="form-control" name="gruppo" placeholder="Nome gruppo (es. FINCANTIERI)">
            </div>
            <div class="mb-2">
              <textarea class="form-control" name="emails" rows="3" placeholder="email1@...; email2@..."></textarea>
            </div>
            <button class="btn btn-success btn-sm w-100">Salva Gruppo</button>
          </form>

          <div class="accordion" id="accGruppi">
            {% for g, emails in rubrica.gruppi.items() %}
            <div class="accordion-item">
              <h2 class="accordion-header" id="h{{ loop.index }}">
                <button class="accordion-button collapsed" type="button" data-bs-toggle="collapse" data-bs-target="#c{{ loop.index }}">
                  {{ g }} <span class="badge bg-secondary ms-2">{{ emails|length }}</span>
                </button>
              </h2>
              <div id="c{{ loop.index }}" class="accordion-collapse collapse" data-bs-parent="#accGruppi">
                <div class="accordion-body">
                  <div class="small text-muted mb-2">Email:</div>
                  <div class="border rounded p-2 small" style="white-space: pre-wrap;">{{ emails|join('; ') }}</div>
                  <form method="post" class="mt-2">
                    <input type="hidden" name="action" value="delete_group">
                    <input type="hidden" name="gruppo" value="{{ g }}">
                    <button class="btn btn-outline-danger btn-sm" onclick="return confirm('Eliminare gruppo?')">
                      <i class="bi bi-trash"></i> Elimina gruppo
                    </button>
                  </form>
                </div>
              </div>
            </div>
            {% else %}
              <div class="text-muted">Nessun gruppo.</div>
            {% endfor %}
          </div>

        </div>
      </div>

    </div>
  </div>
</div>
{% endblock %}
"""





# --- TEMPLATE PAGINA PICKING / LAVORAZIONI (Senza Emoji, usa Icone Bootstrap) ---

LAVORAZIONI_HTML = """
{% extends "base.html" %}
{% block content %}
<div class="container-fluid mt-4">
    <div class="d-flex justify-content-between align-items-center mb-3">
        <h2><i class="bi bi-gear"></i> Gestione Picking / Lavorazioni</h2>
        <a href="{{ url_for('home') }}" class="btn btn-secondary shadow-sm"><i class="bi bi-box-arrow-left"></i> Esci</a>
    </div>

    <div class="card p-3 mb-4 bg-light border shadow-sm">
        {% if edit_row %}
            <h5 class="mb-3 text-primary"><i class="bi bi-pencil-square"></i> Modifica Picking ID {{ edit_row.id }}</h5>
        {% else %}
            <h5 class="mb-3">Inserisci Nuovo Picking</h5>
        {% endif %}

        <form method="POST" class="row g-2">
            {% if edit_row %}
                <input type="hidden" name="edit_lavorazione" value="1">
                <input type="hidden" name="id" value="{{ edit_row.id }}">
            {% else %}
                <input type="hidden" name="add_lavorazione" value="1">
            {% endif %}

            <div class="col-md-2">
                <label class="small fw-bold">Data</label>
                <input type="date" name="data" class="form-control" required
                       value="{% if edit_row and edit_row.data %}{{ edit_row.data }}{% else %}{{ today }}{% endif %}">
            </div>

            <div class="col-md-2"><label class="small fw-bold">Cliente</label>
                <input type="text" name="cliente" class="form-control"
                       value="{{ edit_row.cliente if edit_row else '' }}">
            </div>

            <div class="col-md-3"><label class="small fw-bold">Descrizione</label>
                <input type="text" name="descrizione" class="form-control"
                       value="{{ edit_row.descrizione if edit_row else '' }}">
            </div>

            <div class="col-md-2"><label class="small fw-bold">Richiesta Di</label>
                <input type="text" name="richiesta_di" class="form-control"
                       value="{{ edit_row.richiesta_di if edit_row else '' }}">
            </div>

            <div class="col-md-3"><label class="small fw-bold">Seriali</label>
                <input type="text" name="seriali" class="form-control"
                       value="{{ edit_row.seriali if edit_row else '' }}">
            </div>

            <div class="col-md-1"><label class="small fw-bold">Colli</label>
                <input type="number" name="colli" class="form-control"
                       value="{{ edit_row.colli if edit_row else '' }}">
            </div>

            <div class="col-md-1"><label class="small fw-bold">Pallet Entrati</label>
                <input type="number" name="pallet_forniti" class="form-control"
                       value="{{ edit_row.pallet_forniti if edit_row else '' }}">
            </div>

            <div class="col-md-1"><label class="small fw-bold">Pallet Usciti</label>
                <input type="number" name="pallet_uscita" class="form-control"
                       value="{{ edit_row.pallet_uscita if edit_row else '' }}">
            </div>

            <div class="col-md-1"><label class="small fw-bold">Ore Blue</label>
                <input type="number" step="0.5" name="ore_blue_collar" class="form-control"
                       value="{{ edit_row.ore_blue_collar if edit_row else '' }}">
            </div>

            <div class="col-md-1"><label class="small fw-bold">Ore White</label>
                <input type="number" step="0.5" name="ore_white_collar" class="form-control"
                       value="{{ edit_row.ore_white_collar if edit_row else '' }}">
            </div>

            <div class="col-md-12 text-end mt-2">
                {% if edit_row %}
                    <a href="{{ url_for('lavorazioni') }}" class="btn btn-secondary fw-bold">Annulla</a>
                    <button type="submit" class="btn btn-primary fw-bold">
                        <i class="bi bi-save"></i> Salva Modifica
                    </button>
                {% else %}
                    <button type="submit" class="btn btn-success fw-bold">
                        <i class="bi bi-plus-lg"></i> Aggiungi
                    </button>
                {% endif %}
            </div>
        </form>
    </div>

    <div class="card p-3 mb-3 border-warning shadow-sm">
        <h6 class="text-warning-emphasis fw-bold"><i class="bi bi-printer"></i> Stampa Report Picking</h6>
        <form action="{{ url_for('stampa_picking_pdf') }}" method="POST" target="_blank" class="row g-2 align-items-end">
            <div class="col-md-2">
                <label class="small">Mese (es. 2025-01)</label>
                <input type="month" name="mese" class="form-control form-control-sm">
            </div>
            <div class="col-md-3">
                <label class="small">Cliente</label>
                <input type="text" name="cliente" class="form-control form-control-sm" placeholder="Tutti">
            </div>
            <div class="col-md-2">
                <button type="submit" class="btn btn-warning btn-sm w-100 fw-bold">Genera PDF</button>
            </div>
        </form>
    </div>

    <div class="card shadow-sm">
        <div class="table-responsive">
            <table class="table table-bordered table-hover mb-0 align-middle">
                <thead class="table-light" style="color:#000;">
                    <tr>
                        <th>Data</th><th>Cliente</th><th>Descrizione</th>
                        <th>Richiesta</th><th>Seriali</th><th>Colli</th>
                        <th>Pallet Entrati</th><th>Pallet Usciti</th>
                        <th>Blue</th><th>White</th>
                        <th>Azioni</th>
                    </tr>
                </thead>
                <tbody>
                    {% for l in lavorazioni %}
                    <tr>
                        <td>{{ l.data or '' }}</td>
                        <td>{{ l.cliente or '' }}</td>
                        <td>{{ l.descrizione or '' }}</td>
                        <td>{{ l.richiesta_di or '' }}</td>
                        <td>{{ l.seriali or '' }}</td>
                        <td>{{ l.colli or '' }}</td>
                        <td>{{ l.pallet_forniti or '' }}</td>
                        <td>{{ l.pallet_uscita or '' }}</td>
                        <td>{{ l.ore_blue_collar or '' }}</td>
                        <td>{{ l.ore_white_collar or '' }}</td>
                        <td class="d-flex gap-1">
                            {% if session.get('role') == 'admin' %}
                            <a href="{{ url_for('lavorazioni', edit_id=l.id) }}"
                               class="btn btn-sm btn-primary"
                               title="Modifica">
                               <i class="bi bi-pencil"></i>
                            </a>
                            <a href="{{ url_for('elimina_record', table='lavorazioni', id=l.id) }}"
                               class="btn btn-sm btn-danger"
                               onclick="return confirm('Sei sicuro di voler eliminare?')"
                               title="Elimina"><i class="bi bi-trash"></i></a>
                            {% else %}
                                -
                            {% endif %}
                        </td>
                    </tr>
                    {% else %}
                    <tr><td colspan="11" class="text-center text-muted">Nessuna attività registrata.</td></tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
    </div>
</div>
{% endblock %}
"""


# ==========================================================
# TEMPLATE ADMIN BACKUPS (gestito dentro al file Python)
# ==========================================================

ADMIN_BACKUPS_HTML = """
{% extends "base.html" %}
{% block content %}

<div class="container-fluid mt-4">
  <h3><i class="bi bi-hdd-stack"></i> Backup & Ripristino</h3>

  <div class="alert alert-info">
    I backup sono salvati su disco persistente Render:<br>
    <b>/var/data/app/backups</b>
  </div>

  {% if backups %}
    <div class="card shadow-sm">
      <div class="table-responsive">
        <table class="table table-striped align-middle mb-0">
          <thead style="background:#f0f0f0;">
            <tr>
              <th>File Backup</th>
              <th class="text-center">Data</th>
              <th class="text-center">Dimensione (MB)</th>
              <th class="text-end">Azioni</th>
            </tr>
          </thead>

          <tbody>
            {% for b in backups %}
            <tr>
              <td><code>{{ b.name }}</code></td>
              <td class="text-center">{{ b.mtime }}</td>
              <td class="text-center">{{ b.size_mb }}</td>

              <td class="text-end">

                <!-- DOWNLOAD -->
                <a class="btn btn-sm btn-outline-primary"
                   href="{{ url_for('admin_backup_download', filename=b.name) }}">
                  <i class="bi bi-download"></i> Scarica
                </a>

                <!-- RIPRISTINA DB + JSON -->
                <form method="post"
                      style="display:inline-block"
                      onsubmit="return confirm('Confermi ripristino di questo backup?');">
                  <input type="hidden" name="action" value="restore">
                  <input type="hidden" name="filename" value="{{ b.name }}">
                  <input type="hidden" name="restore_media" value="0">

                  <button type="submit" class="btn btn-sm btn-warning">
                    <i class="bi bi-arrow-counterclockwise"></i>
                    Ripristina DB
                  </button>
                </form>

                <!-- RIPRISTINO COMPLETO -->
                <form method="post"
                      style="display:inline-block"
                      onsubmit="return confirm('Ripristino completo (DB+PDF+Foto). Confermi?');">
                  <input type="hidden" name="action" value="restore">
                  <input type="hidden" name="filename" value="{{ b.name }}">
                  <input type="hidden" name="restore_media" value="1">

                  <button type="submit" class="btn btn-sm btn-danger">
                    <i class="bi bi-exclamation-triangle"></i>
                    Ripristina Completo
                  </button>
                </form>

              </td>
            </tr>
            {% endfor %}
          </tbody>

        </table>
      </div>
    </div>

  {% else %}
    <div class="alert alert-warning">
      Nessun backup trovato nella cartella backups.
    </div>
  {% endif %}

  <a href="{{ url_for('home') }}" class="btn btn-outline-secondary mt-3">
    <i class="bi bi-arrow-left"></i> Torna alla Home
  </a>
</div>

{% endblock %}
"""


CALCOLA_COSTI_HTML = """
{% extends 'base.html' %}
{% block content %}

<div class="container-fluid mt-4">

  <h2>
    <i class="bi bi-calculator"></i>
    {% if metric == 'colli' %}
      Report Costi Magazzino (Colli in giacenza)
    {% else %}
      Report Costi Magazzino (M² per cliente)
    {% endif %}
  </h2>

  <div class="card p-3 mb-3 bg-light border shadow-sm">
    <form method="post" class="row g-2 align-items-end">
      <div class="col-md-2">
        <label class="form-label fw-bold">Data da:</label>
        <input type="date" name="data_da" class="form-control" value="{{ data_da }}" required>
      </div>

      <div class="col-md-2">
        <label class="form-label fw-bold">Data a:</label>
        <input type="date" name="data_a" class="form-control" value="{{ data_a }}" required>
      </div>

      <div class="col-md-4">
        <label class="form-label fw-bold">Cliente (contiene):</label>
        <input type="text" name="cliente" class="form-control" value="{{ cliente_filtro or '' }}" placeholder="es. FINCANTIERI">
      </div>

      <div class="col-md-3">
        <label class="form-label fw-bold d-block">Raggruppa:</label>
        <div class="form-check form-check-inline">
          <input class="form-check-input" type="radio" name="raggruppamento" id="rg_mese" value="mese"
                 {% if raggruppamento != 'giorno' %}checked{% endif %}>
          <label class="form-check-label" for="rg_mese">Per mese</label>
        </div>

        <div class="form-check form-check-inline">
          <input class="form-check-input" type="radio" name="raggruppamento" id="rg_giorno" value="giorno"
                 {% if raggruppamento == 'giorno' %}checked{% endif %}>
          <label class="form-check-label" for="rg_giorno">Per giorno</label>
        </div>
      </div>

      <div class="col-md-1 d-grid">
        <button type="submit" class="btn btn-secondary">Calcola</button>
      </div>
    </form>
  </div>

  {% if risultati %}
    <div class="card shadow-sm">
      <div class="table-responsive">
        <table class="table table-striped table-hover mb-0 align-middle">
          <thead style="background:#f0f0f0;">
            <tr>
              <th class="text-center">Periodo</th>
              <th>Cliente</th>

              {% if metric == 'colli' %}
                <th class="text-end">Colli Tot</th>
                <th class="text-end">Colli Medio</th>
              {% else %}
                <th class="text-end">M² Tot</th>
                <th class="text-end">M² Medio</th>
              {% endif %}

              <th class="text-center">Giorni</th>
            </tr>
          </thead>
          <tbody>
            {% for r in risultati %}
            <tr>
              <td class="text-center">{{ r.periodo }}</td>
              <td>{{ r.cliente }}</td>
              <td class="text-end">{{ r.tot }}</td>
              <td class="text-end">{{ r.medio }}</td>
              <td class="text-center">{{ r.giorni }}</td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  {% else %}
    {% if calcolato %}
      <div class="alert alert-warning">Nessun dato trovato per i criteri selezionati.</div>
    {% endif %}
  {% endif %}

  <a href="{{ url_for('home') }}" class="btn btn-outline-secondary mt-3">Torna alla Home</a>
</div>

{% endblock %}
"""



# --- TEMPLATE PAGINA TRASPORTI (Senza Emoji, usa Icone Bootstrap) ---
TRASPORTI_HTML = """
{% extends "base.html" %}
{% block content %}
<div class="container-fluid mt-4">

    <div class="d-flex justify-content-between align-items-center mb-3">
        <h2 class="m-0"><i class="bi bi-truck"></i> Gestione Trasporti</h2>
        <a href="{{ url_for('home') }}" class="btn btn-secondary shadow-sm">
            <i class="bi bi-box-arrow-left"></i> Esci
        </a>
    </div>

    <div class="card p-3 mb-4 bg-light border shadow-sm">
        {% if edit_row %}
            <h5 class="mb-3 text-primary"><i class="bi bi-pencil-square"></i> Modifica Trasporto ID {{ edit_row.id }}</h5>
        {% else %}
            <h5 class="mb-3 text-success"><i class="bi bi-plus-circle"></i> Inserisci Nuovo Trasporto</h5>
        {% endif %}

        <form method="POST" class="row g-2">
            {% if edit_row %}
                <input type="hidden" name="edit_trasporto" value="1">
                <input type="hidden" name="id" value="{{ edit_row.id }}">
            {% else %}
                <input type="hidden" name="add_trasporto" value="1">
            {% endif %}

            <div class="col-md-2">
                <label class="small fw-bold">Data</label>
                <input type="date" name="data" class="form-control" required
                       value="{% if edit_row and edit_row.data %}{{ edit_row.data }}{% else %}{{ today }}{% endif %}">
            </div>

            <div class="col-md-2"><label class="small fw-bold">Tipo Mezzo</label>
                <input type="text" name="tipo_mezzo" class="form-control"
                       value="{{ edit_row.tipo_mezzo if edit_row else '' }}" placeholder="es. Bilico">
            </div>

            <div class="col-md-2"><label class="small fw-bold">Cliente</label>
                <input type="text" name="cliente" class="form-control"
                       value="{{ edit_row.cliente if edit_row else '' }}">
            </div>

            <div class="col-md-2"><label class="small fw-bold">Trasportatore</label>
                <input type="text" name="trasportatore" class="form-control"
                       value="{{ edit_row.trasportatore if edit_row else '' }}">
            </div>

            <div class="col-md-1"><label class="small fw-bold">N. DDT</label>
                <input type="text" name="ddt_uscita" class="form-control"
                       value="{{ edit_row.ddt_uscita if edit_row else '' }}">
            </div>

            <div class="col-md-1"><label class="small fw-bold">Magazzino</label>
                <input type="text" name="magazzino" class="form-control"
                       value="{{ edit_row.magazzino if edit_row else '' }}">
            </div>

            <div class="col-md-1"><label class="small fw-bold">Consolidato</label>
                <input type="text" name="consolidato" class="form-control"
                       value="{{ edit_row.consolidato if edit_row else '' }}">
            </div>

            <div class="col-md-1"><label class="small fw-bold">Costo €</label>
                <input type="text" name="costo" class="form-control" placeholder="0,00"
                       value="{% if edit_row and edit_row.costo is not none %}{{ '%.2f'|format(edit_row.costo) }}{% endif %}">
            </div>

            <div class="col-md-12 text-end mt-2">
                {% if edit_row %}
                    <a href="{{ url_for('trasporti') }}" class="btn btn-secondary fw-bold">
                        Annulla
                    </a>
                    <button type="submit" class="btn btn-primary fw-bold px-4">
                        <i class="bi bi-save"></i> Salva Modifica
                    </button>
                {% else %}
                    <button type="submit" class="btn btn-success fw-bold px-4">
                        <i class="bi bi-save"></i> Salva
                    </button>
                {% endif %}
            </div>
        </form>
    </div>

    <div class="card p-3 mb-3 border-primary shadow-sm">
        <h6 class="text-primary fw-bold"><i class="bi bi-printer"></i> Stampa Report PDF</h6>
        <form action="{{ url_for('report_trasporti') }}" method="POST" target="_blank" class="row g-2 align-items-end">
            <div class="col-md-2">
                <label class="small">Seleziona Mese</label>
                <input type="month" name="mese" class="form-control form-control-sm">
            </div>
            <div class="col-md-2">
                <label class="small">Filtra Cliente</label>
                <input type="text" name="cliente" class="form-control form-control-sm" placeholder="Tutti">
            </div>
            <div class="col-md-2">
                <label class="small">Filtra Mezzo</label>
                <input type="text" name="tipo_mezzo" class="form-control form-control-sm" placeholder="Tutti">
            </div>
            <div class="col-md-2">
                <button type="submit" class="btn btn-primary btn-sm w-100 fw-bold">
                    Genera PDF
                </button>
            </div>
        </form>
    </div>

    <div class="card shadow-sm">
        <div class="table-responsive">
            <table class="table table-striped table-hover mb-0 align-middle">
                <thead class="table-light" style="color:#000;">
                    <tr>
                        <th>Data</th><th>Mezzo</th><th>Cliente</th>
                        <th>Trasportatore</th><th>DDT</th><th>Mag.</th>
                        <th>Consolidato</th><th>Costo</th>
                        <th>Azioni</th>
                    </tr>
                </thead>
                <tbody>
                    {% for t in trasporti %}
                    <tr>
                        <td>{{ t.data or '' }}</td>
                        <td>{{ t.tipo_mezzo or '' }}</td>
                        <td>{{ t.cliente or '' }}</td>
                        <td>{{ t.trasportatore or '' }}</td>
                        <td>{{ t.ddt_uscita or '' }}</td>
                        <td>{{ t.magazzino or '' }}</td>
                        <td>{{ t.consolidato or '' }}</td>
                        <td>€ {{ '%.2f'|format(t.costo) if t.costo is not none else '' }}</td>
                        <td class="d-flex gap-1">
                            {% if session.get('role') == 'admin' %}
                            <a href="{{ url_for('trasporti', edit_id=t.id) }}"
                               class="btn btn-sm btn-primary"
                               title="Modifica">
                               <i class="bi bi-pencil"></i>
                            </a>
                            <a href="{{ url_for('elimina_record', table='trasporti', id=t.id) }}"
                               class="btn btn-sm btn-danger"
                               onclick="return confirm('Sei sicuro di voler eliminare questo trasporto?')"
                               title="Elimina">
                               <i class="bi bi-trash"></i>
                            </a>
                            {% else %}
                                -
                            {% endif %}
                        </td>
                    </tr>
                    {% else %}
                    <tr><td colspan="9" class="text-center text-muted py-3">Nessun trasporto inserito.</td></tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
    </div>
</div>
{% endblock %}
"""


INVIA_EMAIL_HTML = """
{% extends "base.html" %}
{% block content %}
<div class="row justify-content-center">
    <div class="col-md-9">
        <div class="card p-4 shadow">
            <h4 class="mb-3"><i class="bi bi-envelope"></i> Invia Email</h4>

            {% if selected_ids %}
            <div class="alert alert-info py-2">
                <i class="bi bi-info-circle"></i> Hai selezionato <strong>{{ selected_ids.split(',')|length }}</strong> articoli.
            </div>
            {% endif %}

            <form method="post" enctype="multipart/form-data">
                <input type="hidden" name="selected_ids" value="{{ selected_ids }}">

                <div class="mb-3">
                    <label class="form-label">Destinatari</label>
                    <div class="mb-2">
                        <label class="form-label">Gruppo (rubrica)</label>
                        <select id="gruppo_email" class="form-select form-select-sm">
                            <option value="">-- Seleziona un gruppo --</option>
                            {% for g, emails in (email_groups or {}).items() %}
                                <option value="{{ emails|join('; ') }}">{{ g }} ({{ emails|length }})</option>
                            {% endfor %}
                        </select>
                        <div class="form-text">Se scegli un gruppo, le email verranno inserite automaticamente qui sotto.</div>
                    </div>

                    <input
                        type="text"
                        name="destinatario"
                        class="form-control"
                        placeholder="email1@dominio.it; email2@dominio.it"
                        required
                    >
                    <div class="form-text">
                        Puoi inserire più destinatari separandoli con <b>;</b> oppure <b>,</b>
                    </div>
                </div>

                <div class="mb-3">
                    <label class="form-label fw-bold">Oggetto</label>
                    <input type="text" name="oggetto" class="form-control" value="Documentazione Merce - Camar S.r.l." required>
                </div>

                <div class="mb-3">
                    <label class="form-label fw-bold">Messaggio</label>
                    <textarea name="messaggio" rows="6" class="form-control" style="font-family: Arial, sans-serif;">Buongiorno,

Di seguito inviamo il riepilogo della merce in oggetto.

Cordiali saluti,</textarea>
                    <div class="form-text text-muted">
                        <i class="bi bi-info-circle"></i> Il logo e la firma legale verranno aggiunti automaticamente sotto questo testo.
                    </div>
                </div>

                <div class="card bg-light mb-3 border-0">
                    <div class="card-body opacity-75">
                        <small class="text-uppercase fw-bold text-muted mb-2 d-block">Anteprima piè di pagina automatico:</small>
                        <div class="d-flex align-items-center gap-3 mb-2">
                            <img src="{{ url_for('static', filename='logo camar.jpg') }}" alt="Logo" style="height:50px;">
                            <div>
                                <strong>Camar S.r.l.</strong><br>
                                <span class="text-muted" style="font-size: 0.8rem;">Via Balleydier 52r – 16149 GENOVA</span>
                            </div>
                        </div>
                        <div style="font-size: 0.7rem; color: #666; max-height: 60px; overflow: hidden; text-overflow: ellipsis;">
                            (Seguono disclaimer legale, contatti uffici, telefoni, ecc...)
                        </div>
                    </div>
                </div>

                <div class="card bg-light mb-3">
                    <div class="card-body">
                        <h6 class="card-title">Opzioni</h6>

                        <div class="form-check">
                            <input class="form-check-input" type="checkbox" name="genera_ddt" id="genera_ddt" checked>
                            <label class="form-check-label" for="genera_ddt">Inserisci Riepilogo Merci in email (tabella)</label>
                        </div>

                        <div class="form-check mb-2">
                            <input class="form-check-input" type="checkbox" name="allega_file" id="allega_file" checked>
                            <label class="form-check-label" for="allega_file">Includi allegati esistenti (Foto/PDF degli articoli)</label>
                        </div>

                        <label class="form-label mt-2"><strong>Aggiungi altro allegato (dal PC):</strong></label>
                        <input type="file" name="allegati_extra" class="form-control" multiple>
                    </div>
                </div>

                <div class="d-flex justify-content-between">
                    <a href="{{ url_for('giacenze') }}" class="btn btn-secondary">Annulla</a>
                    <button type="submit" class="btn btn-primary px-4"><i class="bi bi-send"></i> Invia Email</button>
                </div>
            </form>
        </div>
    </div>
</div>

<script>
document.addEventListener('DOMContentLoaded', function(){
  const sel = document.getElementById('gruppo_email');
  const inp = document.querySelector('input[name="destinatario"]');
  if(sel && inp){
    sel.addEventListener('change', function(){
      if(this.value){
        inp.value = this.value;
      }
    });
  }
});
</script>
{% endblock %}
"""


templates = {
        'base.html': BASE_HTML,
        'login.html': LOGIN_HTML,
        'home.html': HOME_HTML,
        'giacenze.html': GIACENZE_HTML,
        
        
        'edit.html': EDIT_HTML,  
        
        'bulk_edit.html': BULK_EDIT_HTML,
        'buono_preview.html': BUONO_PREVIEW_HTML,
        'ddt_preview.html': DDT_PREVIEW_HTML,
        'labels_form.html': LABELS_FORM_HTML,
        'labels_preview.html': LABELS_PREVIEW_HTML,
        
        'import_excel.html': IMPORT_EXCEL_HTML,
        'import_pdf.html': IMPORT_PDF_HTML,
        'mappe_excel.html': MAPPE_EXCEL_HTML,
        
        # NUOVI MODULI (Trasporti e Picking)
        'trasporti.html': TRASPORTI_HTML,
        'lavorazioni.html': LAVORAZIONI_HTML,

        # ⚠️ MANCAVANO QUESTI DUE PER LE STAMPE:
        'report_trasporti_print.html': REPORT_TRASPORTI_HTML,
        'report_inventario_print.html': REPORT_INVENTARIO_HTML,

        # ALTRI MODULI (Se hai le variabili definite sopra, lasciali)
        'invia_email.html': INVIA_EMAIL_HTML,
        'export_client.html': EXPORT_CLIENT_HTML,
        'destinatari.html': DESTINATARI_HTML,
        'rubrica_email.html': RUBRICA_EMAIL_HTML,
        'calcoli.html': CALCOLI_HTML
    }

# ========================================================
# CONFIGURAZIONE FINALE (SENZA RICREARE L'APP)
# ========================================================
app.jinja_loader = DictLoader(templates)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret")
app.jinja_env.globals['getattr'] = getattr
app.jinja_env.filters['fmt_date'] = fmt_date


def logo_url():
    if not LOGO_PATH:
        return None
    p = Path(LOGO_PATH)
    if p.exists() and p.parent == STATIC_DIR:
        return url_for('static', filename=p.name)
    try:
        target = STATIC_DIR / Path(LOGO_PATH).name
        if not target.exists():
            target.write_bytes(p.read_bytes())
        return url_for('static', filename=target.name)
    except Exception:
        return None

@app.context_processor
def inject_globals():
    return dict(logo_url=logo_url())

# --- ROUTE PRINCIPALI E AUTH ---
@app.route('/')
def index():
    if not session.get('user'):
        return redirect(url_for('login'))
    return redirect(url_for('home'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Supporta sia 'username' che 'user' come nomi campo nel form
        username = (request.form.get('username') or request.form.get('user') or '').strip().upper()
        password = (request.form.get('password') or request.form.get('pwd') or '').strip()
        
        users_db = get_users()
        
        if username in users_db and verify_password(users_db[username], password):
            # 1. Crea l'oggetto utente
            role = 'admin' if username in ADMIN_USERS else 'client'
            user = User(username, role)
            
            # 2. PUNTO FONDAMENTALE: Effettua il login formale con Flask-Login
            login_user(user)
            
            # 3. Imposta variabili di sessione accessorie
            session['role'] = role
            session['user'] = username
            session['user_name'] = username
            
            flash(f"Benvenuto {username}", "success")
            
            # 4. Dopo il login vai SEMPRE in Home (non alle giacenze)
            return redirect(url_for('home'))
        else:
            flash("Credenziali non valide", "danger")
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("Logout effettuato con successo", "success")
    return redirect(url_for('login'))


@app.route('/')
@app.route('/home')
@login_required
def home():
    try:
        # Recupera dati per la dashboard (con gestione errori se il DB è vuoto)
        tot_articoli = 0
        tot_m2 = 0.0
        
        try:
            db = SessionLocal()
            tot_articoli = db.query(Articolo).count()
            # Calcolo somma M2 sicuro
            result = db.query(func.sum(Articolo.m2)).scalar()
            if result:
                tot_m2 = float(result)
            db.close()
        except Exception as e_db:
            print(f"Errore Dashboard DB: {e_db}")
            # Non bloccare l'app, mostra 0
            tot_articoli = 0
            tot_m2 = 0

        return render_template('home.html', 
                               tot_articoli=tot_articoli, 
                               tot_m2=round(tot_m2, 2),
                               today=date.today())
                               
    except Exception as e:
        # Se c'è un errore grave nel template o altro
        print(f"CRITICAL ERROR HOME: {e}")
        import traceback
        traceback.print_exc()
        # Fallback estremo: pagina bianca con errore leggibile
        return f"<h1>Errore Caricamento Home</h1><p>{e}</p><a href='/logout'>Logout</a>"

# ========================================================
# GESTIONE MAPPE EXCEL (CORRETTA + LOG DEBUG)
# ========================================================

def load_mappe():
    """Carica mappe_excel.json: prima da config. (con punto), poi fallback su root."""
    config_path = APP_DIR / "config." / "mappe_excel.json"   # <-- cartella con il punto
    root_path = APP_DIR / "mappe_excel.json"

    json_path = config_path if config_path.exists() else root_path
    print(f"DEBUG scelto json_path: {json_path}")

    if not json_path.exists():
        print("DEBUG NESSUN mappe_excel.json trovato -> {}")
        print("=== FINE DEBUG load_mappe() ===\n")
        return {}

    try:
        raw = json_path.read_text(encoding="utf-8")
        data = json.loads(raw)
        print(f"DEBUG size={len(raw)} bytes | profili={len(data) if isinstance(data, dict) else 'N/A'}")
        print("=== FINE DEBUG load_mappe() ===\n")
        return data if isinstance(data, dict) else {}
    except Exception as e:
        print(f"DEBUG ERRORE parsing mappe: {e}")
        print("=== FINE DEBUG load_mappe() ===\n")
        return {}


@app.route('/manage_mappe', methods=['GET', 'POST'])
@login_required
def manage_mappe():
    json_path = APP_DIR / "mappe_excel.json"

    print("\n=== DEBUG manage_mappe() ===")
    print(f"DEBUG manage_mappe: json_path={json_path}")

    if request.method == 'POST':
        content = request.form.get('json_content', '')
        try:
            json.loads(content)  # Validazione
            print(f"DEBUG manage_mappe: scrivo su {json_path}")

            json_path.write_text(content, encoding='utf-8')

            try:
                size = json_path.stat().st_size
            except Exception:
                size = "N/A"
            print(f"DEBUG manage_mappe: scritto OK. md5={_file_digest(json_path)} size={size}")

            flash("Mappa aggiornata con successo.", "success")
        except json.JSONDecodeError as e:
            print(f"DEBUG manage_mappe: JSON NON valido: {e}")
            flash(f"Errore nel formato JSON: {e}", "danger")
        except Exception as e:
            print(f"DEBUG manage_mappe: ERRORE scrittura file: {e}")
            flash(f"Errore scrittura mappa: {e}", "danger")

        print("=== FINE DEBUG manage_mappe() ===\n")
        return redirect(url_for('manage_mappe'))

    # GET
    content = ""
    if json_path.exists():
        try:
            content = json_path.read_text(encoding='utf-8')
            print(f"DEBUG manage_mappe GET: file esiste, md5={_file_digest(json_path)} size={len(content)}")
        except Exception as e:
            print(f"DEBUG manage_mappe GET: errore lettura file: {e}")
    else:
        print("DEBUG manage_mappe GET: file NON esiste")

    print("=== FINE DEBUG manage_mappe() ===\n")
    return render_template('mappe_excel.html', content=content)


@app.post('/upload_mappe_json')
@login_required
def upload_mappe_json():
    import json
    import hashlib

    if 'json_file' not in request.files:
        flash("Nessun file selezionato", "warning")
        return redirect(url_for('manage_mappe'))

    f = request.files['json_file']
    if not f or f.filename == '':
        flash("Nessun file selezionato", "warning")
        return redirect(url_for('manage_mappe'))

    target = APP_DIR / "mappe_excel.json"

    def file_digest(path):
        try:
            h = hashlib.md5()
            with open(path, "rb") as fp:
                for chunk in iter(lambda: fp.read(8192), b""):
                    h.update(chunk)
            return h.hexdigest()
        except Exception:
            return "N/A"

    print("\n=== DEBUG upload_mappe_json() ===")
    print(f"DEBUG upload_mappe_json: filename={f.filename}")
    print(f"DEBUG upload_mappe_json: target={target}")

    try:
        raw = f.read()

        # ✅ UTF-8 con BOM (capita spesso con file creati da Windows/Excel)
        try:
            content = raw.decode("utf-8-sig")
        except Exception:
            content = raw.decode("utf-8")

        # ✅ validazione JSON (se non è JSON valido -> eccezione)
        json.loads(content)

        # ✅ assicura cartella esista
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        target.write_text(content, encoding="utf-8")

        size = target.stat().st_size if target.exists() else "N/A"
        print(f"DEBUG upload_mappe_json: scritto OK. md5={file_digest(target)} size={size}")

        flash("File mappe_excel.json caricato correttamente.", "success")

    except Exception as e:
        print(f"DEBUG upload_mappe_json: ERRORE: {e}")
        flash(f"Errore nel file caricato: {e}", "danger")

    print("=== FINE DEBUG upload_mappe_json() ===\n")
    return redirect(url_for('manage_mappe'))


# --- GESTIONE TRASPORTI (ADMIN) ---

@app.route('/trasporti', methods=['GET', 'POST'])
@login_required
def trasporti():
    db = SessionLocal()
    try:
        # --- MODIFICA TRASPORTO ---
        if request.method == 'POST' and request.form.get('edit_trasporto'):
            if session.get('role') != 'admin':
                flash("ACCESSO NEGATO: Solo Admin.", "danger")
                return redirect(url_for('trasporti'))

            try:
                tid = int(request.form.get('id') or 0)
                rec = db.query(Trasporto).filter(Trasporto.id == tid).first()
                if not rec:
                    flash("Trasporto non trovato.", "danger")
                    return redirect(url_for('trasporti'))

                data_str = (request.form.get('data') or '').strip()
                rec.data = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None

                costo_str = (request.form.get('costo') or '').strip()
                rec.costo = float(costo_str.replace(',', '.')) if costo_str != '' else None

                rec.tipo_mezzo = (request.form.get('tipo_mezzo') or '').strip() or None
                rec.cliente = (request.form.get('cliente') or '').strip() or None
                rec.trasportatore = (request.form.get('trasportatore') or '').strip() or None
                rec.ddt_uscita = (request.form.get('ddt_uscita') or '').strip() or None
                rec.magazzino = (request.form.get('magazzino') or '').strip() or None
                rec.consolidato = (request.form.get('consolidato') or '').strip() or None

                db.commit()
                flash("Trasporto modificato!", "success")
            except Exception as e:
                db.rollback()
                flash(f"Errore modifica trasporto: {e}", "danger")

            return redirect(url_for('trasporti'))

        # --- AGGIUNGI NUOVO TRASPORTO ---
        if request.method == 'POST' and request.form.get('add_trasporto'):
            if session.get('role') != 'admin':
                flash("ACCESSO NEGATO: Solo Admin.", "danger")
                return redirect(url_for('trasporti'))

            try:
                data_str = (request.form.get('data') or '').strip()
                data_val = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None

                costo_str = (request.form.get('costo') or '').strip()
                costo_val = float(costo_str.replace(',', '.')) if costo_str != '' else None

                nuovo = Trasporto(
                    data=data_val,
                    tipo_mezzo=(request.form.get('tipo_mezzo') or '').strip() or None,
                    cliente=(request.form.get('cliente') or '').strip() or None,
                    trasportatore=(request.form.get('trasportatore') or '').strip() or None,
                    ddt_uscita=(request.form.get('ddt_uscita') or '').strip() or None,
                    magazzino=(request.form.get('magazzino') or '').strip() or None,
                    consolidato=(request.form.get('consolidato') or '').strip() or None,
                    costo=costo_val
                )

                db.add(nuovo)
                db.commit()
                flash("Trasporto salvato!", "success")
            except Exception as e:
                db.rollback()
                flash(f"Errore salvataggio trasporto: {e}", "danger")

            return redirect(url_for('trasporti'))

        # --- EDIT MODE (GET ?edit_id=) ---
        edit_id = request.args.get('edit_id')
        edit_row = None
        if edit_id and session.get('role') == 'admin':
            try:
                edit_row = db.query(Trasporto).filter(Trasporto.id == int(edit_id)).first()
            except:
                edit_row = None

        # --- VISUALIZZA LISTA ---
        dati = db.query(Trasporto).order_by(
            Trasporto.data.desc().nullslast(),
            Trasporto.id.desc()
        ).all()

        return render_template(
            'trasporti.html',
            trasporti=dati,
            today=date.today(),
            edit_row=edit_row
        )

    finally:
        db.close()


@app.route('/report_trasporti', methods=['POST'])
@login_required
def report_trasporti():
    if session.get('role') != 'admin':
        return "No Access", 403

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from sqlalchemy import func

    # Recupera i filtri dal form
    mese = (request.form.get('mese') or '').strip()              # es '2026-01'
    mezzo = (request.form.get('tipo_mezzo') or '').strip()
    cliente = (request.form.get('cliente') or '').strip()
    ddt_uscita = (request.form.get('ddt_uscita') or '').strip()
    consolidato = (request.form.get('consolidato') or '').strip()

    db = SessionLocal()
    try:
        query = db.query(Trasporto)

        # ✅ FILTRO MESE (compatibile se Trasporto.data è TEXT)
        if mese:
            try:
                year, month = mese.split("-")
                y = int(year)
                m = int(month)
                start = date(y, m, 1)
                end = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)

                # Converte 'data' (testo) in data: to_date(data, 'YYYY-MM-DD')
                data_as_date = func.to_date(Trasporto.data, 'YYYY-MM-DD')
                query = query.filter(data_as_date >= start, data_as_date < end)

            except Exception:
                # fallback
                query = query.filter(Trasporto.data.like(f"{mese}%"))

        if mezzo:
            query = query.filter(Trasporto.tipo_mezzo.ilike(f"%{mezzo}%"))
        if cliente:
            query = query.filter(Trasporto.cliente.ilike(f"%{cliente}%"))
        if ddt_uscita:
            query = query.filter(Trasporto.ddt_uscita.ilike(f"%{ddt_uscita}%"))
        if consolidato:
            query = query.filter(Trasporto.consolidato.ilike(f"%{consolidato}%"))

        dati = query.order_by(Trasporto.data.asc().nullslast(), Trasporto.id.asc()).all()

        # --- CREA EXCEL ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Trasporti"

        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_fill = PatternFill("solid", fgColor="D9E1F2")

        ws["A1"] = "REPORT TRASPORTI"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:H1")
        ws["A1"].alignment = center

        ws["A3"] = "Filtri:"
        ws["A3"].font = bold
        ws["B3"] = f"Mese={mese or 'Tutti'} | Cliente={cliente or 'Tutti'} | Mezzo={mezzo or 'Tutti'} | DDT={ddt_uscita or 'Tutti'} | Consolidato={consolidato or 'Tutti'}"
        ws.merge_cells("B3:H3")

        headers = ["Data", "Mezzo", "Cliente", "Trasportatore", "DDT", "Magazzino", "Consolidato", "Costo (€)"]
        start_row = 5
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center

        riga = start_row + 1
        totale = 0.0

        for t in dati:
            d_val = ""
            if t.data:
                try:
                    d_val = t.data.strftime("%Y-%m-%d")
                except:
                    d_val = str(t.data)[:10]

            costo_val = float(t.costo or 0.0)
            totale += costo_val

            ws.cell(riga, 1, d_val).alignment = center
            ws.cell(riga, 2, (t.tipo_mezzo or "")).alignment = left
            ws.cell(riga, 3, (t.cliente or "")).alignment = left
            ws.cell(riga, 4, (t.trasportatore or "")).alignment = left
            ws.cell(riga, 5, (t.ddt_uscita or "")).alignment = center
            ws.cell(riga, 6, (t.magazzino or "")).alignment = center
            ws.cell(riga, 7, (t.consolidato or "")).alignment = center

            c = ws.cell(riga, 8, costo_val)
            c.number_format = '#,##0.00'
            c.alignment = center

            riga += 1

        ws.cell(riga, 1, "TOTALE").font = bold
        ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=7)
        ws.cell(riga, 1).alignment = Alignment(horizontal="right", vertical="center")
        tot_cell = ws.cell(riga, 8, totale)
        tot_cell.font = bold
        tot_cell.number_format = '#,##0.00'
        tot_cell.alignment = center

        col_widths = [12, 16, 22, 22, 14, 14, 16, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        safe_mese = mese.replace("-", "_") if mese else "TUTTO"
        filename = f"Report_Trasporti_{safe_mese}.xlsx"

        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return f"Errore export Trasporti Excel: {e}", 500
    finally:
        db.close()

# --- GESTIONE LAVORAZIONI (ADMIN) ---
@app.route('/lavorazioni', methods=['GET', 'POST'])
@login_required
def lavorazioni():
    db = SessionLocal()

    # --- MODIFICA LAVORAZIONE ---
    if request.method == 'POST' and request.form.get('edit_lavorazione'):
        if session.get('role') != 'admin':
            flash("ACCESSO NEGATO: Solo Admin.", "danger")
            return redirect(url_for('lavorazioni'))

        try:
            lid = int(request.form.get('id') or 0)
            rec = db.query(Lavorazione).filter(Lavorazione.id == lid).first()
            if not rec:
                flash("Record non trovato.", "danger")
                return redirect(url_for('lavorazioni'))

            d_val = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
            rec.data = d_val
            rec.cliente = request.form.get('cliente')
            rec.descrizione = request.form.get('descrizione')
            rec.richiesta_di = request.form.get('richiesta_di')
            rec.seriali = request.form.get('seriali')
            rec.colli = int(request.form.get('colli') or 0)
            rec.pallet_forniti = int(request.form.get('pallet_forniti') or 0)
            rec.pallet_uscita = int(request.form.get('pallet_uscita') or 0)
            rec.ore_blue_collar = float(request.form.get('ore_blue_collar') or 0)
            rec.ore_white_collar = float(request.form.get('ore_white_collar') or 0)

            db.commit()
            flash("Picking modificato!", "success")
        except Exception as e:
            db.rollback()
            flash(f"Errore modifica: {e}", "danger")

        return redirect(url_for('lavorazioni'))

    # --- INSERIMENTO ---
    if request.method == 'POST' and request.form.get('add_lavorazione'):
        if session.get('role') != 'admin':
            flash("ACCESSO NEGATO: Solo Admin.", "danger")
            return redirect(url_for('lavorazioni'))

        try:
            d_val = datetime.strptime(request.form.get('data'), '%Y-%m-%d').date()
            nuovo = Lavorazione(
                data=d_val,
                cliente=request.form.get('cliente'),
                descrizione=request.form.get('descrizione'),
                richiesta_di=request.form.get('richiesta_di'),
                seriali=request.form.get('seriali'),
                colli=int(request.form.get('colli') or 0),
                pallet_forniti=int(request.form.get('pallet_forniti') or 0),
                pallet_uscita=int(request.form.get('pallet_uscita') or 0),
                ore_blue_collar=float(request.form.get('ore_blue_collar') or 0),
                ore_white_collar=float(request.form.get('ore_white_collar') or 0)
            )
            db.add(nuovo)
            db.commit()
            flash("Picking aggiunto!", "success")
        except Exception as e:
            db.rollback()
            flash(f"Errore inserimento: {e}", "danger")
        return redirect(url_for('lavorazioni'))

    # --- EDIT MODE (GET ?edit_id=) ---
    edit_id = request.args.get('edit_id')
    edit_row = None
    if edit_id and session.get('role') == 'admin':
        try:
            edit_row = db.query(Lavorazione).filter(Lavorazione.id == int(edit_id)).first()
        except:
            edit_row = None

    # --- VISUALIZZAZIONE ---
    dati = db.query(Lavorazione).order_by(Lavorazione.data.desc()).all()
    db.close()

    return render_template('lavorazioni.html', lavorazioni=dati, today=date.today(), edit_row=edit_row)



@app.route('/stampa_picking_pdf', methods=['POST'])
@login_required
def stampa_picking_pdf():
    if session.get('role') != 'admin':
        return "No Access", 403

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from sqlalchemy import func

    mese = (request.form.get('mese') or '').strip()       # es '2026-01'
    cliente = (request.form.get('cliente') or '').strip()

    db = SessionLocal()
    try:
        query = db.query(Lavorazione)

        # ✅ Qui trattiamo lavorazioni.data come TEXT -> la convertiamo in DATE (Postgres)
        data_as_date = func.to_date(func.left(Lavorazione.data, 10), 'YYYY-MM-DD')

        # ✅ FILTRO MESE (con range corretto)
        if mese:
            try:
                year, month = mese.split("-")
                y = int(year)
                m = int(month)

                start_str = f"{y:04d}-{m:02d}-01"
                if m == 12:
                    end_str = f"{y+1:04d}-01-01"
                else:
                    end_str = f"{y:04d}-{m+1:02d}-01"

                query = query.filter(
                    data_as_date >= func.to_date(start_str, 'YYYY-MM-DD'),
                    data_as_date < func.to_date(end_str, 'YYYY-MM-DD')
                )
            except Exception:
                # fallback se mese non è valido
                pass

        # ✅ FILTRO CLIENTE
        if cliente:
            query = query.filter(Lavorazione.cliente.ilike(f"%{cliente}%"))

        # ✅ ORDINAMENTO SICURO (per data convertita)
        rows = query.order_by(data_as_date.asc().nullslast(), Lavorazione.id.asc()).all()

        # --- CREA EXCEL ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Picking"

        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_fill = PatternFill("solid", fgColor="D9E1F2")

        ws["A1"] = "REPORT PICKING / LAVORAZIONI"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:J1")
        ws["A1"].alignment = center

        ws["A3"] = "Filtri:"
        ws["A3"].font = bold
        ws["B3"] = f"Mese={mese or 'Tutti'} | Cliente={cliente or 'Tutti'}"
        ws.merge_cells("B3:J3")

        headers = [
            "Data", "Cliente", "Descrizione", "Richiesta di", "Seriali",
            "Colli", "Pallet Entrati", "Pallet Usciti", "Ore Blue", "Ore White"
        ]

        start_row = 5
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center

        riga = start_row + 1

        # Totali
        t_colli = 0
        t_pin = 0
        t_pout = 0
        t_blue = 0.0
        t_white = 0.0

        for r in rows:
            d_str = (str(r.data)[:10] if r.data else "")

            colli = int(r.colli or 0)
            pin = int(r.pallet_forniti or 0)
            pout = int(r.pallet_uscita or 0)
            blue = float(r.ore_blue_collar or 0.0)
            white = float(r.ore_white_collar or 0.0)

            t_colli += colli
            t_pin += pin
            t_pout += pout
            t_blue += blue
            t_white += white

            ws.cell(riga, 1, d_str).alignment = center
            ws.cell(riga, 2, (r.cliente or "")).alignment = left
            ws.cell(riga, 3, (r.descrizione or "")).alignment = left
            ws.cell(riga, 4, (r.richiesta_di or "")).alignment = left
            ws.cell(riga, 5, (r.seriali or "")).alignment = left
            ws.cell(riga, 6, colli).alignment = center
            ws.cell(riga, 7, pin).alignment = center
            ws.cell(riga, 8, pout).alignment = center

            c9 = ws.cell(riga, 9, blue);  c9.number_format = '0.00'; c9.alignment = center
            c10 = ws.cell(riga, 10, white); c10.number_format = '0.00'; c10.alignment = center

            riga += 1

        # Riga Totali
        ws.cell(riga, 1, "TOTALI").font = bold
        ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=5)
        ws.cell(riga, 1).alignment = Alignment(horizontal="right", vertical="center")

        ws.cell(riga, 6, t_colli).font = bold
        ws.cell(riga, 7, t_pin).font = bold
        ws.cell(riga, 8, t_pout).font = bold

        tc9 = ws.cell(riga, 9, t_blue); tc9.font = bold; tc9.number_format = '0.00'; tc9.alignment = center
        tc10 = ws.cell(riga, 10, t_white); tc10.font = bold; tc10.number_format = '0.00'; tc10.alignment = center

        widths = [12, 18, 40, 20, 22, 10, 14, 14, 10, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A6"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        safe_mese = mese.replace("-", "_") if mese else "TUTTO"
        filename = f"Report_Picking_{safe_mese}.xlsx"

        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return f"Errore export Picking Excel: {e}", 500
    finally:
        db.close()


# --- NUOVO: EXPORT INVENTARIO EXCEL ---

@app.post('/report_inventario_excel')
@login_required
def report_inventario_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import io
    from datetime import datetime, date
    from collections import defaultdict

    # ✅ Client forzato sul proprio utente; Admin può scegliere cliente dal form
    if session.get('role') == 'client':
        cliente_rif = (current_user.id or '').strip()
    else:
        cliente_rif = (request.form.get('cliente_inventario') or '').strip()

    if not cliente_rif:
        return "Cliente mancante", 400

    # ✅ Data inventario FACOLTATIVA:
    # - se vuota => oggi
    # - se compilata => usa quella data come limite
    data_rif_str = (request.form.get('data_inventario') or '').strip()

    def parse_d(v):
        if not v:
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        s = str(v).strip().split(' ')[0][:10]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        return None

    if data_rif_str:
        d_limit = parse_d(data_rif_str)
        if not d_limit:
            return "Formato data inventario non valido", 400
    else:
        d_limit = date.today()

    # ✅ SOLO GALVANO TECNICA USA PEZZI (al posto dei colli)
    usa_pezzi = (cliente_rif.strip().upper() == "GALVANO TECNICA")

    # ✅ DUFERCO: nell'inventario serve vedere anche il Serial Number
    is_duferco = (cliente_rif.strip().upper() == "DUFERCO")

    db = SessionLocal()
    try:
        articoli = (
            db.query(Articolo)
            .filter(Articolo.cliente.ilike(f"%{cliente_rif}%"))
            .all()
        )

        # ✅ Aggregazione per CODICE ARTICOLO
        agg = defaultdict(lambda: {"descrizione": "", "serial_number": "", "entrata": 0, "uscita": 0}) if is_duferco else defaultdict(lambda: {"descrizione": "", "entrata": 0, "uscita": 0})

        for art in articoli:
            codice = (art.codice_articolo or "").strip()
            if not codice:
                continue

            serial = (getattr(art, 'serial_number', None) or '').strip()
            key = (codice, serial) if is_duferco else codice

            descr = (art.descrizione or "").strip()

            # ✅ quantità: PEZZI solo Galvano, COLLI per gli altri
            if usa_pezzi:
                q_raw = getattr(art, "pezzi", None)
                if q_raw is None:
                    q_raw = getattr(art, "pezzo", None)
            else:
                q_raw = getattr(art, "n_colli", None)

            try:
                qty = int(q_raw or 0)
            except Exception:
                qty = 0

            if qty <= 0:
                continue

            if descr and not agg[key]["descrizione"]:
                agg[key]["descrizione"] = descr

            if is_duferco and serial and not agg[key]["serial_number"]:
                agg[key]["serial_number"] = serial

            d_ing = parse_d(getattr(art, "data_ingresso", None))
            d_usc = parse_d(getattr(art, "data_uscita", None))

            # ✅ Logica inventario "al d_limit":
            # - Conta ENTRATA solo se ingresso <= d_limit
            # - Conta USCITA solo se uscita <= d_limit
            if d_ing and d_ing <= d_limit:
                agg[key]["entrata"] += qty

            if d_usc and d_usc <= d_limit:
                agg[key]["uscita"] += qty

        # ✅ righe finali
        righe = []
        for k in sorted(agg.keys()):
            data = agg[k]
            entrata = data.get("entrata", 0)
            uscita = data.get("uscita", 0)
            rimanenza = entrata - uscita

            # Se vuoi escludere righe a zero totale, lascia questo:
            if entrata == 0 and uscita == 0 and rimanenza == 0:
                continue

            if is_duferco:
                codice, serial = k
            else:
                codice, serial = k, ""

            righe.append({
                "codice": codice,
                "serial_number": serial,
                "descrizione": data.get("descrizione", ""),
                "entrata": entrata,
                "uscita": uscita,
                "rimanenza": rimanenza
            })

        # ============================
        # ✅ CREAZIONE EXCEL
        # ============================
        wb = Workbook()
        ws = wb.active
        ws.title = "INVENTARIO"

        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        oggi_str = datetime.now().strftime("%Y-%m-%d")
        data_limite_str = d_limit.strftime("%Y-%m-%d")
        tipo = "PEZZI" if usa_pezzi else "COLLI"

        ws["A1"] = "ELENCO ARTICOLI"
        ws["A2"] = f"Cliente: {cliente_rif}"
        ws["A3"] = f"Inventario basato su: {tipo}"
        ws["A4"] = f"Inventario al: {data_limite_str}"
        ws["A5"] = f"Generato il: {oggi_str}"

        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"].font = bold
        ws["A3"].font = bold
        ws["A4"].font = bold
        ws["A5"].font = bold

        headers = [
            "ID",
            "CODICE ARTICOLO",
        ]

        if is_duferco:
            headers.append("SERIAL NUMBER")

        headers += [
            "DESCRIZIONE",
            f"Q.TA ENTRATA ({tipo})",
            f"Q.TA USCITA ({tipo})",
            f"RIMANENZA ({tipo})"
        ]

        start_row = 7
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=c, value=h)
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border

        r = start_row + 1
        idx = 1
        for row in righe:
            values = [
                str(idx).zfill(3),
                row["codice"],
            ]

            if is_duferco:
                values.append(row.get("serial_number", ""))

            values += [
                row["descrizione"],
                row["entrata"],
                row["uscita"],
                row["rimanenza"]
            ]

            for c, v in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = left if c in (2, 3) else center
                cell.border = border

            r += 1
            idx += 1

        ws.freeze_panes = "A9"

        if r > start_row + 1:
            tab = Table(displayName="TabInventario", ref=f"A{start_row}:F{r-1}")
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True
            )
            ws.add_table(tab)

        # larghezze colonne
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 22

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"Inventario_{cliente_rif.replace(' ', '_')}_{data_limite_str}.xlsx"

        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    finally:
        db.close()

# =========================
# IMPORT EXCEL (con log)
# =========================

@app.route('/import_excel', methods=['GET', 'POST'])
@login_required
@require_admin
def import_excel():
    mappe = load_mappe()
    profiles = list(mappe.keys()) if mappe else []

    if request.method == 'GET':
        return render_template('import_excel.html', profiles=profiles)

    # POST logic
    profile_name = request.form.get('profile')
    if not profile_name or profile_name not in mappe:
        flash("Seleziona un profilo valido.", "warning")
        return redirect(request.url)

    if 'excel_file' not in request.files:
        return redirect(request.url)

    file = request.files['excel_file']
    if not file or file.filename == '':
        return redirect(request.url)

    db = SessionLocal()
    try:
        config = mappe[profile_name]
        header_row_idx = int(config.get('header_row', 1)) - 1
        column_map = config.get('column_map', {}) or {}

        import pandas as pd
        import numpy as np
        from datetime import datetime, date

        xls = pd.ExcelFile(file, engine="openpyxl")
        df = xls.parse(0, header=header_row_idx)

        # Mappa colonne (case-insensitive)
        df_cols_upper = {str(c).strip().upper(): c for c in df.columns}

        # ✅ HELPER DATA SUPER-ROBUSTO
        def to_date_db(val):
            """
            Ritorna data in formato YYYY-MM-DD oppure None.
            Gestisce:
            - datetime / pd.Timestamp / date
            - numpy.datetime64
            - seriale Excel (int/float)
            - stringhe (dd/mm/yyyy, dd/mm/yy, dd.mm.yyyy, ecc.)
            """
            if val is None:
                return None

            # NaN/NaT
            try:
                if pd.isna(val):
                    return None
            except Exception:
                pass

            # date/datetime/pandas timestamp
            if isinstance(val, (datetime, pd.Timestamp)):
                return val.strftime("%Y-%m-%d")
            if isinstance(val, date) and not isinstance(val, datetime):
                return val.strftime("%Y-%m-%d")

            # numpy datetime64
            if isinstance(val, np.datetime64):
                try:
                    dt = pd.to_datetime(val, errors="coerce")
                    if pd.isna(dt):
                        return None
                    return dt.strftime("%Y-%m-%d")
                except Exception:
                    return None

            # Seriali Excel (giorni da 1899-12-30)
            # Pandas a volte legge le date come float/int
            if isinstance(val, (int, float)) and val > 0:
                try:
                    dt = pd.to_datetime(val, unit="D", origin="1899-12-30", errors="coerce")
                    if pd.isna(dt):
                        return None
                    return dt.strftime("%Y-%m-%d")
                except Exception:
                    pass  # continua sotto a tentare come stringa

            # Stringhe
            s = str(val).strip()
            if not s:
                return None

            # prova formati comuni
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d", "%d-%m-%Y", "%d-%m-%y", "%d.%m.%Y", "%d.%m.%y"):
                try:
                    return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
                except Exception:
                    pass

            # fallback Pandas (dayfirst=True per Italia)
            try:
                dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
                if pd.isna(dt):
                    return None
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return None

        imported_count = 0

        for _, row in df.iterrows():
            if row.isnull().all():
                continue

            new_art = Articolo()
            has_data = False

            for excel_header, db_field in column_map.items():
                key = str(excel_header).strip().upper()
                col_name = df_cols_upper.get(key)
                if col_name is None:
                    continue

                val = row[col_name]

                # se vuoto, skip
                try:
                    if pd.isna(val) or str(val).strip() == "":
                        continue
                except Exception:
                    pass

                # Conversioni
                if db_field in ['larghezza', 'lunghezza', 'altezza', 'peso', 'm2', 'm3']:
                    try:
                        val = float(str(val).replace(',', '.'))
                    except Exception:
                        val = 0.0

                elif db_field in ['n_colli', 'pezzo']:
                    try:
                        val = int(float(str(val).replace(',', '.')))
                    except Exception:
                        val = 1

                elif db_field in ['data_ingresso', 'data_uscita']:
                    val = to_date_db(val)  # ✅ QUI ORA PRENDE ANCHE SERIALI EXCEL

                else:
                    val = str(val).strip()

                if val is not None and str(val).strip() != "":
                    setattr(new_art, db_field, val)
                    has_data = True

            if has_data:
                # Calcoli automatici se mancano
                try:
                    if not new_art.m2 or float(new_art.m2) == 0:
                        l = float(new_art.lunghezza or 0)
                        w = float(new_art.larghezza or 0)
                        h = float(new_art.altezza or 0)
                        c = int(new_art.n_colli or 1)

                        if l > 0 and w > 0:
                            new_art.m2 = round(l * w * c, 3)
                            new_art.m3 = round(l * w * (h if h > 0 else 0) * c, 3)
                except Exception:
                    pass

                db.add(new_art)
                imported_count += 1

        db.commit()
        flash(f"{imported_count} articoli importati con successo.", "success")
        return redirect(url_for('giacenze'))

    except Exception as e:
        db.rollback()
        flash(f"Errore import: {e}", "danger")
        return redirect(request.url)
    finally:
        db.close()


def get_all_fields_map():
    return {
        'codice_articolo': 'Codice Articolo', 'pezzo': 'Pezzi','lotto':'Lotto',
        'descrizione': 'Descrizione', 'cliente': 'Cliente','ordine':'Ordine',
        'protocollo': 'Protocollo', 'peso': 'Peso (Kg)',
        'n_colli': 'N° Colli', 'posizione': 'Posizione', 'stato': 'Stato',
        'n_arrivo': 'N° Arrivo', 'buono_n': 'Buono N°',
        'fornitore': 'Fornitore', 'magazzino': 'Magazzino',
        'data_ingresso': 'Data Ingresso', 'data_uscita': 'Data Uscita',
        'n_ddt_ingresso': 'N° DDT Ingresso', 'n_ddt_uscita': 'N° DDT Uscita',
        'larghezza': 'Larghezza (m)', 'lunghezza': 'Lunghezza (m)',
        'altezza': 'Altezza (m)', 'serial_number': 'Serial Number',
        'ns_rif': 'NS Rif', 'mezzi_in_uscita': 'Mezzi in Uscita', 'note': 'Note'
    }

# --- ROUTE IMPORT PDF (PROTETTA ADMIN) ---
@app.route('/import_pdf', methods=['GET', 'POST'])
@login_required
@require_admin
def import_pdf():
    # PROTEZIONE ADMIN
    if session.get('role') != 'admin':
        flash("Accesso negato: Funzione riservata agli amministratori.", "danger")
        return redirect(url_for('giacenze'))

    if request.method == 'POST':
        if 'file' not in request.files: return redirect(request.url)
        f = request.files['file']
        if f.filename:
            temp_path = os.path.join(DOCS_DIR, f"temp_{uuid.uuid4().hex}.pdf")
            f.save(temp_path)
            try:
                meta, rows = extract_data_from_ddt_pdf(temp_path)
                # Pulisce file temp
                if os.path.exists(temp_path): os.remove(temp_path)
                return render_template('import_pdf.html', meta=meta, rows=rows)
            except Exception as e:
                flash(f"Errore PDF: {e}", "danger")
                return redirect(url_for('giacenze'))
                
    return render_template('import_pdf.html', meta={}, rows=[])

@app.route('/save_pdf_import', methods=['POST'])
@login_required
@require_admin
def save_pdf_import():
    if session.get('role') != 'admin':
        return "Accesso Negato", 403

    db = SessionLocal()
    try:
        codici = request.form.getlist('codice[]')
        descrizioni = request.form.getlist('descrizione[]')
        colli_list = request.form.getlist('colli[]')
        pezzi_list = request.form.getlist('pezzi[]')

        c = 0
        for i in range(len(codici)):
            codice = (codici[i] or "").strip()
            descr = (descrizioni[i] or "").strip()
            if not codice and not descr:
                continue

            art = Articolo()

            # testata
            art.cliente = request.form.get('cliente')
            art.fornitore = request.form.get('fornitore')
            art.commessa = request.form.get('commessa')
            art.n_ddt_ingresso = request.form.get('n_ddt')
            art.data_ingresso = parse_date_ui(request.form.get('data_ingresso'))
            art.stato = "DOGANALE"

            # riga
            art.codice_articolo = codice
            art.descrizione = descr
            art.n_colli = to_int_eu(colli_list[i] if i < len(colli_list) else 1)
            # pezzo è String nel modello, ma in UI lo vuoi numerico: salviamo come stringa “pulita”
            pz = pezzi_list[i] if i < len(pezzi_list) else ""
            art.pezzo = str(pz).strip()

            db.add(art)
            c += 1

        db.commit()
        flash(f"Importati {c} articoli.", "success")
        return redirect(url_for('giacenze'))
    except Exception as e:
        db.rollback()
        flash(f"Errore import PDF: {e}", "danger")
        return redirect(url_for('import_pdf'))
    finally:
        db.close()



# --- EXPORTAZIONE EXCEL ---
@app.get('/export_excel')
@login_required
@require_admin
def export_excel():
    db = SessionLocal()
    df = pd.read_sql(db.query(Articolo).statement, db.bind)
    bio = io.BytesIO()
    df.to_excel(bio, index=False, engine='openpyxl')
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name='Giacenze_Totali.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export_client', methods=['GET', 'POST'])
@login_required
def export_client():
    db = SessionLocal()
    clienti = [c[0] for c in db.query(Articolo.cliente).distinct().filter(Articolo.cliente != None, Articolo.cliente != '').order_by(Articolo.cliente).all()]
    if session.get('role') == 'client':
        clienti = [(current_user.id or '').strip()]
    
    if request.method == 'POST':
        cliente = request.form.get('cliente')
        if session.get('role') == 'client':
            cliente = (current_user.id or '').strip()
        if not cliente:
            flash("Seleziona un cliente.", "warning")
            return redirect(request.url)
        
        rows = db.query(Articolo).filter(Articolo.cliente == cliente).all()
        if not rows:
            flash(f"Nessun articolo trovato per il cliente {cliente}.", "info")
            return redirect(request.url)
        
        df = pd.DataFrame([vars(r) for r in rows])
        bio = io.BytesIO()
        df.to_excel(bio, index=False, engine='openpyxl')
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name=f"Giacenze_{cliente}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('export_client.html', clienti=clienti)
    
# ==============================================================================
#  FUNZIONE INVIA EMAIL (CORRETTA CON FIRMA COMPLETA E LOGO)
# ==============================================================================

@app.route('/invia_email', methods=['GET', 'POST'])
@login_required
@require_admin
def invia_email():
    from email.header import Header
    from email.mime.image import MIMEImage
    import html

    # =========================
    # Helper: Riepilogo Merci (tabella in email)
    # =========================
    def _build_riepilogo_schema_html(rows):
        def esc(x):
            return html.escape("" if x is None else str(x))

        def fnum(x, nd=2):
            try:
                return f"{float(x):.{nd}f}"
            except:
                return ""

        total_colli = 0
        total_peso = 0.0
        trs = []

        for r in rows:
            try:
                total_colli += int(r.n_colli or 0)
            except:
                pass
            try:
                total_peso += float(r.peso or 0)
            except:
                pass

            misure = f"{fnum(r.larghezza,2)} × {fnum(r.lunghezza,2)} × {fnum(r.altezza,2)}"

            trs.append(f"""
            <tr>
              <td style="border:1px solid #ddd;padding:6px;">{esc(r.commessa)}</td>
              <td style="border:1px solid #ddd;padding:6px;">{esc(r.ordine)}</td>
              <td style="border:1px solid #ddd;padding:6px;">{esc(misure)}</td>
              <td style="border:1px solid #ddd;padding:6px;">{esc(r.cliente)}</td>
              <td style="border:1px solid #ddd;padding:6px;">{esc(r.fornitore)}</td>
              <td style="border:1px solid #ddd;padding:6px;text-align:right;">{fnum(r.peso,2)}</td>
              <td style="border:1px solid #ddd;padding:6px;">{esc(r.descrizione)}</td>
              <td style="border:1px solid #ddd;padding:6px;">{esc(r.codice_articolo)}</td>
              <td style="border:1px solid #ddd;padding:6px;text-align:right;">{esc(r.n_colli)}</td>
              <td style="border:1px solid #ddd;padding:6px;">{esc(r.n_arrivo)}</td>
            </tr>
            """)

        return f"""
        <div style="margin:12px 0 20px 0; font-family: Arial, sans-serif;">
          <b>Riepilogo merce selezionata</b>

          <table style="border-collapse:collapse;width:100%;font-size:12px;margin-top:6px;">
            <thead>
              <tr style="background:#f2f2f2;">
                <th style="border:1px solid #ddd;padding:6px;">Commessa</th>
                <th style="border:1px solid #ddd;padding:6px;">Ordine</th>
                <th style="border:1px solid #ddd;padding:6px;">Misure pallet (L×P×H)</th>
                <th style="border:1px solid #ddd;padding:6px;">Cliente</th>
                <th style="border:1px solid #ddd;padding:6px;">Fornitore</th>
                <th style="border:1px solid #ddd;padding:6px;text-align:right;">Peso (kg)</th>
                <th style="border:1px solid #ddd;padding:6px;">Descrizione</th>
                <th style="border:1px solid #ddd;padding:6px;">Codice Articolo</th>
                <th style="border:1px solid #ddd;padding:6px;text-align:right;">Colli</th>
                <th style="border:1px solid #ddd;padding:6px;">N. Arrivo</th>
              </tr>
            </thead>
            <tbody>
              {''.join(trs)}
            </tbody>
          </table>

          <div style="margin-top:8px;font-size:12px;">
            <b>Totali:</b> Colli = {total_colli} | Peso = {total_peso:.2f} kg
          </div>
        </div>
        """

    # =========================
    # Firma completa + Avviso Importante (come da testo)
    # =========================
    firma_completa_html = """
    <div style="font-size:12px;color:#444;line-height:1.4;">
      <div style="margin-top:10px;">
        <b>Numero Ufficio :</b> 010265995<br>
        <b>Numero Fax:</b> 010 4550943<br>
        <b>Mobili:</b><br><br>

        Sig.  Tazio Marcellino&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; +39 334 6892992<br>
        Sig.ra Alessia Moncalvo&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; +39 324 9255537<br>
        Sig. Giorgio Cabella&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; +39 338 7255224<br>
        Sig.  Hugo Esviza&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; +39 327 4573767<br><br>

        <i>a simple but ingenious company ®</i><br><br>

        <b>INDIRIZZO CONTABILITA':</b> <a href="mailto:contabilita@camarsrl.net">contabilita@camarsrl.net</a><br><br>

        HEAD OFFICE: Via Balleydier 52r – 16149 GENOVA<br>
        BRANCH OFFICE: La Spezia - Savona - Vado Ligure - Civitavecchia - Marina Di Carrara - Venezia<br><br>

        Tutte le parti accettano il presente documento e stabiliscono che per ogni eventuale e futura controversia derivante dal presente accordo, o connesse allo stesso, è competente il Tribunale di Roma .<br><br>

        Si ritiene accettato con la conferma del trasporto o la conferma della vendita .<br><br>

        All the parts agree upon the present document and establish that for any possible future controversy related to the present agreement, or connected to it, the Tribunal of Rome is in charge.<br>
        This is considered as accepted once the transport or the sale has been confermed
      </div>

      <hr style="border:0;border-top:1px solid #ccc;margin:15px 0;">

      <p style="font-size:10px;color:#777;text-align:justify;margin-top:10px;">
      <b>AVVISO IMPORTANTE.</b>Le informazioni contenute nella presente comunicazione e i relativi allegati possono essere riservate e sono, comunque, destinate esclusivamente alle persone o alla Società sopraindicati. La comunicazione, diffusione, distribuzione e/o copiatura del documento trasmesso nonché qualsiasi forma di trattamento dei dati ivi contenuti da parte di qualsiasi soggetto diverso dal destinatario è proibita, sia ai sensi dell’art. 616 c.p., che ai sensi del D. Lgs. n. 196/2003, ed in ogni caso espressamente inibita. Le informazioni e tutte le indicazioni, dati, contenuti in questo messaggio hanno una scadenza decennale. Se avete ricevuto questo messaggio per errore, vi preghiamo di distruggerlo e di informarci immediatamente per telefono allo 010 265995 o inviando un messaggio. L’operazione eseguita per vostro conto, segue l’accordo/le tariffe stabilite appositamente, fa parte di un appalto di servizi in esclusiva per le operazioni marittime della vostra azienda. La sopracitata operazione, che sarà effettuata con il massimo dell’attenzione e più velocemente possibile, viene eseguita tramite Autorizzazione Doganale, di Polizia ,o di Capitaneria, ed è riconducibile e discrezionale solo da parte dell’Autorità Ministeriale/Statale, pertanto la nostra azienda si manleva da qualsiasi responsabilità relativa all’esito della stessa. Le disposizioni di cui sopra si ritengono accettate dalle controparti, dal momento dell’incarico e dello svolgimento del lavoro sopra menzionato nella email. Questo messaggio, con gli eventuali allegati e informazioni contiene documentazione, dati, notizie, nomi, riservate esclusivamente per fini lavorativi al destinatario inteso come azienda, e alla sua direzione. La nostra azienda non accetta nessun tipo di addebito per ritardi o errori, deficienze o negligenze, nella compilazione o nell’esecuzione, assistenza della documentazione richiesta o fornita.La scrivente agisce come intermediario tra IMPORTANTE. Mandato di trasporto e assicurativo: eseguiamo l’ordine di trasporto e assicuriamo la merce al valore dichiarato. La risposta a questa email è da considerare come mandato assicurativo( quello assicurativo se esplicitamente manifestato dal cliente) e di trasporto a tutti gli effetti.Vi preghiamo di avvisarci nel caso di imprevisti.Comunichiamo che il cambio della data di consegna da noi indicata, non deve essere soggetta a richieste danni o spese. Comunichiamo, inoltre, che dall’uscita dei varchi doganali sino a Vs destinazione, le spese e i costi derivanti da eventuali blocchi traffico, soste, verbali, sanzioni, incidenti non sono a noi imputabili.Se il valore della merce trasportata non è stato dichiarato, il cliente anche per conto dei propri mandatari rinuncia a far valere nei confronti della società e del vettore qualsiasi credito per danni o perdita delle merci in misura superiore al valore indicato dal decreto riportato. Si obbliga a tenere indenne e manlevare la società e il vettore a fronte di qualsiasi richiesta di risarcimento da parte di terzi a fronte di perdite delle merci in misura superiore al valore indicato dal decreto sotto riportato.Il trasporto oggetto della presente prenotazione è disciplinato dalle disposizioni del decreto legislativo 21.11.2005 n.286. Tali disposizioni, tra l’altro, prevedono a carico del committente, caricatore, e proprietario delle merci responsabilità e sanzioni in relazione a violazione delle disposizioni in materia di sicurezza della circolazione quali quelle relative alla massa limite e alla sistemazione del carico sui veicoli. Il cliente garantisce l’esattezza e la completezza delle informazioni fornite alla società in merito alle merci oggetto della prenotazione, nonché, laddove vi preveda l’accuratezza e l’idoneità della sistemazione del carico sui veicoli nel rispetto delle norme descritte si terrà indenne e manleverà la società e il vettore da quest’ultima incaricato per suo conto a fronte di qualsiasi sanzione e responsabilità che dovesse derivare dall’inesattezza incompletezza o inidoneità delle predette informazioni e sistemazioni.È a conoscenza e quindi manleva da qualsiasi danno o addebito la scrivente, nel caso che l’ordine di trasporto venga disdetto da quest’ultima per motivi logistici.La nostra azienda si occupa d’intermediazione nel campo della logistica e trasporti.Eseguiamo operazioni solo ed esclusivamente per Vs conto senza alcuna responsabilità civile, economica, legale.Le disposizioni di cui sopra si ritengono accettate dal momento dell’incarico.
      </p>
    </div>
    """

    # =========================
    # GET
    # =========================
    if request.method == 'GET':
        selected_ids = request.args.getlist('ids')
        return render_template('invia_email.html', selected_ids=",".join(selected_ids), email_groups=load_rubrica_email().get('gruppi', {}))

    # =========================
    # POST
    # =========================
    selected_ids = request.form.get('selected_ids', '')
    ids_list = [int(i) for i in selected_ids.split(',') if i.isdigit()]

    destinatari = [
        e.strip() for e in request.form.get('destinatario', '').replace(";", ",").split(",") if e.strip()
    ]

    if not destinatari:
        flash("Inserire almeno un destinatario valido", "danger")
        return redirect(url_for('giacenze'))

    oggetto = request.form.get('oggetto')
    messaggio = request.form.get('messaggio') or ""
    genera_ddt = 'genera_ddt' in request.form
    allega_file = 'allega_file' in request.form
    allegati_extra = request.files.getlist('allegati_extra')

    SMTP_SERVER = os.environ.get("MAIL_SERVER") or os.environ.get("SMTP_SERVER", "smtp.gmail.com")
    SMTP_PORT = int(os.environ.get("MAIL_PORT") or os.environ.get("SMTP_PORT", 587))
    SMTP_USER = os.environ.get("MAIL_USERNAME") or os.environ.get("SMTP_USER", "")
    SMTP_PASS = os.environ.get("MAIL_PASSWORD") or os.environ.get("SMTP_PASS", "")

    if not SMTP_USER or not SMTP_PASS:
        flash("Configurazione email mancante.", "warning")
        return redirect(url_for('giacenze'))

    try:
        riepilogo_html = ""
        if genera_ddt and ids_list:
            db = SessionLocal()
            try:
                rows = db.query(Articolo).filter(Articolo.id_articolo.in_(ids_list)).all()
                if rows:
                    riepilogo_html = _build_riepilogo_schema_html(rows)
            finally:
                db.close()

        msg_root = MIMEMultipart('related')
        msg_root['From'] = SMTP_USER
        msg_root['To'] = ", ".join(destinatari)
        msg_root['Subject'] = Header(oggetto, 'utf-8')

        msg_alt = MIMEMultipart('alternative')
        msg_root.attach(msg_alt)

        msg_alt.attach(MIMEText(messaggio, 'plain', 'utf-8'))

        html_body = f"""
        <html>
          <head><meta http-equiv="Content-Type" content="text/html; charset=utf-8"></head>
          <body style="font-family:Arial, sans-serif; font-size:14px; color:#333;">
            <div style="margin-bottom:18px;">{html.escape(messaggio).replace(chr(10), '<br>')}</div>
            {riepilogo_html}

            <div style="margin: 16px 0 12px 0;">
              <img src="cid:logo_camar" alt="Camar S.r.l." style="height:65px; width:auto; display:block;">
            </div>

            {firma_completa_html}
          </body>
        </html>
        """
        msg_alt.attach(MIMEText(html_body, 'html', 'utf-8'))

        # ✅ Allega LOGO inline (CID)
        possible_logos = ["logo camar.jpg", "logo_camar.jpg", "logo.jpg"]
        logo_found = False
        for name in possible_logos:
            logo_path = os.path.join(app.root_path, "static", name)
            if os.path.exists(logo_path):
                with open(logo_path, "rb") as f:
                    img = MIMEImage(f.read())
                img.add_header('Content-ID', '<logo_camar>')
                img.add_header('Content-Disposition', 'inline', filename='logo_camar.jpg')
                msg_root.attach(img)
                logo_found = True
                break

        if not logo_found:
            print("⚠️ Logo non trovato in static: l'email partirà senza logo.")

        # ✅ Allegati esistenti (foto/pdf articoli)
        if allega_file and ids_list:
            db = SessionLocal()
            try:
                rows = db.query(Articolo).filter(Articolo.id_articolo.in_(ids_list)).all()
                for r in rows:
                    for att in r.attachments:
                        fname = att.filename
                        path = (DOCS_DIR if att.kind == 'doc' else PHOTOS_DIR) / fname
                        if not path.exists():
                            from urllib.parse import unquote
                            path = (DOCS_DIR if att.kind == 'doc' else PHOTOS_DIR) / unquote(fname)

                        if path.exists():
                            with open(path, "rb") as f:
                                part = MIMEBase('application', "octet-stream")
                                part.set_payload(f.read())
                            encoders.encode_base64(part)
                            part.add_header('Content-Disposition', f'attachment; filename="{fname}"')
                            msg_root.attach(part)
            finally:
                db.close()

        # ✅ Allegati extra
        for file in allegati_extra:
            if file and file.filename:
                part = MIMEBase('application', "octet-stream")
                part.set_payload(file.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{secure_filename(file.filename)}"')
                msg_root.attach(part)

        # ✅ Invio SMTP
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg_root, from_addr=SMTP_USER, to_addrs=destinatari)
        server.quit()

        flash("Email inviata correttamente", "success")

    except Exception as e:
        print(f"DEBUG EMAIL EXCEPTION: {e}")
        import traceback
        traceback.print_exc()
        flash(str(e), "danger")

    return redirect(url_for('giacenze'))


# --- FUNZIONE UPLOAD FILE MULTIPLI (CORRETTA PER EDIT_RECORD) ---
@app.route('/upload/<int:id_articolo>', methods=['POST'])
@login_required
def upload_file(id_articolo):
    # 1. Controllo Permessi
    if session.get('role') != 'admin':
        flash("Solo Admin può caricare file", "danger")
        return redirect(url_for('edit_record', id_articolo=id_articolo))

    # 2. Recupera LISTA di file
    files = request.files.getlist('file')
    
    if not files or all(f.filename == '' for f in files):
        flash("Nessun file selezionato", "warning")
        return redirect(url_for('edit_record', id_articolo=id_articolo))

    db = SessionLocal()
    count = 0
    try:
        from werkzeug.utils import secure_filename
        
        for file in files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                
                # Crea nome univoco: ID_UUID_Nome
                unique_name = f"{id_articolo}_{uuid.uuid4().hex[:6]}_{filename}"
                
                ext = filename.rsplit('.', 1)[-1].lower()
                if ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
                    kind = 'photo'
                    save_path = PHOTOS_DIR / unique_name
                else:
                    kind = 'doc'
                    save_path = DOCS_DIR / unique_name
                    
                file.save(str(save_path))

                # Salva nel DB
                att = Attachment(
                    articolo_id=id_articolo,
                    filename=unique_name,
                    kind=kind
                )
                db.add(att)
                count += 1
        
        db.commit()
        if count > 0:
            flash(f"Caricati {count} file correttamente!", "success")
        else:
            flash("Nessun file valido caricato.", "warning")
        
    except Exception as e:
        db.rollback()
        print(f"ERRORE UPLOAD: {e}") 
        flash(f"Errore caricamento: {e}", "danger")
    finally:
        db.close()

    # --- MODIFICA QUI: RITORNA A EDIT_RECORD ---
    return redirect(url_for('edit_record', id_articolo=id_articolo))
    
    

@app.route('/delete_file/<int:id_file>')
@login_required
def delete_file(id_file):
    db = SessionLocal()
    att = db.query(Attachment).get(id_file)
    if att:
        id_art = att.articolo_id
        path = (DOCS_DIR if att.kind=='doc' else PHOTOS_DIR) / att.filename
        try:
            if path.exists(): os.remove(path)
        except: pass
        db.delete(att)
        db.commit()
        db.close()
        return redirect(url_for('edit_record', id_articolo=id_art))
    db.close()
    return redirect(url_for('giacenze'))



# --- FIX VISUALIZZAZIONE ALLEGATI ---
from urllib.parse import unquote
import os

@app.route('/serve_file/<path:filename>')
@login_required
def serve_uploaded_file(filename):
    # 1. Decodifica standard (es. %20 -> spazio)
    decoded_name = unquote(filename)
    
    # 2. Lista di possibili nomi da cercare (Originale, Decodificato, Con Underscore)
    candidates = [
        filename,                   
        decoded_name,               
        filename.replace(' ', '_'), 
        decoded_name.replace(' ', '_'),
        secure_filename(decoded_name) # Prova anche la versione "sicura"
    ]
    
    # 3. Cerca in entrambe le cartelle (Foto e Documenti)
    # Usa os.walk o listdir se necessario, ma qui proviamo i path diretti
    for folder in [PHOTOS_DIR, DOCS_DIR]:
        for name in candidates:
            p = folder / name
            if p.exists():
                return send_file(p)
            
            # Tentativo case-insensitive (per sistemi Linux sensibili alle maiuscole)
            try:
                for existing_file in os.listdir(folder):
                    if existing_file.lower() == name.lower():
                        return send_file(folder / existing_file)
            except: pass

    # Se arriviamo qui, il file non c'è. Stampa debug nei log di Render.
    print(f"DEBUG: File '{filename}' non trovato. Cercato candidati: {candidates}")
    return f"File '{decoded_name}' non trovato sul server (potrebbe essere stato cancellato dal riavvio di Render).", 404
    
# --- GESTIONE ARTICOLI (CRUD) ---
# ========================================================
# 8. CRUD (NUOVO / MODIFICA)
# ========================================================

@app.route('/new', methods=['GET', 'POST'])
@login_required
@require_admin
def nuovo_articolo():
    # 1. Controllo permessi
    if session.get('role') != 'admin':
        flash("Accesso negato: Solo Admin.", "danger")
        return redirect(url_for('giacenze'))

    if request.method == 'POST':
        db = SessionLocal()
        try:
            # --- RECUPERO DATI FORM ---
            # Determina quante righe creare in base al numero di colli
            n_colli_input = to_int_eu(request.form.get('n_colli'))
            if n_colli_input < 1: n_colli_input = 1
            
            created_articles = []

            # --- CICLO DI CREAZIONE (Crea N righe identiche) ---
            for _ in range(n_colli_input):
                art = Articolo()
                
                # Popola i campi testuali
                art.codice_articolo = request.form.get('codice_articolo')
                art.descrizione = request.form.get('descrizione')
                # Cliente: per i client e' bloccato sul proprio utente
            if session.get('role') == 'client':
                art.cliente = current_user.id
            else:
                art.cliente = request.form.get('cliente')
                art.fornitore = request.form.get('fornitore')
                art.commessa = request.form.get('commessa')
                art.ordine = request.form.get('ordine')
                art.protocollo = request.form.get('protocollo')
                art.buono_n = request.form.get('buono_n')
                art.n_arrivo = request.form.get('n_arrivo')
                art.magazzino = request.form.get('magazzino')
                art.posizione = request.form.get('posizione')
                art.stato = request.form.get('stato')
                art.note = request.form.get('note')
                art.serial_number = request.form.get('serial_number')
                art.mezzi_in_uscita = request.form.get('mezzi_in_uscita')
                art.lotto = request.form.get('lotto')

                # Date
                art.data_ingresso = parse_date_ui(request.form.get('data_ingresso'))
                art.data_uscita = parse_date_ui(request.form.get('data_uscita'))
                art.n_ddt_ingresso = request.form.get('n_ddt_ingresso')
                art.n_ddt_uscita = request.form.get('n_ddt_uscita')
                
                # Numeri
                art.pezzo = request.form.get('pezzo')
                art.n_colli = 1  # FORZA 1 COLLO PER OGNI RIGA CREATA
                art.peso = to_float_eu(request.form.get('peso'))
                art.lunghezza = to_float_eu(request.form.get('lunghezza'))
                art.larghezza = to_float_eu(request.form.get('larghezza'))
                art.altezza = to_float_eu(request.form.get('altezza'))
                
                # Calcolo M2/M3 (su 1 collo)
                art.m2, art.m3 = calc_m2_m3(art.lunghezza, art.larghezza, art.altezza, 1)

                db.add(art)
                created_articles.append(art)

            # Salva tutto per ottenere gli ID
            db.commit() 
            
            # --- GESTIONE ALLEGATI (Duplicazione su tutte le righe) ---
            # getlist permette di prendere PIÙ file selezionati
            files = request.files.getlist('new_files') 
            valid_files = [f for f in files if f and f.filename]
            
            if valid_files:
                import shutil
                from werkzeug.utils import secure_filename
                
                for file in valid_files:
                    fname = secure_filename(file.filename)
                    ext = fname.rsplit('.', 1)[-1].lower()
                    kind = 'photo' if ext in ['jpg', 'jpeg', 'png', 'webp'] else 'doc'
                    folder = PHOTOS_DIR if kind == 'photo' else DOCS_DIR
                    
                    # 1. Salva il file fisico per il PRIMO articolo
                    first_art = created_articles[0]
                    first_final_name = f"{first_art.id_articolo}_{fname}"
                    src_path = folder / first_final_name
                    
                    # Importante: seek(0) se il file è stato letto parzialmente
                    file.seek(0) 
                    file.save(str(src_path))
                    
                    # Collega al primo articolo nel DB
                    db.add(Attachment(articolo_id=first_art.id_articolo, filename=first_final_name, kind=kind))
                    
                    # 2. Copia il file fisico e il record DB per gli ALTRI articoli
                    for other_art in created_articles[1:]:
                        other_final_name = f"{other_art.id_articolo}_{fname}"
                        dst_path = folder / other_final_name
                        
                        try:
                            # Copia fisica del file
                            shutil.copy2(src_path, dst_path)
                            # Nuovo record nel DB
                            db.add(Attachment(articolo_id=other_art.id_articolo, filename=other_final_name, kind=kind))
                        except Exception as e:
                            print(f"Errore copia file per ID {other_art.id_articolo}: {e}")

                db.commit()

            flash(f"Operazione completata: Creati {len(created_articles)} articoli distinti con allegati.", "success")
            
            # Se ne abbiamo creato uno solo, vai alla modifica, altrimenti torna alla lista
            if len(created_articles) == 1:
                # Reindirizza alla funzione edit_articolo corretta
                return redirect(url_for('edit_articolo', id=created_articles[0].id_articolo))
            else:
                return redirect(url_for('giacenze'))
            
        except Exception as e:
            db.rollback()
            flash(f"Errore creazione: {e}", "danger")
            return redirect(url_for('giacenze'))
        finally:
            db.close()

    # GET: Mostra form vuoto
    dummy_art = Articolo() 
    dummy_art.data_ingresso = date.today().strftime("%d/%m/%Y")
    return render_template('edit.html', row=dummy_art)
    
# 1. ELIMINA ARTICOLO (Per la pagina Magazzino)
@app.route('/delete_articolo/<int:id>')
@login_required
@require_admin
def delete_articolo(id):
    if session.get('role') != 'admin':
        flash("Accesso Negato: Solo Admin può eliminare.", "danger")
        return redirect(url_for('giacenze'))
        
    db = SessionLocal()
    try:
        record = db.query(Articolo).get(id)
        if record:
            db.delete(record)
            db.commit()
            flash("Articolo eliminato.", "success")
        else:
            flash("Articolo non trovato.", "warning")
    except Exception as e:
        db.rollback()
        flash(f"Errore eliminazione: {e}", "danger")
    finally:
        db.close()
    return redirect(url_for('giacenze'))

# --- DUPLICAZIONE SINGOLA (Quella che mancava e dava errore) ---

# --- DUPLICAZIONE SINGOLA (COMPLETA: copia TUTTI i campi tranne ID) ---
@app.route('/duplica_articolo/<int:id_articolo>')
@login_required
def duplica_articolo(id_articolo):
    if session.get('role') != 'admin':
        flash('Accesso negato: Solo Admin.', 'danger')
        return redirect(url_for('giacenze'))

    db = SessionLocal()
    try:
        originale = db.query(Articolo).filter(Articolo.id_articolo == id_articolo).first()

        if not originale:
            flash("Articolo non trovato", "danger")
            return redirect(url_for('giacenze'))

        # ✅ Copia tutti i campi della tabella Articolo eccetto la PK
        data_copy = {}
        for col in Articolo.__table__.columns:
            if col.name == 'id_articolo':
                continue
            data_copy[col.name] = getattr(originale, col.name)

        nuovo = Articolo(**data_copy)

        # ✅ Modifiche volute sulla copia
        nuovo.note = f"Copia di ID {originale.id_articolo}"
        nuovo.data_ingresso = date.today()

        db.add(nuovo)
        db.commit()

        flash("Articolo duplicato con successo!", "success")

    except Exception as e:
        db.rollback()
        flash(f"Errore duplicazione: {e}", "danger")

    finally:
        db.close()

    return redirect(url_for('giacenze'))


# ==============================================================================
#  1. MODIFICA ARTICOLO (Solo per TABELLA GIACENZE)
# ==============================================================================
# --- MODIFICA ARTICOLO SINGOLO (UNICA FUNZIONE CORRETTA) ---
@app.route('/edit_articolo/<int:id>', methods=['GET', 'POST'])
@login_required
@require_admin
def edit_articolo(id):
    db = SessionLocal()
    try:
        art = db.query(Articolo).get(id)
        if not art:
            flash("Articolo non trovato", "danger")
            return redirect(url_for('giacenze'))

        # Sicurezza HARD: un client non puo' accedere/modificare articoli di altri clienti
        if session.get('role') == 'client':
            art_cliente = (art.cliente or '').strip().upper()
            user_key = (current_user.id or '').strip().upper()
            art_norm = re.sub(r'[^A-Z0-9]+', '', art_cliente)
            user_norm = re.sub(r'[^A-Z0-9]+', '', user_key)
            # Se non contiene l'identificativo cliente -> blocca
            if user_norm not in art_norm:
                abort(403)

        if request.method == 'POST':
            # 1. Recupera Colli (per eventuale split)
            colli_input = to_int_eu(request.form.get('n_colli'))
            if colli_input < 1: colli_input = 1

            # 2. Aggiorna tutti i campi
            art.codice_articolo = request.form.get('codice_articolo')
            art.descrizione = request.form.get('descrizione')
            # Cliente: per i client e' bloccato sul proprio utente
            if session.get('role') == 'client':
                art.cliente = current_user.id
            else:
                art.cliente = request.form.get('cliente')
            art.fornitore = request.form.get('fornitore')
            art.commessa = request.form.get('commessa')
            art.ordine = request.form.get('ordine')
            art.protocollo = request.form.get('protocollo')
            art.buono_n = request.form.get('buono_n')
            art.n_arrivo = request.form.get('n_arrivo')
            art.magazzino = request.form.get('magazzino')
            art.posizione = request.form.get('posizione')
            art.stato = request.form.get('stato')
            art.note = request.form.get('note')
            art.serial_number = request.form.get('serial_number')
            art.mezzi_in_uscita = request.form.get('mezzi_in_uscita')
            art.lotto = request.form.get('lotto')
            art.ns_rif = request.form.get('ns_rif')

            # Date
            art.data_ingresso = parse_date_ui(request.form.get('data_ingresso'))
            art.data_uscita = parse_date_ui(request.form.get('data_uscita'))
            art.n_ddt_ingresso = request.form.get('n_ddt_ingresso')
            art.n_ddt_uscita = request.form.get('n_ddt_uscita')

            # Numeri
            art.pezzo = request.form.get('pezzo')
            # In modifica, l'articolo corrente resta sempre 1 collo (perché è una riga sola)
            # Se l'utente ha scritto 3, creiamo 2 copie e lasciamo questo a 1
            art.n_colli = 1 
            
            art.peso = to_float_eu(request.form.get('peso'))
            art.lunghezza = to_float_eu(request.form.get('lunghezza'))
            art.larghezza = to_float_eu(request.form.get('larghezza'))
            art.altezza = to_float_eu(request.form.get('altezza'))
            
            # Calcoli
            m2_calc, m3_calc = calc_m2_m3(art.lunghezza, art.larghezza, art.altezza, 1)
            art.m2 = m2_calc
            m3_man = request.form.get('m3')
            art.m3 = to_float_eu(m3_man) if m3_man and to_float_eu(m3_man) > 0 else m3_calc

            # 3. LOGICA SPLIT (Se colli > 1, crea copie)
            if colli_input > 1:
                import shutil
                # Recupera gli allegati attuali per copiarli sulle nuove righe
                current_attachments = db.query(Attachment).filter_by(articolo_id=art.id_articolo).all()
                
                for _ in range(colli_input - 1):
                    # Clona l'articolo
                    clone = Articolo()
                    for c in Articolo.__table__.columns:
                        if c.name not in ['id_articolo', 'attachments']:
                            setattr(clone, c.name, getattr(art, c.name))
                    
                    db.add(clone)
                    db.flush() # Ottieni ID del clone
                    
                    # Clona anche gli allegati
                    for att in current_attachments:
                        fname = att.filename
                        ext = fname.rsplit('.', 1)[-1].lower()
                        kind = att.kind
                        folder = PHOTOS_DIR if kind == 'photo' else DOCS_DIR
                        
                        src_path = folder / fname
                        if src_path.exists():
                            new_name = f"{clone.id_articolo}_{uuid.uuid4().hex[:6]}_{fname.split('_',1)[-1]}"
                            dst_path = folder / new_name
                            try:
                                shutil.copy2(src_path, dst_path)
                                db.add(Attachment(articolo_id=clone.id_articolo, filename=new_name, kind=kind))
                            except: pass

                flash(f"Articolo aggiornato e create {colli_input - 1} copie aggiuntive.", "success")
            else:
                flash("Articolo aggiornato con successo.", "success")

            db.commit()
            return redirect(url_for('giacenze'))

        # GET: Mostra template modifica
        return render_template('edit.html', row=art)

    except Exception as e:
        db.rollback()
        flash(f"Errore modifica: {e}", "danger")
        return redirect(url_for('giacenze'))
    finally:
        db.close()

@app.route('/edit/<int:id_articolo>', methods=['GET', 'POST'])
@login_required
def edit_record(id_articolo):
    db = SessionLocal()
    try:
        art = db.query(Articolo).filter(Articolo.id_articolo == id_articolo).first()
        if not art:
            flash("Articolo non trovato", "danger")
            return redirect(url_for('giacenze'))

        if request.method == 'POST':
            # --- SALVATAGGIO MODIFICHE ---
            colli_input = to_int_eu(request.form.get('n_colli'))
            if colli_input < 1: colli_input = 1

            # Aggiornamento campi
            art.codice_articolo = request.form.get('codice_articolo')
            art.descrizione = request.form.get('descrizione')
            # Cliente: per i client e' bloccato sul proprio utente
            if session.get('role') == 'client':
                art.cliente = current_user.id
            else:
                art.cliente = request.form.get('cliente')
            art.fornitore = request.form.get('fornitore')
            art.commessa = request.form.get('commessa')
            art.ordine = request.form.get('ordine')
            art.protocollo = request.form.get('protocollo')
            art.buono_n = request.form.get('buono_n')
            art.n_arrivo = request.form.get('n_arrivo')
            art.magazzino = request.form.get('magazzino')
            art.posizione = request.form.get('posizione')
            art.stato = request.form.get('stato')
            art.note = request.form.get('note')
            art.serial_number = request.form.get('serial_number')
            art.mezzi_in_uscita = request.form.get('mezzi_in_uscita')
            
            art.data_ingresso = parse_date_ui(request.form.get('data_ingresso'))
            art.data_uscita = parse_date_ui(request.form.get('data_uscita'))
            art.n_ddt_ingresso = request.form.get('n_ddt_ingresso')
            art.n_ddt_uscita = request.form.get('n_ddt_uscita')
            
            art.pezzo = request.form.get('pezzo')
            # Nella modifica singola, n_colli resta 1 per coerenza, se split gestito sotto
            art.n_colli = 1 
            art.peso = to_float_eu(request.form.get('peso'))
            art.lunghezza = to_float_eu(request.form.get('lunghezza'))
            art.larghezza = to_float_eu(request.form.get('larghezza'))
            art.altezza = to_float_eu(request.form.get('altezza'))
            
            # Calcoli
            m2_calc, m3_calc = calc_m2_m3(art.lunghezza, art.larghezza, art.altezza, 1)
            art.m2 = m2_calc
            m3_man = request.form.get('m3')
            art.m3 = to_float_eu(m3_man) if m3_man and to_float_eu(m3_man) > 0 else m3_calc

            # --- DUPLICAZIONE SE COLLI > 1 ---
            if colli_input > 1:
                for _ in range(colli_input - 1):
                    clone = Articolo()
                    for c in Articolo.__table__.columns:
                        if c.name not in ['id_articolo', 'attachments']:
                            setattr(clone, c.name, getattr(art, c.name))
                    db.add(clone)
                flash(f"Salvataggio OK. Generate {colli_input - 1} copie aggiuntive.", "success")
            else:
                flash("Modifiche salvate.", "success")

            db.commit()
            return redirect(url_for('giacenze'))

        return render_template('edit.html', row=art)
    except Exception as e:
        db.rollback()
        flash(f"Errore: {e}", "danger")
        return redirect(url_for('giacenze'))
    finally:
        db.close()

@app.route('/edit/<int:id>', methods=['GET','POST'])
@login_required
def edit_row(id):
    db = SessionLocal()
    row = db.get(Articolo, id)
    if not row:
        abort(404)

    if request.method == 'POST':
        for f, label in get_all_fields_map().items():
            v = request.form.get(f)
            if v is not None:
                if f in ('data_ingresso','data_uscita'):
                    v = parse_date_ui(v) if v else None
                elif f in ('larghezza','lunghezza','altezza','peso'):
                    v = to_float_eu(v)
                elif f in ('n_colli', 'pezzo'):
                    v = to_int_eu(v)
                setattr(row, f, v if v != '' else None)
        row.m2, row.m3 = calc_m2_m3(row.lunghezza, row.larghezza, row.altezza, row.n_colli)
        if 'files' in request.files:
            for f in request.files.getlist('files'):
                if not f or not f.filename: continue
                safe_name = f"{id}_{uuid.uuid4().hex}_{f.filename.replace(' ','_')}"
                ext = os.path.splitext(safe_name)[1].lower()
                kind = 'doc' if ext == '.pdf' else 'foto'
                folder = DOCS_DIR if kind == 'doc' else PHOTOS_DIR
                f.save(str(folder / safe_name))
                db.add(Attachment(articolo_id=id, kind=kind, filename=safe_name))
        db.commit()
        flash('Riga salvata', 'success')
        return redirect(url_for('giacenze'))

    return render_template('edit.html', row=row, fields=get_all_fields_map().items())



# --- MEDIA & ALLEGATI ---
@app.get('/media/<int:att_id>')
@login_required
def media(att_id):
    db = SessionLocal()
    att = db.get(Attachment, att_id)
    if not att: abort(404)
    path = (DOCS_DIR if att.kind=='doc' else PHOTOS_DIR) / att.filename
    if not path.exists():
        flash(f"File allegato non trovato sul server: {att.filename}", "danger")
        return redirect(request.referrer or url_for('giacenze'))
    return send_file(path, as_attachment=False)

# --- ROUTE PER ELIMINARE UN ALLEGATO ---
@app.route('/delete_attachment/<int:id_attachment>')
@login_required
def delete_attachment(id_attachment):
    # Protezione Ruolo
    if session.get('role') != 'admin':
        flash("Solo gli admin possono eliminare file.", "danger")
        return redirect(url_for('giacenze'))

    db = SessionLocal()
    try:
        att = db.query(Attachment).filter(Attachment.id == id_attachment).first()
        
        if att:
            article_id = att.articolo_id # Salva ID per il redirect
            
            # Percorsi possibili
            folder = PHOTOS_DIR if att.kind == 'photo' else DOCS_DIR
            file_path = folder / att.filename
            
            # Prova a cancellare il file fisico
            if file_path.exists():
                try:
                    os.remove(file_path)
                except Exception as e:
                    print(f"Avviso: Errore rimozione file fisico {e}")
            
            # ELIMINA SEMPRE DAL DATABASE (Pulizia)
            db.delete(att)
            db.commit()
            
            flash("Allegato eliminato.", "success")
            return redirect(url_for('edit_record', id_articolo=article_id))
        else:
            flash("Allegato non trovato nel database.", "warning")
            return redirect(url_for('giacenze'))
            
    except Exception as e:
        db.rollback()
        flash(f"Errore eliminazione: {e}", "danger")
        return redirect(url_for('giacenze'))
    finally:
        db.close()


    # Redirect back to the correct page
    if table == 'trasporti': return redirect(url_for('trasporti'))
    if table == 'lavorazioni': return redirect(url_for('lavorazioni'))
    return redirect(url_for('home'))

# ==============================================================================
#  1. FUNZIONE GIACENZE (Visualizzazione Magazzino)
# ==============================================================================
@app.route('/giacenze', methods=['GET', 'POST'])
@login_required
def giacenze():
    import logging
    import re
    import math
    from sqlalchemy.orm import selectinload
    from sqlalchemy import func
    from datetime import datetime, date

    db = SessionLocal()
    try:
        # Configurazione Paginazione
        PER_PAGE = 50
        page = request.args.get('page', 1, type=int)
        args = request.args

        # 1) Query Base
        qs = (
            db.query(Articolo)
            .options(selectinload(Articolo.attachments))
            .order_by(Articolo.id_articolo.desc())
        )

        # 2) Filtri Base (cliente)
        if session.get('role') == 'client':
            user_key = (current_user.id or '').strip().upper()
            user_key_norm = re.sub(r'[^A-Z0-9]+', '', user_key)

            cliente_db_norm = func.upper(func.trim(Articolo.cliente))
            for char in [' ', '.', '-', '_']:
                cliente_db_norm = func.replace(cliente_db_norm, char, '')

            qs = qs.filter(cliente_db_norm.like(f"%{user_key_norm}%"))
        else:
            if args.get('cliente'):
                qs = qs.filter(Articolo.cliente.ilike(f"%{args.get('cliente')}%"))

        # 3) Filtro ID
        if args.get('id'):
            try:
                qs = qs.filter(Articolo.id_articolo == int(args.get('id')))
            except:
                pass

        # 4) Filtri Testuali
        text_filters = [
            'commessa', 'descrizione', 'posizione', 'buono_n', 'protocollo', 'lotto',
            'fornitore', 'ordine', 'magazzino', 'mezzi_in_uscita', 'stato',
            'n_ddt_ingresso', 'n_ddt_uscita', 'codice_articolo', 'serial_number', 'n_arrivo'
        ]
        for field in text_filters:
            val = args.get(field)
            if val and val.strip():
                qs = qs.filter(getattr(Articolo, field).ilike(f"%{val.strip()}%"))

        # 5) Recupero righe (per filtro date in Python)
        all_rows = qs.all()
        filtered_rows = []

        # 6) Filtri Date
        def get_date_arg(k):
            v = args.get(k)
            try:
                return datetime.strptime(v, "%Y-%m-%d").date() if v else None
            except:
                return None

        d_ing_da, d_ing_a = get_date_arg('data_ing_da'), get_date_arg('data_ing_a')
        d_usc_da, d_usc_a = get_date_arg('data_usc_da'), get_date_arg('data_usc_a')

        def parse_d(val):
            if isinstance(val, date):
                return val
            if not val:
                return None
            if isinstance(val, str):
                try:
                    return datetime.strptime(val[:10], "%Y-%m-%d").date()
                except:
                    return None
            return None

        if any([d_ing_da, d_ing_a, d_usc_da, d_usc_a]):
            for r in all_rows:
                keep = True

                # Ingresso
                if d_ing_da or d_ing_a:
                    rd = parse_d(r.data_ingresso)
                    if not rd or (d_ing_da and rd < d_ing_da) or (d_ing_a and rd > d_ing_a):
                        keep = False

                # Uscita
                if keep and (d_usc_da or d_usc_a):
                    rd = parse_d(r.data_uscita)
                    if not rd or (d_usc_da and rd < d_usc_da) or (d_usc_a and rd > d_usc_a):
                        keep = False

                if keep:
                    filtered_rows.append(r)
        else:
            filtered_rows = all_rows

        # ✅ 7) NUOVO FILTRO: SOLO IN GIACENZA
        # In giacenza = NON ha data_uscita e NON ha n_ddt_uscita
        if args.get("solo_giacenza") == "1":
            tmp = []
            for r in filtered_rows:
                has_data_usc = parse_d(r.data_uscita) is not None
                has_ddt_usc = bool((r.n_ddt_uscita or "").strip())
                if (not has_data_usc) and (not has_ddt_usc):
                    tmp.append(r)
            filtered_rows = tmp

        # 8) Totali (sui risultati filtrati)
        total_colli = 0
        total_m2 = 0.0
        total_peso = 0.0

        for r in filtered_rows:
            try:
                total_colli += int(r.n_colli or 0)
            except:
                pass
            try:
                total_m2 += float(r.m2) if r.m2 else 0.0
            except:
                pass
            try:
                total_peso += float(r.peso) if r.peso else 0.0
            except:
                pass

        # 9) Paginazione
        total_items = len(filtered_rows)
        total_pages = math.ceil(total_items / PER_PAGE) if total_items else 1

        if page < 1:
            page = 1
        if page > total_pages:
            page = total_pages

        start = (page - 1) * PER_PAGE
        end = start + PER_PAGE
        current_page_rows = filtered_rows[start:end]

        # ✅ FIX: parametri senza "page"
        search_params = request.args.copy()
        if 'page' in search_params:
            del search_params['page']

        return render_template(
            'giacenze.html',
            rows=current_page_rows,
            result=current_page_rows,
            page=page,
            total_pages=total_pages,
            total_items=total_items,
            total_colli=total_colli,
            total_m2=it_num(total_m2, 2),
            total_peso=it_num(total_peso, 2),
            today=date.today(),
            search_params=search_params
        )

    except Exception as e:
        logging.error(f"ERRORE GIACENZE: {e}")
        return f"<h1>Errore: {e}</h1>"
    finally:
        db.close()

# ==============================================================================
#  3. FUNZIONE ELIMINA (Risolve l'errore 'endpoint elimina_record')
# ==============================================================================
@app.route('/elimina_record/<table>/<int:id>')
@login_required
def elimina_record(table, id):
    # Solo Admin può eliminare
    if session.get('role') != 'admin':
        flash("Accesso Negato: Solo Admin può eliminare.", "danger")
        return redirect(url_for('giacenze'))

    db = SessionLocal()
    try:
        record = None
        redirect_url = 'home'

        if table == 'articoli':
            record = db.query(Articolo).get(id)
            redirect_url = 'giacenze'
        elif table == 'trasporti':
            record = db.query(Trasporto).get(id)
            redirect_url = 'trasporti'
        elif table == 'lavorazioni':
            record = db.query(Lavorazione).get(id)
            redirect_url = 'lavorazioni'
        
        if record:
            db.delete(record)
            db.commit()
            flash("Elemento eliminato.", "success")
        else:
            flash("Elemento non trovato.", "warning")

        return redirect(url_for(redirect_url))

    except Exception as e:
        db.rollback()
        flash(f"Errore eliminazione: {e}", "danger")
        return redirect(url_for('home'))
    finally:
        db.close()

# --- MODIFICA MULTIPLA COMPLETA CON CALCOLI ---

@app.route('/bulk_edit', methods=['GET', 'POST'])
@login_required
@require_admin
def bulk_edit():
    db = SessionLocal()
    try:
        # Recupera ID (da form POST o query string GET)
        ids = request.form.getlist('ids') or request.args.getlist('ids')

        # Filtra ID validi
        ids = [int(i) for i in ids if str(i).isdigit()]

        if not ids:
            flash("Nessun articolo selezionato.", "warning")
            return redirect(url_for('giacenze'))

        articoli = db.query(Articolo).filter(Articolo.id_articolo.in_(ids)).all()

        # Configurazione Campi Modificabili
        editable_fields = [
            ('Cliente', 'cliente'),
            ('Fornitore', 'fornitore'),
            ('N. DDT Ingresso', 'n_ddt_ingresso'),
            ('Data Ingresso', 'data_ingresso'),
            ('Data Uscita', 'data_uscita'),
            ('N. DDT Uscita', 'n_ddt_uscita'),
            ('Protocollo', 'protocollo'),
            ('N. Buono', 'buono_n'),
            ('Magazzino', 'magazzino'),
            ('Commessa', 'commessa'),
            ('Mezzo Uscita', 'mezzi_in_uscita'),
            ('N. Arrivo', 'n_arrivo'),
            ('Peso', 'peso'),
            ('Lotto', 'lotto'),
            ('Ordine', 'ordine'),
            ('Stato', 'stato'),
            ('Descrizione', 'descrizione'),
            ('Codice Articolo', 'codice_articolo'),
            ('Serial Number', 'serial_number'),
            ('Colli', 'n_colli'),
            ('Pezzi', 'pezzo'),
            ('Lunghezza', 'lunghezza'),
            ('Larghezza', 'larghezza'),
            ('Altezza', 'altezza'),
        ]

        if request.method == 'POST' and request.form.get('save_bulk') == 'true':
            updates = {}
            recalc_dims = False

            # 1) Applica Modifiche Campi
            for key in request.form:
                if not key.startswith('chk_'):
                    continue

                field_name = key.replace('chk_', '').strip()

                # Accetta SOLO campi presenti nella lista editable_fields
                if not any(f[1] == field_name for f in editable_fields):
                    continue

                val = request.form.get(field_name)

                if field_name in ['n_colli', 'pezzo']:
                    val = to_int_eu(val)
                elif field_name in ['lunghezza', 'larghezza', 'altezza', 'peso']:
                    val = to_float_eu(val)
                elif 'data' in field_name:
                    val = parse_date_ui(val) if val else None
                else:
                    val = (val or "").strip()

                updates[field_name] = val

                if field_name in ['lunghezza', 'larghezza', 'altezza', 'n_colli']:
                    recalc_dims = True

            if updates:
                for art in articoli:
                    for k, v in updates.items():
                        if hasattr(art, k):
                            setattr(art, k, v)

                    if recalc_dims:
                        # Ricalcola M2/M3 usando i nuovi valori (o quelli esistenti se non cambiati)
                        L = updates.get('lunghezza', art.lunghezza)
                        W = updates.get('larghezza', art.larghezza)
                        H = updates.get('altezza', art.altezza)
                        C = updates.get('n_colli', art.n_colli)
                        art.m2, art.m3 = calc_m2_m3(L, W, H, C)

            # 2) UPLOAD MASSIVO MULTIPLO (più file) ✅ CORRETTO
            files = request.files.getlist('bulk_files')
            count_uploaded = 0

            if files:
                from werkzeug.utils import secure_filename

                for file in files:
                    if not file or not file.filename:
                        continue

                    raw_name = secure_filename(file.filename)

                    # Legge UNA volta (poi lo duplica su ogni articolo)
                    content = file.read()
                    if not content:
                        continue

                    ext = os.path.splitext(raw_name)[1].lower()

                    # Coerenza: doc=PDF, photo=immagine
                    if ext == '.pdf':
                        kind = 'doc'
                        dest_dir = DOCS_DIR
                    elif ext in ['.jpg', '.jpeg', '.png', '.webp']:
                        kind = 'photo'
                        dest_dir = PHOTOS_DIR
                    else:
                        kind = 'doc'
                        dest_dir = DOCS_DIR

                    # Salva una copia per ogni articolo selezionato
                    for art in articoli:
                        new_name = f"{art.id_articolo}_{uuid.uuid4().hex[:6]}_{raw_name}"
                        save_path = dest_dir / new_name

                        with open(save_path, 'wb') as f_out:
                            f_out.write(content)

                        db.add(Attachment(articolo_id=art.id_articolo, filename=new_name, kind=kind))

                    count_uploaded += 1

            db.commit()
            flash(
                f"Aggiornati {len(articoli)} articoli e caricati {count_uploaded} file (copiati su ciascun articolo).",
                "success"
            )
            return redirect(url_for('giacenze'))

        return render_template(
            'bulk_edit.html',
            rows=articoli,
            ids_csv=",".join(map(str, ids)),
            fields=editable_fields
        )

    except Exception as e:
        db.rollback()
        print(f"ERRORE BULK: {e}")
        flash(f"Errore: {e}", "danger")
        return redirect(url_for('giacenze'))
    finally:
        db.close()


@app.post('/delete_rows')
@login_required
@require_admin
def delete_rows():
    # Controllo Permessi: Solo Admin può cancellare
    if session.get('role') != 'admin':
        flash("Non hai i permessi per eliminare le righe.", "danger")
        return redirect(url_for('giacenze'))

    ids = request.form.getlist('ids')
    if not ids:
        flash("Nessuna riga selezionata per l'eliminazione.", "warning")
        return redirect(url_for('giacenze'))

    db = SessionLocal()
    try:
        # Filtra solo ID numerici validi
        clean_ids = [int(x) for x in ids if x.isdigit()]
        
        # Esegue la cancellazione
        affected = db.query(Articolo).filter(Articolo.id_articolo.in_(clean_ids)).delete(synchronize_session=False)
        db.commit()
        
        flash(f"Eliminati {affected} articoli.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Errore durante l'eliminazione: {e}", "danger")
    finally:
        db.close()
        
    return redirect(url_for('giacenze'))
        
@app.post('/bulk/delete')
@login_required
def bulk_delete():
    ids = [int(i) for i in request.form.getlist('ids') if i.isdigit()]
    if not ids:
        flash("Nessun articolo selezionato per l'eliminazione.", "warning")
        return redirect(url_for('giacenze'))
    
    db = SessionLocal()
    articoli_da_eliminare = db.query(Articolo).filter(Articolo.id_articolo.in_(ids)).all()
    for art in articoli_da_eliminare:
        for att in art.attachments:
            path = (DOCS_DIR if att.kind=='doc' else PHOTOS_DIR) / att.filename
            try:
                if path.exists(): path.unlink()
            except Exception: pass

    db.query(Articolo).filter(Articolo.id_articolo.in_(ids)).delete(synchronize_session=False)
    db.commit()
    flash(f"{len(ids)} articoli e i loro allegati sono stati eliminati.", "success")
    return redirect(url_for('giacenze'))

@app.post('/bulk/duplicate')
@login_required
@require_admin
def bulk_duplicate():
    # Controllo Permessi
    if session.get('role') != 'admin':
        flash("Non hai i permessi per duplicare.", "danger")
        return redirect(url_for('giacenze'))

    ids = request.form.getlist('ids')
    if not ids:
        flash("Nessun articolo selezionato.", "warning")
        return redirect(url_for('giacenze'))

    db = SessionLocal()
    try:
        count = 0
        for id_str in ids:
            if not id_str.isdigit(): continue
            original = db.query(Articolo).get(int(id_str))
            if original:
                # Clona tutti i campi tranne ID
                new_art = Articolo(
                    codice_articolo=original.codice_articolo,
                    descrizione=original.descrizione,
                    cliente=original.cliente,
                    fornitore=original.fornitore,
                    commessa=original.commessa,
                    protocollo=original.protocollo,
                    buono_n=original.buono_n,
                    ordine=original.ordine,
                    lotto=original.lotto,
                    data_ingresso=original.data_ingresso,
                    n_ddt_ingresso=original.n_ddt_ingresso,
                    data_uscita=original.data_uscita,
                    n_ddt_uscita=original.n_ddt_uscita,
                    pezzo=original.pezzo,
                    n_colli=original.n_colli,
                    peso=original.peso,
                    lunghezza=original.lunghezza,
                    larghezza=original.larghezza,
                    altezza=original.altezza,
                    m2=original.m2,
                    m3=original.m3,
                    n_arrivo=original.n_arrivo,
                    stato=original.stato,
                    magazzino=original.magazzino,
                    posizione=original.posizione,
                    note=original.note
                )
                db.add(new_art)
                count += 1
        db.commit()
        flash(f"{count} articoli duplicati.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Errore duplicazione: {e}", "danger")
    finally:
        db.close()
    return redirect(url_for('giacenze'))
# --- ANTEPRIME HTML (BUONO / DDT) ---
def _get_rows_from_ids(ids_list):
    if not ids_list: return []
    db=SessionLocal()
    return db.query(Articolo).filter(Articolo.id_articolo.in_(ids_list)).all()

@app.post('/buono/preview')
@login_required
def buono_preview():

    if session.get('role') != 'admin':
        flash('Accesso negato.', 'danger')
        return redirect(url_for('giacenze'))
    # Recupera gli ID selezionati
    ids_str_list = request.form.getlist('ids')
    ids = [int(i) for i in ids_str_list if i.isdigit()]
    
    db = SessionLocal()
    try:
        rows = db.query(Articolo).filter(Articolo.id_articolo.in_(ids)).all()
        
        # --- LOGICA DI AUTO-COMPILAZIONE ---
        
        # 1. PROTOCOLLO
        protocolli_trovati = set()
        for r in rows:
            if r.protocollo and str(r.protocollo).strip():
                protocolli_trovati.add(str(r.protocollo).strip())
        protocollo_auto = ", ".join(sorted(protocolli_trovati))

        # 2. COMMESSA
        commessa_auto = next((r.commessa for r in rows if r.commessa), "")
        
        # 3. FORNITORE
        fornitore_auto = next((r.fornitore for r in rows if r.fornitore), "")

        # 4. N. BUONO (ristampa)
        buono_n_auto = next((r.buono_n for r in rows if r.buono_n), "")
        
        # 5. ORDINE (NUOVO!)
        ordine_auto = next((r.ordine for r in rows if r.ordine), "")

        meta = {
            "buono_n": buono_n_auto, 
            "data_em": datetime.today().strftime("%d/%m/%Y"),
            "commessa": commessa_auto, 
            "fornitore": fornitore_auto,
            "protocollo": protocollo_auto,
            "ordine": ordine_auto, # ✅ CAMPO AGGIUNTO
        }
        
        return render_template('buono_preview.html', rows=rows, meta=meta, ids=",".join(map(str, ids)))
        
    finally:
        db.close()

from datetime import date
from flask import request, redirect, url_for, flash, jsonify, render_template
from flask_login import login_required

@app.post('/ddt/preview')
@login_required
def ddt_preview():
    if session.get('role') != 'admin':
        flash('Accesso negato.', 'danger')
        return redirect(url_for('giacenze'))

    # ⚠️ ids può arrivare in due modi:
    # - lista di checkbox: ids=1 ids=2 ids=3
    # - stringa csv in un campo hidden: ids="1,2,3"
    ids = []

    ids_list = request.form.getlist('ids')
    if ids_list:
        # caso checkbox
        for i in ids_list:
            if str(i).isdigit():
                ids.append(int(i))
    else:
        # caso hidden "1,2,3"
        ids_csv = (request.form.get('ids') or '').strip()
        if ids_csv:
            for part in ids_csv.split(','):
                part = part.strip()
                if part.isdigit():
                    ids.append(int(part))

    if not ids:
        flash("Seleziona almeno un articolo per creare il DDT.", "warning")
        return redirect(url_for('giacenze'))

    rows = _get_rows_from_ids(ids)

    return render_template(
        'ddt_preview.html',
        rows=rows,
        ids=",".join(map(str, ids)),
        destinatari=load_destinatari(),
        n_ddt=peek_next_ddt_number(),
        oggi=date.today().isoformat()
    )


@app.get('/next_ddt_number')
@login_required
def get_next_ddt_number():
    """
    Restituisce il numero successivo rispetto a quello attualmente scritto nel campo,
    SENZA memorizzarlo su DB (serve solo per cambiare il valore con le frecce).
    Accetta ?current=NN/YY oppure ?current=NN.
    """
    current = (request.args.get('current') or '').strip()

    # base = numero + anno
    base_num = None
    base_year = None

    # prova formato NN/YY
    m = re.match(r'^(\d+)\s*/\s*(\d{2})$', current)
    if m:
        try:
            base_num = int(m.group(1))
            base_year = m.group(2)
        except Exception:
            base_num = None
            base_year = None

    # prova solo numero
    if base_num is None and current.isdigit():
        base_num = int(current)

    # se non valido, usa il "peek" (es. 01/26)
    if base_num is None or base_year is None:
        try:
            p = peek_next_ddt_number()
            pm = re.match(r'^(\d+)\s*/\s*(\d{2})$', (p or '').strip())
            if pm:
                base_num = int(pm.group(1))
                base_year = pm.group(2)
        except Exception:
            pass

    if base_num is None:
        base_num = 1
    if base_year is None:
        base_year = str(date.today().year)[-2:]

    nxt = base_num + 1
    return jsonify({'next_ddt': f"{nxt:02d}/{base_year}"})


@app.get('/prev_ddt_number')
@login_required
def get_prev_ddt_number():
    """
    Restituisce il numero precedente rispetto a quello attualmente scritto nel campo,
    SENZA memorizzarlo su DB (serve solo per cambiare il valore con le frecce).
    Accetta ?current=NN/YY oppure ?current=NN.
    """
    current = (request.args.get('current') or '').strip()

    base_num = None
    base_year = None

    m = re.match(r'^(\d+)\s*/\s*(\d{2})$', current)
    if m:
        try:
            base_num = int(m.group(1))
            base_year = m.group(2)
        except Exception:
            base_num = None
            base_year = None

    if base_num is None and current.isdigit():
        base_num = int(current)

    if base_num is None or base_year is None:
        try:
            p = peek_next_ddt_number()
            pm = re.match(r'^(\d+)\s*/\s*(\d{2})$', (p or '').strip())
            if pm:
                base_num = int(pm.group(1))
                base_year = pm.group(2)
        except Exception:
            pass

    if base_num is None:
        base_num = 1
    if base_year is None:
        base_year = str(date.today().year)[-2:]

    prv = max(1, base_num - 1)
    return jsonify({'prev_ddt': f"{prv:02d}/{base_year}"})


@app.route('/manage_destinatari', methods=['GET', 'POST'])
@login_required
def manage_destinatari():
    import json

    dest_file = APP_DIR / "destinatari_saved.json"
    destinatari = load_destinatari()

    if request.method == 'POST':

        # =========================
        # ELIMINAZIONE DESTINATARIO
        # =========================
        if 'delete_key' in request.form:
            key_to_delete = (request.form.get('delete_key') or '').strip()

            if key_to_delete and key_to_delete in destinatari:
                del destinatari[key_to_delete]
                try:
                    dest_file.write_text(
                        json.dumps(destinatari, ensure_ascii=False, indent=4),
                        encoding="utf-8"
                    )
                    flash(f"Destinatario '{key_to_delete}' eliminato.", "success")
                except Exception as e:
                    flash(f"Errore salvataggio file: {e}", "danger")
            else:
                flash("Destinatario non trovato.", "warning")

        # =========================
        # AGGIUNTA / MODIFICA
        # =========================
        else:
            key_name = (request.form.get('key_name') or '').strip()

            if not key_name:
                flash("Il Nome Chiave è obbligatorio.", "warning")
            else:
                destinatari[key_name] = {
                    "ragione_sociale": (request.form.get('ragione_sociale') or '').strip(),
                    "indirizzo": (request.form.get('indirizzo') or '').strip(),
                    "piva": (request.form.get('piva') or '').strip()
                }

                try:
                    dest_file.write_text(
                        json.dumps(destinatari, ensure_ascii=False, indent=4),
                        encoding="utf-8"
                    )
                    flash(f"Destinatario '{key_name}' salvato.", "success")
                except Exception as e:
                    flash(f"Errore salvataggio file: {e}", "danger")

        return redirect(url_for('manage_destinatari'))

    # =========================
    # GET
    # =========================
    return render_template(
        'destinatari.html',
        destinatari=destinatari
    )


# ========================================================
#  RUBRICA EMAIL (UI) + BACKUP (download)
# ========================================================

@app.route('/rubrica_email', methods=['GET', 'POST'])
@login_required
@require_admin
def rubrica_email():
    data = load_rubrica_email()

    if request.method == 'POST':
        action = request.form.get('action', 'save')

        if action == 'save_contact':
            nome = (request.form.get('nome') or '').strip()
            email = (request.form.get('email') or '').strip()
            if not nome or not email:
                flash("Nome ed email sono obbligatori.", "warning")
            else:
                data["contatti"][nome] = {"email": email}
                save_rubrica_email(data)
                flash("Contatto salvato.", "success")

        elif action == 'delete_contact':
            nome = (request.form.get('nome') or '').strip()
            if nome in data.get("contatti", {}):
                del data["contatti"][nome]
                # rimuovi anche dai gruppi
                for g, emails in list(data.get("gruppi", {}).items()):
                    data["gruppi"][g] = [e for e in emails if e != nome and e != data.get("contatti", {}).get(nome, {}).get("email")]
                save_rubrica_email(data)
                flash("Contatto eliminato.", "success")

        elif action == 'save_group':
            gruppo = (request.form.get('gruppo') or '').strip()
            raw = (request.form.get('emails') or '').strip()
            if not gruppo:
                flash("Nome gruppo obbligatorio.", "warning")
            else:
                emails = _parse_emails(raw)
                data["gruppi"][gruppo] = emails
                save_rubrica_email(data)
                flash("Gruppo salvato.", "success")

        elif action == 'delete_group':
            gruppo = (request.form.get('gruppo') or '').strip()
            if gruppo in data.get("gruppi", {}):
                del data["gruppi"][gruppo]
                save_rubrica_email(data)
                flash("Gruppo eliminato.", "success")

        return redirect(url_for('rubrica_email'))

    return render_template('rubrica_email.html', rubrica=data)

@app.route('/backup', methods=['GET'])
@login_required
@require_admin
def backup_download():
    try:
        p = create_backup_zip(include_media=True)
        return send_file(p, as_attachment=True, download_name=p.name, mimetype="application/zip")
    except Exception as e:
        flash(f"Errore backup: {e}", "danger")
        return redirect(url_for('home'))



def _pdf_table(data, col_widths=None, header=True, hAlign='LEFT', style=None):
    t = Table(data, colWidths=col_widths, hAlign=hAlign)
    base_style = [
        ('FONT', (0,0), (-1,-1), 'Helvetica', 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
    ]
    if style:
        base_style.extend(style)
    else:
        base_style.append(('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey))

    if header and data:
        base_style.extend([
            ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
            ('FONT', (0,0), (-1,0), 'Helvetica-Bold', 8)
        ])
    t.setStyle(TableStyle(base_style))
    return t

def _copyright_para():
    tiny_style = _styles['Normal'].clone('copyright')
    tiny_style.fontSize = 7; tiny_style.textColor = colors.grey; tiny_style.alignment = TA_CENTER
    return Paragraph("Camar S.r.l. - Gestionale Web - © Alessia Moncalvo", tiny_style)

def _generate_buono_pdf(form_data, rows):
    import io
    from pathlib import Path
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER

    def _to_int_safe(v, default=0):
        try:
            if v is None:
                return default
            s = str(v).strip()
            if s == "":
                return default
            s = s.replace(",", ".")
            return int(float(s))
        except Exception:
            return default

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=10*mm, bottomMargin=10*mm
    )
    story = []

    styles = getSampleStyleSheet()
    s_norm  = ParagraphStyle('Norm', parent=styles['Normal'], fontSize=9, leading=11, textColor=colors.black)
    s_bold  = ParagraphStyle('Bold', parent=s_norm, fontName='Helvetica-Bold')
    s_title = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=10, textColor=colors.black)
    s_note  = ParagraphStyle('Note', parent=s_norm, fontSize=9, textColor=colors.darkblue)

    # 1) Logo
    if 'LOGO_PATH' in globals() and LOGO_PATH and Path(LOGO_PATH).exists():
        story.append(Image(str(LOGO_PATH), width=50*mm, height=16*mm, hAlign='CENTER'))
    else:
        story.append(Paragraph("<b>Ca.mar. srl</b>", s_title))

    story.append(Spacer(1, 5*mm))
    story.append(Paragraph("BUONO DI PRELIEVO", s_title))
    story.append(Spacer(1, 5*mm))

    # 2) Dati Testata
    meta_data = [
        [Paragraph("<b>Data Emissione:</b>", s_bold), Paragraph(str(form_data.get('data_em','')), s_norm)],
        [Paragraph("<b>Cliente:</b>", s_bold), Paragraph(str(rows[0].cliente if rows else ''), s_norm)],
        [Paragraph("<b>Fornitore:</b>", s_bold), Paragraph(str(form_data.get('fornitore','')), s_norm)],
        [Paragraph("<b>Commessa:</b>", s_bold), Paragraph(str(form_data.get('commessa','')), s_norm)],
        [Paragraph("<b>Ordine:</b>", s_bold), Paragraph(str(form_data.get('ordine','')), s_norm)],
        [Paragraph("<b>Protocollo:</b>", s_bold), Paragraph(str(form_data.get('protocollo','')), s_norm)],
        [Paragraph("<b>N. Buono:</b>", s_bold), Paragraph(str(form_data.get('buono_n','')), s_norm)],
    ]

    t_meta = Table(meta_data, colWidths=[40*mm, 140*mm])
    t_meta.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (0,-1), colors.whitesmoke),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 8*mm))

    # 3) Articoli
    header = [
        Paragraph('<b>Codice</b>', s_bold),
        Paragraph('<b>Descrizione</b>', s_bold),
        Paragraph('<b>Q.tà</b>', s_bold),
        Paragraph('<b>N.Arr</b>', s_bold)
    ]
    table_data = [header]

    for r in rows:
        # ✅ Q.tà: prende il valore inserito nel form (q_ID), altrimenti usa PEZZI (r.pezzo)
        q_form = form_data.get(f"q_{r.id_articolo}")
        if q_form is not None and str(q_form).strip() != "":
            q = _to_int_safe(q_form, default=0)
        else:
            q = _to_int_safe(getattr(r, "pezzo", None), default=0)

        desc = str(r.descrizione or '')
        note_user = form_data.get(f"note_{r.id_articolo}")
        if note_user is None:
            note_user = r.note

        table_data.append([
            Paragraph(str(r.codice_articolo or ''), s_norm),
            Paragraph(desc, s_norm),
            str(q),
            Paragraph(str(r.n_arrivo or ''), s_norm)
        ])

        if note_user:
            table_data.append([
                '',
                Paragraph(f"<i>Note: {note_user}</i>", s_note),
                '', ''
            ])

    t = Table(table_data, colWidths=[40*mm, 100*mm, 15*mm, 25*mm], repeatRows=1)
    t.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 4)
    ]))
    story.append(t)

    story.append(Spacer(1, 20*mm))
    sig_data = [[
        Paragraph("Firma Magazzino:<br/><br/>__________________", s_norm),
        Paragraph("Firma Cliente:<br/><br/>__________________", s_norm)
    ]]
    t_sig = Table(sig_data, colWidths=[90*mm, 90*mm])
    story.append(t_sig)

    doc.build(story)
    bio.seek(0)
    return bio



# --- GENERAZIONE PDF DDT (LAYOUT RICHIESTO) ---

def _genera_pdf_ddt_file(ddt_data, righe, filename_out):
    import io
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from pathlib import Path

    bio = filename_out
    doc = SimpleDocTemplate(bio, pagesize=A4, leftMargin=10*mm, rightMargin=10*mm, topMargin=5*mm, bottomMargin=5*mm)
    story = []
    
    styles = getSampleStyleSheet()
    s_norm = styles['Normal']
    s_small = ParagraphStyle('s', parent=s_norm, fontSize=9, leading=11)
    s_bold = ParagraphStyle('b', parent=s_small, fontName='Helvetica-Bold')
    s_white = ParagraphStyle('w', parent=s_bold, textColor=colors.white, alignment=TA_CENTER, fontSize=14)

    def clean(val):
        if val is None: return ""
        s = str(val).strip()
        if s.lower() == 'none': return ""
        return s

    # 1. Logo
    if LOGO_PATH and Path(LOGO_PATH).exists():
        story.append(Image(LOGO_PATH, width=50*mm, height=16*mm, hAlign='CENTER'))
    story.append(Spacer(1, 2*mm))

    # 2. Titolo
    t_title = Table([[Paragraph("DOCUMENTO DI TRASPORTO (DDT)", s_white)]], colWidths=[190*mm])
    t_title.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#4682B4")), 
        ('PADDING', (0,0), (-1,-1), 6)
    ]))
    story.append(t_title)
    story.append(Spacer(1, 2*mm))

    # 3. Testata
    dest_ragione = clean(ddt_data.get('destinatario'))
    dest_ind = clean(ddt_data.get('dest_indirizzo')).replace('\n', '<br/>')
    dest_citta = clean(ddt_data.get('dest_citta'))
    
    mittente_html = "<b>Mittente</b><br/>Camar srl<br/>Via Luigi Canepa 2<br/>16165 Genova Struppa (GE)"
    dest_html = f"<b>Destinatario</b><br/>{dest_ragione}<br/>{dest_ind}<br/>{dest_citta}"
    
    t_md = Table([[Paragraph(mittente_html, s_small), Paragraph(dest_html, s_small)]], colWidths=[95*mm, 95*mm])
    t_md.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 5)
    ]))
    story.append(t_md)
    story.append(Spacer(1, 2*mm))

    # 4. Dati Aggiuntivi
    t_bar = Table([[Paragraph("Dati Aggiuntivi", ParagraphStyle('wb', parent=s_white, fontSize=10, alignment=TA_LEFT))]], colWidths=[190*mm])
    t_bar.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#4682B4")), ('PADDING', (0,0), (-1,-1), 2)]))
    story.append(t_bar)

    first = righe[0] if righe else {}
    commessa = clean(first.get('commessa', ''))
    ordine = clean(first.get('ordine', ''))
    buono = clean(first.get('buono', ''))
    protocollo = clean(first.get('protocollo', ''))
    n_ddt = clean(ddt_data.get('n_ddt'))
    data_ddt = clean(ddt_data.get('data_uscita'))
    targa = clean(ddt_data.get('vettore')) 
    causale = clean(ddt_data.get('causale'))

    dati_agg = [
        [Paragraph("<b>Commessa</b>", s_bold), Paragraph(commessa, s_small), Paragraph("<b>N. DDT</b>", s_bold), Paragraph(n_ddt, s_small)],
        [Paragraph("<b>Ordine</b>", s_bold), Paragraph(ordine, s_small), Paragraph("<b>Data Uscita</b>", s_bold), Paragraph(data_ddt, s_small)],
        [Paragraph("<b>Buono</b>", s_bold), Paragraph(buono, s_small), Paragraph("<b>Targa</b>", s_bold), Paragraph(targa, s_small)],
        [Paragraph("<b>Protocollo</b>", s_bold), Paragraph(protocollo, s_small), Paragraph("<b>Causale</b>", s_bold), Paragraph(causale, s_small)]
    ]
    t_agg = Table(dati_agg, colWidths=[25*mm, 70*mm, 25*mm, 70*mm])
    t_agg.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    story.append(t_agg)
    story.append(Spacer(1, 5*mm))

    # 5. Articoli
    story.append(Paragraph("<b>Articoli nel DDT</b>", ParagraphStyle('h', parent=s_bold, fontSize=10)))
    
    header = [Paragraph(x, s_bold) for x in ['ID', 'Cod.Art.', 'Descrizione', 'Pezzi', 'Colli', 'Peso', 'N.Arrivo']]
    data = [header]
    
    tot_pezzi = 0; tot_colli = 0; tot_peso = 0.0
    note_da_stampare = []

    for r in righe:
        pz = int(r.get('pezzo') or 0)
        cl = int(r.get('n_colli') or 0)
        we = float(r.get('peso') or 0.0)
        tot_pezzi += pz; tot_colli += cl; tot_peso += we
        
        # Note fuori (MODIFICATO: Solo il testo della nota, senza codice articolo)
        nota = r.get('note')
        if nota and str(nota).strip():
            # Prima era: f"NOTE ARTICOLO ({r.get('codice_articolo')}): {nota}"
            # Ora è solo: f"NOTE: {nota}"
            note_da_stampare.append(f"NOTE: {nota}")

        data.append([
            Paragraph(str(r.get('id_articolo','')), s_small),
            Paragraph(clean(r.get('codice_articolo')), s_small),
            Paragraph(clean(r.get('descrizione')), s_small),
            str(pz),
            str(cl),
            f"{we:.0f}",
            Paragraph(clean(r.get('n_arrivo')), s_small)
        ])

    t_items = Table(data, colWidths=[15*mm, 35*mm, 80*mm, 15*mm, 15*mm, 15*mm, 15*mm], repeatRows=1)
    t_items.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 4)
    ]))
    story.append(t_items)
    story.append(Spacer(1, 3*mm))

    # Note
    if note_da_stampare:
        story.append(Spacer(1, 2*mm))
        for n in note_da_stampare:
            story.append(Paragraph(f"<i>{n}</i>", s_small))
        story.append(Spacer(1, 3*mm))

    # 6. Footer
    porto = clean(ddt_data.get('porto'))
    aspetto = clean(ddt_data.get('aspetto'))

    footer_data = [
        [
            Paragraph(f"<b>Porto:</b> {porto}", s_small),
            Paragraph(f"<b>Aspetto:</b> {aspetto}", s_small)
        ],
        [
            Paragraph(f"<b>TOTALE:</b> Pezzi {tot_pezzi} - Colli {tot_colli} - Peso {tot_peso:.0f} Kg", s_bold),
            Paragraph("Firma Vettore: _______________________", s_small)
        ]
    ]
    
    t_foot = Table(footer_data, colWidths=[95*mm, 95*mm])
    t_foot.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BACKGROUND', (0,1), (0,1), colors.whitesmoke),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t_foot)

    doc.build(story)

@app.route('/buono/finalize_and_get_pdf', methods=['POST'])
@login_required
def buono_finalize_and_get_pdf():
    db = SessionLocal()
    try:
        req_data = request.form
        ids = [int(i) for i in req_data.get('ids','').split(',') if i.isdigit()]
        rows = db.query(Articolo).filter(Articolo.id_articolo.in_(ids)).all()
        
        action = req_data.get('action')
        
        # AGGIORNAMENTO DATI (Sia per anteprima che per salvataggio)
        # È importante salvare le note temporaneamente o definitivamente
        # Qui le salviamo nel DB se l'azione è 'save'
        
        bn = req_data.get('buono_n')
        
        for r in rows:
            # Se stiamo salvando, aggiorna il numero buono
            if action == 'save' and bn:
                r.buono_n = bn
            
            # SALVA LE NOTE! (Così il DDT le troverà dopo)
            note_inserite = req_data.get(f"note_{r.id_articolo}")
            if note_inserite is not None:
                r.note = note_inserite

        # Commit delle note (importante per il passaggio al DDT)
        if action == 'save':
            db.commit()
            flash(f"Buono salvato. Note aggiornate.", "info")
        
        # Genera PDF
        pdf_bio = _generate_buono_pdf(req_data, rows)
        
        return send_file(
            pdf_bio, 
            as_attachment=(action == 'save'), 
            download_name=f'Buono_{bn}.pdf', 
            mimetype='application/pdf'
        )

    except Exception as e:
        db.rollback()
        print(f"ERRORE BUONO: {e}") 
        return f"Errore server: {e}", 500
    finally:
        db.close()
        

@app.route('/ddt/finalize', methods=['POST'])
@login_required
def ddt_finalize():
    import io
    db = SessionLocal()
    try:
        # 1. Recupera ID e Azione
        ids_str = request.form.get('ids', '')
        ids = [int(i) for i in ids_str.split(',') if i.strip().isdigit()]
        action = request.form.get('action', 'preview')

        # ✅ MEZZO IN USCITA (colonna: Mezzo Usc / campo DB: mezzi_in_uscita)
        mezzo_uscita = (request.form.get('mezzi_in_uscita') or '').strip()

        # ✅ obbligatorio SOLO quando finalizzi
        if action == 'finalize' and not mezzo_uscita:
            flash("Seleziona il Mezzo in uscita (Motrice / Bilico / Furgone) prima di finalizzare.", "danger")
            return redirect(url_for('giacenze'))

        # 2. Dati Testata
        n_ddt = request.form.get('n_ddt', '').strip()
        data_ddt_str = request.form.get('data_ddt')

        # ✅ Progressivo DDT: viene salvato SOLO quando si preme "Finalizza"
        # In anteprima mostriamo il prossimo numero senza consumarlo.
        if action == 'finalize':
            try:
                if (not n_ddt) or (n_ddt == peek_next_ddt_number()):
                    n_ddt = next_ddt_number()
            except Exception:
                # fallback: se qualcosa va storto, non blocchiamo la finalizzazione
                if not n_ddt:
                    n_ddt = next_ddt_number()

        # ✅ Se l'utente ha scelto un numero diverso (con le frecce),
        # aggiorniamo comunque il progressivo per evitare riutilizzi futuri.
        if action == 'finalize':
            consume_specific_ddt_number(n_ddt)

        try:
            data_ddt_obj = datetime.strptime(data_ddt_str, "%Y-%m-%d").date()
            data_formatted = data_ddt_obj.strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            data_ddt_obj = date.today()
            data_formatted = date.today().strftime("%d/%m/%Y")
            data_ddt_str = date.today().strftime("%Y-%m-%d")

        # 3. Recupera Destinatario
        dest_ragione = request.form.get('dest_ragione', '')
        dest_indirizzo = request.form.get('dest_indirizzo', '')
        dest_citta = request.form.get('dest_citta', '')

        # Sovrascrittura da eventuale rubrica
        dest_key = request.form.get('dest_key')
        if dest_key:
            try:
                dest_info = load_destinatari().get(dest_key, {})
                if dest_info:
                    dest_ragione = dest_info.get('ragione_sociale', '')
                    dest_indirizzo = dest_info.get('indirizzo', '')
                    dest_citta = dest_info.get('citta', '')
            except:
                pass

        # 4. Recupera Articoli
        articoli = db.query(Articolo).filter(Articolo.id_articolo.in_(ids)).all()
        righe_per_pdf = []

        # 5. Loop Articoli
        for art in articoli:
            raw_pezzi = request.form.get(f"pezzi_{art.id_articolo}")
            raw_colli = request.form.get(f"colli_{art.id_articolo}")
            raw_peso = request.form.get(f"peso_{art.id_articolo}")
            nuove_note = request.form.get(f"note_{art.id_articolo}", art.note)

            nuovi_pezzi = to_int_eu(raw_pezzi) if raw_pezzi is not None else art.pezzo
            nuovi_colli = to_int_eu(raw_colli) if raw_colli is not None else art.n_colli
            nuovo_peso = to_float_eu(raw_peso) if raw_peso is not None else art.peso

            # ✅ Se Finalizza -> Salva su DB
            if action == 'finalize':
                art.data_uscita = data_ddt_obj
                art.n_ddt_uscita = n_ddt
                art.mezzi_in_uscita = mezzo_uscita  # ✅ QUI COMPILIAMO "MEZZO USC"
                if nuove_note is not None:
                    art.note = nuove_note

            # Prepara righe PDF (PDF NON CAMBIA)
            righe_per_pdf.append({
                'codice_articolo': art.codice_articolo or '',
                'descrizione': art.descrizione or '',
                'pezzo': nuovi_pezzi,
                'n_colli': nuovi_colli,
                'peso': nuovo_peso,
                'n_arrivo': art.n_arrivo or '',
                'note': nuove_note,
                'commessa': art.commessa,
                'ordine': art.ordine,
                'buono': art.buono_n,
                'protocollo': art.protocollo
            })

        # 6. Salvataggio DB
        if action == 'finalize':
            db.commit()
            flash(f"DDT N.{n_ddt} del {data_formatted} salvato con successo. Mezzo uscita: {mezzo_uscita}", "success")

        # 7. Dati Generali PDF
        ddt_data = {
            'n_ddt': n_ddt,
            'data_uscita': data_formatted,
            'destinatario': dest_ragione,
            'dest_indirizzo': dest_indirizzo,
            'dest_citta': dest_citta,
            'causale': request.form.get('causale', ''),
            'vettore': request.form.get('targa', ''),
            'porto': request.form.get('porto', 'FRANCO'),
            'aspetto': request.form.get('aspetto', 'A VISTA')
        }

        # 8. Genera PDF
        pdf_bio = io.BytesIO()
        _genera_pdf_ddt_file(ddt_data, righe_per_pdf, pdf_bio)
        pdf_bio.seek(0)

        safe_n = n_ddt.replace('/', '-').replace('\\', '-')
        filename = f"DDT_{safe_n}_{data_ddt_str}.pdf"

        return send_file(
            pdf_bio,
            as_attachment=(action == 'finalize'),
            download_name=filename,
            mimetype='application/pdf'
        )

    except Exception as e:
        db.rollback()
        print(f"Errore DDT Finalize: {e}")
        return f"Errore durante la creazione del DDT: {e}", 500
    finally:
        db.close()



@app.route('/ddt/mezzo_uscita', methods=['GET', 'POST'])
@login_required
@require_admin
def ddt_mezzo_uscita():
    """
    Popup dopo finalizzazione DDT:
    salva la colonna mezzi_in_uscita (Motrice/Bilico/Furgone) sugli articoli selezionati.
    """
    if request.method == 'GET':
        ids_str = (request.args.get('ids') or '').strip()
        n_ddt = (request.args.get('n_ddt') or '').strip()
        return render_template('ddt_mezzo_uscita.html', ids=ids_str, n_ddt=n_ddt)

    # POST
    ids_str = (request.form.get('ids') or '').strip()
    n_ddt = (request.form.get('n_ddt') or '').strip()
    mezzo = (request.form.get('mezzo') or '').strip()

    ids = [int(i) for i in ids_str.split(',') if i.strip().isdigit()]
    if not ids:
        return "ERRORE: nessun articolo selezionato.", 400

    # ✅ obbligatorio e solo valori ammessi
    allowed = {"Motrice", "Bilico", "Furgone"}
    if mezzo not in allowed:
        flash("Seleziona un Mezzo valido (Motrice / Bilico / Furgone).", "danger")
        return redirect(url_for('ddt_mezzo_uscita', ids=ids_str, n_ddt=n_ddt))

    db = SessionLocal()
    try:
        q = db.query(Articolo).filter(Articolo.id_articolo.in_(ids))

        # (consigliato) aggiorna solo righe che hanno quel DDT di uscita
        if n_ddt:
            q = q.filter(Articolo.n_ddt_uscita == n_ddt)

        rows = q.all()

        for art in rows:
            if hasattr(art, "mezzi_in_uscita"):
                art.mezzi_in_uscita = mezzo
            else:
                raise Exception("Nel modello Articolo manca la colonna 'mezzi_in_uscita'.")

        db.commit()

        return render_template('ddt_mezzo_uscita_ok.html', mezzo=mezzo, count=len(rows), n_ddt=n_ddt)

    except Exception as e:
        db.rollback()
        return f"Errore salvataggio mezzo in uscita: {e}", 500
    finally:
        db.close()


@app.get('/labels')
@login_required
def labels_form():
    # --- PROTEZIONE ADMIN ---
    if session.get('role') != 'admin':
        flash("Accesso negato.", "danger")
        return redirect(url_for('giacenze'))
    # ------------------------

    db = SessionLocal()
    try:
        clienti_query = (
            db.query(Articolo.cliente)
              .distinct()
              .filter(Articolo.cliente != None, Articolo.cliente != '')
              .order_by(Articolo.cliente)
              .all()
        )
        clienti = [c[0] for c in clienti_query]
        return render_template('labels_form.html', clienti=clienti)
    finally:
        db.close()


# ==============================================================================
#  GESTIONE ETICHETTE (PDF) - ROUTE E GENERAZIONE
# ==============================================================================

@app.route('/labels_pdf', methods=['POST'])
@login_required
@require_admin
def labels_pdf():
    # Se ci sono ID selezionati, prendi dal DB
    ids = request.form.getlist('ids')
    articoli = []

    if ids:
        db = SessionLocal()
        try:
            articoli = db.query(Articolo).filter(Articolo.id_articolo.in_(ids)).all()
        finally:
            db.close()
    else:
        # Etichetta Manuale: Crea oggetto al volo con i dati del form
        a = Articolo()
        a.cliente = request.form.get('cliente')
        a.fornitore = request.form.get('fornitore')
        a.ordine = request.form.get('ordine')
        a.commessa = request.form.get('commessa')
        a.n_ddt_ingresso = request.form.get('ddt_ingresso')  # nome campo HTML
        a.data_ingresso = request.form.get('data_ingresso')
        a.n_arrivo = request.form.get('arrivo')  # arrivo (manuale) -> n_arrivo
        a.n_colli = to_int_eu(request.form.get('n_colli'))
        a.posizione = request.form.get('posizione')
        articoli = [a]

    # Genera PDF
    formato = request.form.get('formato', '62x100')
    pdf_bio = _genera_pdf_etichetta(articoli, formato)

    # ✅ FORZA DOWNLOAD (così poi stampi dal file scaricato con formato corretto)
    filename = f"Etichetta_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf_bio.seek(0)
    return send_file(
        pdf_bio,
        as_attachment=True,              # <-- IMPORTANTISSIMO
        download_name=filename,
        mimetype='application/pdf'
    )


# --- FUNZIONE ETICHETTE COMPATTA (100x62) ---

def _genera_pdf_etichetta(articoli, formato, anteprima=False):
    import io
    from pathlib import Path
    from datetime import datetime, date

    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
        Image as RLImage, KeepTogether
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.pagesizes import A4

    bio = io.BytesIO()

    # Formato Brother QL-800: 100mm x 62mm ORIZZONTALE
    if formato == '62x100':
        pagesize = (100 * mm, 62 * mm)
        margin = 1.2 * mm   # ✅ margini più piccoli
    else:
        pagesize = A4
        margin = 10 * mm

    doc = SimpleDocTemplate(
        bio,
        pagesize=pagesize,
        leftMargin=margin, rightMargin=margin,
        topMargin=margin, bottomMargin=margin
    )

    styles = getSampleStyleSheet()

    # ✅ FONT PIÙ PICCOLI per stare in 62mm senza spezzare
    s_lbl = ParagraphStyle(
        'LBL', parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9, leading=10
    )
    s_val = ParagraphStyle(
        'VAL', parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9, leading=10
    )
    s_hi = ParagraphStyle(
        'HI', parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11, leading=12
    )

    # logo path robusto
    if 'LOGO_PATH' in globals() and LOGO_PATH:
        logo_path = Path(LOGO_PATH)
    else:
        logo_path = Path(app.root_path) / "static" / "logo camar.jpg"

    def fmt_date(v):
        if not v:
            return ""
        try:
            if isinstance(v, (datetime, date)):
                return v.strftime("%d/%m/%Y")
            s = str(v).strip()
            if not s:
                return ""
            try:
                return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                pass
            try:
                return datetime.strptime(s[:10], "%d/%m/%Y").strftime("%d/%m/%Y")
            except Exception:
                return s[:10]
        except Exception:
            return str(v)

    story = []

    # calcola pagine totali (per evitare PageBreak finale)
    total_pages = 0
    colli_per_art = []
    for art in articoli:
        try:
            tot = int(getattr(art, "n_colli", None) or 1)
        except Exception:
            tot = 1
        tot = max(1, tot)
        colli_per_art.append(tot)
        total_pages += tot

    page_counter = 0

    for art, tot in zip(articoli, colli_per_art):
        for i in range(1, tot + 1):
            page_counter += 1

            blocco = []

            # ✅ LOGO PIÙ PICCOLO
            if logo_path.exists():
                try:
                    img = RLImage(str(logo_path), width=35 * mm, height=9 * mm)
                    img.hAlign = "LEFT"
                    blocco.append(img)
                    blocco.append(Spacer(1, 1 * mm))  # ✅ meno spazio
                except Exception:
                    pass

            arr_base = (getattr(art, "n_arrivo", "") or "").strip()
            arr_str = f"{arr_base} N.{i}" if arr_base else f"N.{i}"
            collo_str = f"{i}/{tot}"

            dati = [
                [Paragraph("CLIENTE:", s_lbl),   Paragraph((getattr(art, "cliente", "") or ""), s_val)],
                [Paragraph("FORNITORE:", s_lbl), Paragraph((getattr(art, "fornitore", "") or ""), s_val)],
                [Paragraph("ORDINE:", s_lbl),    Paragraph((getattr(art, "ordine", "") or ""), s_val)],
                [Paragraph("COMMESSA:", s_lbl),  Paragraph((getattr(art, "commessa", "") or ""), s_val)],
                [Paragraph("DDT ING.:", s_lbl),  Paragraph((getattr(art, "n_ddt_ingresso", "") or ""), s_val)],
                [Paragraph("DATA ING.:", s_lbl), Paragraph(fmt_date(getattr(art, "data_ingresso", "")), s_val)],
                [Paragraph("ARRIVO:", s_lbl),    Paragraph(arr_str, s_hi)],
                [Paragraph("N. COLLO:", s_lbl),  Paragraph(collo_str, s_hi)],
                [Paragraph("COLLI:", s_lbl),     Paragraph(str(tot), s_hi)],
                [Paragraph("POSIZIONE:", s_lbl), Paragraph((getattr(art, "posizione", "") or ""), s_val)],
            ]

            # ✅ colonne ottimizzate (totale area utile ≈ 100mm - margini)
            t = Table(dati, colWidths=[26 * mm, 70 * mm])
            t.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            blocco.append(t)

            # ✅ evita spezzature interne (se ci sta, resta 1 pagina)
            story.append(KeepTogether(blocco))

            if page_counter < total_pages:
                story.append(PageBreak())

    doc.build(story)
    bio.seek(0)
    return bio


# --- CONFIGURAZIONE FINALE E AVVIO ---
app.jinja_loader = DictLoader(templates)
app.jinja_env.globals['getattr'] = getattr
app.jinja_env.filters['fmt_date'] = fmt_date
    

    
# --- FIX DATABASE SCHEMA (Esegui all'avvio per correggere tipi colonne) ---
# ========================================================
# 🚑 PULSANTE DI EMERGENZA PER FIX DATABASE
# ========================================================
@app.route('/fix_db_schema')
@login_required
def fix_db_schema():
    if session.get('role') != 'admin': return "Accesso Negato", 403
    db = SessionLocal()
    log = []
    try:
        from sqlalchemy import text
        # 1. Crea tabelle fisiche se mancano
        Base.metadata.create_all(bind=db.get_bind())
        log.append("Struttura base verificata.")

        # 2. COLONNE GIACENZE (Articoli)
        # Aggiungiamo tutte le colonne che potrebbero mancare
        cols_art = [
            ("lotto", "TEXT"), ("peso", "FLOAT"), ("m2", "FLOAT"), ("m3", "FLOAT"), 
            ("n_arrivo", "TEXT"), ("data_uscita", "DATE"), ("serial_number", "TEXT"),
            ("lunghezza", "FLOAT"), ("larghezza", "FLOAT"), ("altezza", "FLOAT")
        ]
        for c, t in cols_art:
            try: 
                db.execute(text(f"ALTER TABLE articoli ADD COLUMN {c} {t};"))
                db.commit()
                log.append(f"Aggiunto {c} ad Articoli")
            except: db.rollback()

        # 3. COLONNE TRASPORTI
        cols_tra = [
            ("magazzino", "TEXT"), ("consolidato", "TEXT"), 
            ("tipo_mezzo", "TEXT"), ("ddt_uscita", "TEXT"), ("costo", "FLOAT")
        ]
        for c, t in cols_tra:
            try: 
                db.execute(text(f"ALTER TABLE trasporti ADD COLUMN {c} {t};"))
                db.commit()
                log.append(f"Aggiunto {c} a Trasporti")
            except: db.rollback()

        # 4. COLONNE PICKING (Lavorazioni)
        cols_lav = [
            ("seriali", "TEXT"), ("colli", "INTEGER"), 
            ("pallet_forniti", "INTEGER"), ("pallet_uscita", "INTEGER"), 
            ("ore_blue_collar", "FLOAT"), ("ore_white_collar", "FLOAT"),
            ("richiesta_di", "TEXT")
        ]
        for c, t in cols_lav:
            try: 
                db.execute(text(f"ALTER TABLE lavorazioni ADD COLUMN {c} {t};"))
                db.commit()
                log.append(f"Aggiunto {c} a Picking")
            except: db.rollback()

        return f"<h1>Database Aggiornato!</h1><ul>{''.join(['<li>'+l+'</li>' for l in log])}</ul><a href='/home'>Torna alla Home</a>"
    except Exception as e:
        return f"Errore Fix: {e}"
    finally:
        db.close()

def _parse_data_db_helper(val):
    """
    Accetta:
    - date / datetime
    - stringa 'YYYY-MM-DD'
    - stringa 'DD/MM/YYYY'
    - stringa con orario 'YYYY-MM-DD HH:MM:SS'
    Ritorna date oppure None.
    """
    if val is None:
        return None

    if isinstance(val, datetime):
        return val.date()

    if isinstance(val, date):
        return val

    s = str(val).strip()
    if not s:
        return None

    # prova YYYY-MM-DD
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except Exception:
        pass

    # prova DD/MM/YYYY
    try:
        return datetime.strptime(s[:10], "%d/%m/%Y").date()
    except Exception:
        pass

    return None

# --- LOGICA CALCOLO COSTI (ROBUSTA) ---
def _calcola_logica_costi(articoli, data_da, data_a, raggruppamento, m2_multiplier: float = 1.0, metric: str = "m2"):
    """
    metric:
      - "m2"    => usa art.m2
      - "colli" => usa art.n_colli
      - "pezzi" => usa art.pezzi / art.pezzo
    Ritorna SEMPRE anche m2_tot/m2_medio per compatibilità template.
    """
    from collections import defaultdict
    from datetime import timedelta, date, datetime

    val_per_giorno = defaultdict(float)

    def to_date_obj(d):
        if not d:
            return None
        if isinstance(d, datetime):
            return d.date()
        if isinstance(d, date):
            return d
        s = str(d).strip().split(" ")[0]
        if len(s) < 8 or not s[0].isdigit():
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except:
                pass
        return None

    d_start = to_date_obj(data_da)
    d_end = to_date_obj(data_a)
    if not d_start or not d_end:
        return []

    metric = (metric or "m2").strip().lower()

    def get_qty(art):
        # ✅ M2
        if metric == "m2":
            try:
                val_m2 = str(getattr(art, "m2", "") or "").replace(",", ".")
                m2 = float(val_m2) if val_m2 else 0.0
            except:
                m2 = 0.0
            if m2 <= 0:
                return 0.0

            # Area manovra (solo se metric == m2)
            try:
                m2 = m2 * float(m2_multiplier or 1.0)
            except:
                pass

            return float(m2)

        # ✅ COLLI
        if metric == "colli":
            try:
                return float(int(getattr(art, "n_colli", 0) or 0))
            except:
                return 0.0

        # ✅ PEZZI
        if metric == "pezzi":
            raw = getattr(art, "pezzi", None)
            if raw is None:
                raw = getattr(art, "pezzo", None)
            try:
                return float(int(raw or 0))
            except:
                return 0.0

        # fallback
        return 0.0

    for art in articoli:
        qty = get_qty(art)
        if qty <= 0:
            continue

        d_ingr = to_date_obj(getattr(art, "data_ingresso", None))
        if not d_ingr:
            continue

        d_usc = to_date_obj(getattr(art, "data_uscita", None))

        inizio = max(d_ingr, d_start)
        if d_usc:
            fine = min(d_usc - timedelta(days=1), d_end)
        else:
            fine = d_end

        if fine < inizio:
            continue

        cliente_key = (getattr(art, "cliente", None) or "SCONOSCIUTO").strip().upper()

        curr = inizio
        while curr <= fine:
            val_per_giorno[(cliente_key, curr)] += qty
            curr += timedelta(days=1)

    risultati_finali = []

    def pack_row(periodo, cliente, tot, medio, giorni):
        # ✅ compatibilità: restituisco SEMPRE anche m2_tot/m2_medio
        # così il template admin che stampa r.m2_tot / r.m2_medio funziona sempre.
        tot_s = f"{tot:.3f}" if isinstance(tot, (int, float)) else str(tot)
        med_s = f"{medio:.3f}" if isinstance(medio, (int, float)) else str(medio)

        return {
            "periodo": periodo,
            "cliente": cliente,
            # chiavi nuove "neutre"
            "tot": tot_s,
            "medio": med_s,
            "giorni": giorni,
            # chiavi legacy del template
            "m2_tot": tot_s,
            "m2_medio": med_s,
        }

    if raggruppamento == "giorno":
        sorted_keys = sorted(val_per_giorno.keys(), key=lambda k: (k[0], k[1]))
        for cliente, giorno in sorted_keys:
            val = val_per_giorno[(cliente, giorno)]
            risultati_finali.append(
                pack_row(giorno.strftime("%d/%m/%Y"), cliente, val, val, 1)
            )
    else:
        agg = defaultdict(lambda: {"sum": 0.0, "days": set()})
        for (cli, day), val in val_per_giorno.items():
            k = (cli, day.year, day.month)
            agg[k]["sum"] += val
            agg[k]["days"].add(day)

        sorted_keys = sorted(agg.keys(), key=lambda k: (k[1], k[2], k[0]))
        for (cli, y, m) in sorted_keys:
            dati = agg[(cli, y, m)]
            n_days = len(dati["days"])
            tot = dati["sum"]

            # ✅ M² EFFETTIVI (non medi): valore reale sull'ULTIMO giorno del periodo considerato per quel mese
            if n_days > 0:
                last_day = max(dati["days"])
                eff = float(val_per_giorno.get((cli, last_day), 0.0))
            else:
                eff = 0.0

            risultati_finali.append(
                pack_row(f"{m:02d}/{y}", cli, tot, eff, n_days)
            )

    return risultati_finali


def _calcola_logica_colli_giacenza(articoli, data_da, data_a, raggruppamento):
    """
    Calcola i COLLI in GIACENZA nel periodo (fotografia giornaliera o mensile),
    togliendo quelli già usciti.

    - Per ogni giorno: somma n_colli degli articoli che risultano "presenti" quel giorno.
    - Presente = data_ingresso <= giorno AND (data_uscita è vuota oppure data_uscita > giorno)
    """
    from collections import defaultdict
    from datetime import timedelta, date, datetime

    colli_per_giorno = defaultdict(float)

    def to_date_obj(d):
        if not d:
            return None
        if isinstance(d, datetime):
            return d.date()
        if isinstance(d, date):
            return d
        s = str(d).strip().split(' ')[0]
        if len(s) < 8 or not s[0].isdigit():
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        return None

    d_start = to_date_obj(data_da)
    d_end = to_date_obj(data_a)
    if not d_start or not d_end:
        return []

    for art in articoli:
        try:
            colli = float(int(art.n_colli or 0))
        except Exception:
            colli = 0.0

        if colli <= 0:
            continue

        d_ingr = to_date_obj(art.data_ingresso)
        if not d_ingr:
            continue

        d_usc = to_date_obj(art.data_uscita)

        # Range di verifica nel periodo
        start = max(d_ingr, d_start)
        end = d_end

        if end < start:
            continue

        cliente_key = (art.cliente or "SCONOSCIUTO").strip().upper()

        curr = start
        while curr <= end:
            presente = (d_ingr <= curr) and ((d_usc is None) or (d_usc > curr))
            if presente:
                colli_per_giorno[(cliente_key, curr)] += colli
            curr += timedelta(days=1)

    risultati = []

    if raggruppamento == "giorno":
        keys = sorted(colli_per_giorno.keys(), key=lambda k: (k[0], k[1]))
        for cli, day in keys:
            v = colli_per_giorno[(cli, day)]
            risultati.append({
                "periodo": day.strftime("%d/%m/%Y"),
                "cliente": cli,
                "tot": f"{v:.0f}",
                "medio": f"{v:.0f}",
                "giorni": 1
            })
    else:
        agg = defaultdict(lambda: {"sum": 0.0, "days": set()})
        for (cli, day), v in colli_per_giorno.items():
            k = (cli, day.year, day.month)
            agg[k]["sum"] += v
            agg[k]["days"].add(day)

        keys = sorted(agg.keys(), key=lambda k: (k[1], k[2], k[0]))
        for cli, y, m in keys:
            dati = agg[(cli, y, m)]
            n_days = len(dati["days"])
            tot = dati["sum"]
            avg = tot / n_days if n_days else 0.0
            risultati.append({
                "periodo": f"{m:02d}/{y}",
                "cliente": cli,
                "tot": f"{tot:.0f}",
                "medio": f"{avg:.0f}",
                "giorni": n_days
            })

    return risultati


@app.route('/calcola_costi', methods=['GET', 'POST'])
@login_required
def calcola_costi():
    oggi = date.today()
    data_da_val = (oggi.replace(day=1)).strftime("%Y-%m-%d")
    data_a_val = oggi.strftime("%Y-%m-%d")

    # Admin: può filtrare + area manovra
    # Client: solo il proprio cliente e niente area manovra
    is_admin = (session.get('role') == 'admin')
    cliente_lock = current_cliente()  # stringa se role=client, altrimenti None

    cliente_val = (cliente_lock or "")
    raggruppamento = "mese"
    area_manovra_val = False
    risultati = []
    metric = "m2"

    def _metric_for_cliente(nome_cliente: str) -> str:
        s = (nome_cliente or "").strip().upper()
        # Per Galvano Tecnica calcoliamo COLLI in giacenza
        if "GALVANO" in s:
            return "colli"
        return "m2"

    if request.method == 'POST':
        data_da_str = request.form.get('data_da')
        data_a_str = request.form.get('data_a')
        raggruppamento = request.form.get('raggruppamento', 'mese')

        # Cliente + area manovra
        if is_admin:
            cliente_val = (request.form.get('cliente') or '').strip()
            area_manovra = (request.form.get('area_manovra') == '1')
        else:
            cliente_val = (cliente_lock or '').strip()
            area_manovra = False

        export_excel = ('export_excel' in request.form)

        # metrica (colli o m2)
        metric = _metric_for_cliente(cliente_val)

        try:
            db = SessionLocal()
            query = db.query(Articolo)

            # ✅ filtro sicuro
            if cliente_val:
                if not is_admin:
                    cli_up = cliente_val.strip().upper()
                    query = query.filter(func.upper(func.trim(Articolo.cliente)) == cli_up)
                else:
                    query = query.filter(Articolo.cliente.ilike(f"%{cliente_val}%"))

            articoli = query.all()
            db.close()

            # ✅ calcolo in base a metrica
            if metric == "colli":
                risultati = _calcola_logica_colli_giacenza(
                    articoli,
                    data_da_str,
                    data_a_str,
                    raggruppamento
                )
            else:
                risultati = _calcola_logica_costi(
                    articoli,
                    data_da_str,
                    data_a_str,
                    raggruppamento,
                    m2_multiplier=(1.25 if (is_admin and area_manovra) else 1.0)
                )

            data_da_val = data_da_str
            data_a_val = data_a_str
            area_manovra_val = bool(is_admin and area_manovra and metric == "m2")

            # ✅ Export Excel
            if export_excel:
                try:
                    df = pd.DataFrame(risultati)

                    if metric == "colli":
                        # qui i risultati hanno chiavi: periodo, cliente, tot, medio, giorni
                        df = df.rename(columns={
                            'periodo': 'Periodo',
                            'cliente': 'Cliente',
                            'tot': 'Colli Giacenza (somma)',
                            'medio': 'Colli Medi',
                            'giorni': 'Giorni'
                        })
                        filename = f"Report_Colli_Giacenza_{data_da_val}_to_{data_a_val}.xlsx"
                    else:
                        # qui i risultati hanno chiavi: periodo, cliente, m2_tot, m2_medio, giorni
                        df = df.rename(columns={
                            'periodo': 'Periodo',
                            'cliente': 'Cliente',
                            'm2_tot': 'M2 Tot',
                            'm2_medio': 'M2 Medio',
                            'giorni': 'Giorni'
                        })
                        extra = '_AREA_MANOVRA' if area_manovra_val else ''
                        filename = f"Report_Costi{extra}_{data_da_val}_to_{data_a_val}.xlsx"

                    bio = io.BytesIO()
                    df.to_excel(bio, index=False, engine='openpyxl')
                    bio.seek(0)

                    return send_file(
                        bio,
                        as_attachment=True,
                        download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                except Exception as e:
                    flash(f"Errore export Excel: {e}", "danger")

            if not risultati:
                flash("Nessun dato valido trovato per i criteri selezionati.", "warning")

        except Exception as e:
            flash(f"Errore: {e}", "danger")

    # ✅ anche su GET metrica in base al cliente lock
    metric = _metric_for_cliente(cliente_val)

    return render_template(
        'calcoli.html',
        risultati=risultati,
        data_da=data_da_val,
        data_a=data_a_val,
        cliente_filtro=cliente_val,
        raggruppamento=raggruppamento,
        area_manovra=area_manovra_val,
        is_admin=is_admin,
        metric=metric
    )



# --- AVVIO FLASK APP ---
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    print(f"✅ Avvio Gestionale Camar Web Edition su http://127.0.0.1:{port}")
    app.run(host='0.0.0.0', port=port, debug=True)

